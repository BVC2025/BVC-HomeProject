"""Plain-text extraction per file format for the RAG ingestion pipeline.

Each extractor takes a filesystem Path and returns raw text. Kept as small,
dependency-light functions — no framework, matching this codebase's
preference for plain functions over abstractions."""

from pathlib import Path
import csv


def _extract_pdf(file_path: Path) -> str:

    from pypdf import PdfReader

    reader = PdfReader(str(file_path))

    return "\n\n".join(
        (page.extract_text() or "") for page in reader.pages
    )


def _extract_docx(file_path: Path) -> str:

    from docx import Document

    doc = Document(str(file_path))

    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_txt_md(file_path: Path) -> str:

    return file_path.read_text(encoding="utf-8", errors="ignore")


def _extract_csv(file_path: Path) -> str:

    lines = []

    with file_path.open(newline="", encoding="utf-8", errors="ignore") as f:

        for row in csv.reader(f):

            lines.append(", ".join(cell.strip() for cell in row if cell.strip()))

    return "\n".join(line for line in lines if line)


def _extract_xlsx(file_path: Path) -> str:

    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)

    lines = []

    for sheet in wb.worksheets:

        lines.append(f"# Sheet: {sheet.title}")

        for row in sheet.iter_rows(values_only=True):

            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]

            if cells:

                lines.append(", ".join(cells))

    return "\n".join(lines)


_EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".txt": _extract_txt_md,
    ".md": _extract_txt_md,
    ".csv": _extract_csv,
    ".xlsx": _extract_xlsx,
}

SUPPORTED_EXTENSIONS = set(_EXTRACTORS.keys())


def extract_text(file_path: Path, file_extension: str) -> str:
    """Dispatch to the right extractor by extension (lowercase, with dot,
    e.g. '.pdf'). Raises ValueError for anything unsupported — callers
    should validate against SUPPORTED_EXTENSIONS before this is reached."""

    ext = (file_extension or "").lower()

    fn = _EXTRACTORS.get(ext)

    if fn is None:

        raise ValueError(f"Unsupported file extension for extraction: {ext}")

    text = fn(Path(file_path))

    if not text or not text.strip():

        raise ValueError("No extractable text found in document")

    return text
