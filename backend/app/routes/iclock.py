"""
ADMS-Push endpoints for ZKTeco / ESSL biometric devices.

Real-hardware protocol notes
----------------------------
The device (e.g. ESSL X2008, firmware 8.0.4.7, Push Service 2.0.33S)
talks to the ERP over PLAIN HTTP with a tiny text protocol — NOT
JSON. Every request carries `?SN=<serial>` in the query string; the
serial identifies which physical box is talking. All responses are
`text/plain`.

Endpoints implemented
---------------------
  GET  /iclock/cdata        — handshake, returns server options
  POST /iclock/cdata        — receives ATTLOG (punches) + OPERLOG
                              (user enrolments / operator actions)
  GET  /iclock/getrequest   — device polls for commands. We reply OK
                              (no server-side commands yet)
  POST /iclock/devicecmd    — device reports command results → OK
  GET  /iclock/ping         — heartbeat some firmwares use → OK

Flow for a normal punch
-----------------------
1. Employee places finger on the device.
2. Device POSTs to /iclock/cdata?SN=JNP2255102739&table=ATTLOG&Stamp=...
   Body (tab-separated, one record per line):
       8   2026-07-30 09:12:03   0   1   0   0
       │   │                     │   │
       PIN timestamp             status verify_mode
3. We look up Employee where FINGERPRINT_ID = "8" → BVC008 (Puviyarasi).
4. Write one BiometricEvent row (raw log, keeps unmapped PINs too).
5. Update today's Attendance:
      - device status 0 (or first punch)  → CHECK_IN if empty
      - device status 1 (or later punch)  → CHECK_OUT (overwrites)
      - device status 4 / 5               → OT_CHECK_IN / OT_CHECK_OUT
6. Reply `OK: 1` — device marks the record as delivered and stops
   retrying.
"""

from __future__ import annotations

from datetime import datetime, date, time
from typing import Optional

from fastapi import APIRouter, Depends, Query, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import PlainTextResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from sqlalchemy.exc import IntegrityError

from app.database.database import get_db
from app.models.models import Employee, Attendance, BiometricEvent


router = APIRouter(prefix="/iclock", tags=["ADMS Biometric"])


# ---------------------------------------------------------------------
# Config — tune only if the device misbehaves.
# ---------------------------------------------------------------------

# Anyone scanning in AFTER this time is marked LATE. Matches
# attendance.py / biometric.py so all three code paths agree.
WORK_START = time(9, 15)

# Regular shift ends at 6:00 PM. Work past this time is overtime and
# gets moved to OVERTIME_HOURS / OT_CHECK_IN / OT_CHECK_OUT.
OFFICE_END = time(18, 0)

# Punches strictly before this time-of-day are treated as morning
# arrivals (CHECK_IN candidates). Punches at or after are treated as
# departures (CHECK_OUT). Prevents an accidental verify-scan at
# 09:41 from being recorded as CHECK_OUT the same day the employee
# arrived at 09:20 — the bug the user reported.
MORNING_CUTOFF = time(13, 0)      # 1:00 PM

# Ignore biometric scans that arrive within this many seconds of the
# most recent edge for the same employee. Protects against ESSL's
# common "verify + real punch" double-tap that arrives 5–20s apart.
DEDUP_WINDOW_SECONDS = 60


# ---------------------------------------------------------------------
# GET /iclock/cdata — device handshake on power-up / config-refresh.
#
# The device sends: ?SN=<serial>&options=all&pushver=<v>&language=<l>
# We reply with the standard ADMS option block. The exact keys the
# device cares about are Stamp, OpStamp, TransFlag, TransInterval,
# Realtime — everything else is informational. Keeping Realtime=1 tells
# the device to push each punch immediately (no batching).
# ---------------------------------------------------------------------
@router.get("/cdata", response_class=PlainTextResponse)
@router.get("/cdata.aspx", response_class=PlainTextResponse)
def iclock_handshake(
    SN: str = Query(""),
    options: str = Query("all"),
    pushver: str = Query(""),
    language: str = Query(""),
):
    lines = [
        f"GET OPTION FROM: {SN}",
        "Stamp=9999",
        "OpStamp=9999",
        "ErrorDelay=30",
        "Delay=10",
        "TransTimes=00:00;14:05",
        "TransInterval=1",
        "TransFlag=TransData AttLog OpLog EnrollUser ChgUser EnrollFP ChgFP UserPic",
        "TimeZone=8",
        "Realtime=1",
        "Encrypt=None",
        "ServerVer=2.4.1",
        "PushProtVer=2.4.1",
    ]
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------
# POST /iclock/cdata — punches (table=ATTLOG) + user changes (OPERLOG).
#
# The body is text/plain with lines separated by \n or \r\n, records
# separated inside a line by TABS. We accept latin-1 decoding to survive
# devices that send extended chars for user names.
# ---------------------------------------------------------------------
@router.post("/cdata", response_class=PlainTextResponse)
@router.post("/cdata.aspx", response_class=PlainTextResponse)
async def iclock_push(
    request: Request,
    SN: str = Query(""),
    table: str = Query(""),
    Stamp: str = Query(""),
    db: Session = Depends(get_db),
):
    body_bytes = await request.body()

    # Some firmwares send latin-1; UTF-8 in newer builds. Try UTF-8
    # first, fall back to latin-1 which never fails.
    try:
        body = body_bytes.decode("utf-8")
    except UnicodeDecodeError:
        body = body_bytes.decode("latin-1", errors="replace")

    if table.upper() == "ATTLOG":
        count = _handle_attlog(db, SN, body)
        return f"OK: {count}"

    if table.upper() == "OPERLOG":
        count = _handle_operlog(db, SN, body)
        return f"OK: {count}"

    # Unknown table → don't error, just acknowledge so the device
    # doesn't sit in a retry loop.
    return "OK"


# ---------------------------------------------------------------------
# GET /iclock/getrequest — device polls for server-side commands.
#
# Return "OK" when there's nothing to do. Later we can push a command
# by returning e.g. `C:1:REBOOT` or `C:2:DATA UPDATE USERINFO PIN=1...`
# ---------------------------------------------------------------------
@router.get("/getrequest", response_class=PlainTextResponse)
@router.get("/getrequest.aspx", response_class=PlainTextResponse)
def iclock_get_command(
    SN: str = Query(""),
    INFO: Optional[str] = Query(None),
):
    return "OK"


# ---------------------------------------------------------------------
# POST /iclock/devicecmd — device reports how a previously-sent command
# executed. We don't queue commands yet, but acknowledge so the device
# is happy.
# ---------------------------------------------------------------------
@router.post("/devicecmd", response_class=PlainTextResponse)
@router.post("/devicecmd.aspx", response_class=PlainTextResponse)
async def iclock_command_result(request: Request, SN: str = Query("")):
    return "OK"


# ---------------------------------------------------------------------
# GET /iclock/ping — heartbeat.
# ---------------------------------------------------------------------
@router.get("/ping", response_class=PlainTextResponse)
@router.get("/ping.aspx", response_class=PlainTextResponse)
def iclock_ping(SN: str = Query("")):
    return "OK"


# =====================================================================
# ATTLOG handler
# =====================================================================
def _handle_attlog(db: Session, device_sn: str, body: str) -> int:
    """
    ATTLOG body format (tab-separated, one record per line):

        PIN <TAB> YYYY-MM-DD HH:MM:SS <TAB> Status <TAB> VerifyMode
            <TAB> WorkCode <TAB> Reserved

    Status codes (per ADMS spec):
        0 = check-in       1 = check-out
        2 = break-out      3 = break-in
        4 = OT check-in    5 = OT check-out

    VerifyMode codes:
        0 = password       1 = fingerprint     2 = card
        15 = face

    Returns the number of records the ERP accepted (some rows may be
    duplicates or reference unknown PINs — we still count them so the
    device doesn't resend forever).
    """
    lines = [ln.strip() for ln in body.replace("\r\n", "\n").split("\n") if ln.strip()]
    processed = 0

    for line in lines:

        try:
            _process_one_attlog_line(db, device_sn, line)
            # Commit after each record so one bad line can't roll back
            # a whole batch — critical when a device dumps months of
            # backlogged punches on first successful push.
            db.commit()
            processed += 1

        except IntegrityError:
            # Almost always the (EMPLOYEE_ID, DATE) unique constraint —
            # we're trying to insert a second Attendance row for the
            # same employee on the same day. Roll back the failed
            # add, then update the existing row instead.
            db.rollback()
            try:
                _retry_attlog_line_as_update(db, device_sn, line)
                db.commit()
                processed += 1
            except Exception:
                db.rollback()
                # Give up on this specific line but keep going — return
                # "OK" to the device so it doesn't resend forever.
                processed += 1

        except Exception:
            db.rollback()
            # Any other failure — skip the line so one weird record
            # doesn't kill the whole batch. Device retries handle it.
            processed += 1

    return processed


def _process_one_attlog_line(db: Session, device_sn: str, line: str) -> None:
    """Parse a single ATTLOG record and apply it to BiometricEvent +
    Attendance. Caller handles commit/rollback."""

    parts = line.split("\t")

    if len(parts) < 2:
        return

    pin = parts[0].strip()
    raw_ts = parts[1].strip()

    try:
        event_time = datetime.strptime(raw_ts, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        _write_biometric_event(
            db, device_sn, pin, None,
            verify_mode="ERR", result="BAD_TIMESTAMP",
            raw=line, event_time=datetime.utcnow(),
        )
        return

    status_raw = parts[2].strip() if len(parts) >= 3 else "0"
    verify_raw = parts[3].strip() if len(parts) >= 4 else "1"
    verify_mode = _verify_label(verify_raw)

    emp = (
        db.query(Employee)
          .filter(Employee.FINGERPRINT_ID == pin)
          .first()
    )
    if emp is None:
        # HR-summary Excel imports use EMPLOYEE_CODE ("BVC002") as the
        # PIN column instead of the device enrollment number.
        emp = (
            db.query(Employee)
              .filter(Employee.EMPLOYEE_CODE == pin)
              .first()
        )
    emp_id = emp.ID if emp else None

    already = (
        db.query(BiometricEvent)
          .filter(
              BiometricEvent.DEVICE_ID == device_sn,
              BiometricEvent.FINGERPRINT_ID == pin,
              BiometricEvent.EVENT_TIME == event_time,
          )
          .first()
    )

    if already:
        return

    _write_biometric_event(
        db, device_sn, pin, emp_id,
        verify_mode=verify_mode,
        result="SUCCESS" if emp else "UNKNOWN_USER",
        raw=line,
        event_time=event_time,
    )

    if emp:
        _apply_to_attendance(db, emp, event_time, status_raw)


def _retry_attlog_line_as_update(db: Session, device_sn: str, line: str) -> None:
    """Re-run a single ATTLOG line after an IntegrityError, assuming the
    Attendance row for (employee, day) already exists — so we UPDATE it
    instead of INSERTing. BiometricEvent write is retried too, since the
    session was rolled back."""

    parts = line.split("\t")
    if len(parts) < 2:
        return

    pin = parts[0].strip()
    try:
        event_time = datetime.strptime(parts[1].strip(), "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return

    status_raw = parts[2].strip() if len(parts) >= 3 else "0"
    verify_raw = parts[3].strip() if len(parts) >= 4 else "1"

    emp = (
        db.query(Employee)
          .filter(Employee.FINGERPRINT_ID == pin)
          .first()
    )
    if emp is None:
        emp = (
            db.query(Employee)
              .filter(Employee.EMPLOYEE_CODE == pin)
              .first()
        )
    if not emp:
        return

    # Write the BiometricEvent (audit) — safe on retry, has no unique
    # constraint other than the primary key.
    _write_biometric_event(
        db, device_sn, pin, emp.ID,
        verify_mode=_verify_label(verify_raw),
        result="SUCCESS",
        raw=line,
        event_time=event_time,
    )

    # Now apply to the *existing* Attendance row. Because we've flushed
    # the previous batch and rolled back the failed insert, a fresh
    # query will find the row that was already there before this batch.
    _apply_to_attendance(db, emp, event_time, status_raw)


def _apply_to_attendance(
    db: Session,
    emp: Employee,
    event_time: datetime,
    status_code: str,
) -> None:
    """
    Map an ADMS punch onto today's Attendance row.

    Every employee gets ONE row per day; this function either creates
    it (on first punch) or updates it. The device's status code drives
    the decision when it's a known value; otherwise we auto-decide
    based on what's already filled.
    """
    day = event_time.date()

    row = (
        db.query(Attendance)
          .filter(
              and_(
                  Attendance.EMPLOYEE_ID == emp.ID,
                  Attendance.DATE == day,
              )
          )
          .first()
      )

    if not row:

        row = Attendance(
            EMPLOYEE_ID=emp.ID,
            DATE=day,
            VENDOR_ID=getattr(emp, "VENDOR_ID", 1) or 1,
            STATUS="PRESENT",
        )
        db.add(row)
        # Flush so a subsequent query in the same session sees this
        # row. Without this, two ATTLOG lines for the same
        # (employee, day) both try to INSERT and the second violates
        # uq_attendance_employee_date.
        db.flush()

    # Decide what this punch means, based on time-of-day (not device
    # status). ESSL devices commonly send status=0 for every punch,
    # so trusting the code caused accidental double-scans in the
    # morning to be recorded as CHECK_OUT. See MORNING_CUTOFF for
    # the rule.
    action = _decide_action(row, event_time)

    if action is None:
        # Duplicate / debounced — this punch is ignored, keep the
        # existing edges intact.
        row.DEVICE_INFO = "ESSL_ADMS"
        return

    if action == "CHECK_IN":

        row.CHECK_IN = event_time
        row.STATUS = "LATE" if event_time.time() > WORK_START else "PRESENT"

    elif action == "CHECK_OUT":

        row.CHECK_OUT = event_time

        # Split into regular hours + OT if the employee stayed past
        # 6:00 PM. This is the whole point of the OT feature the user
        # requested — WORKED_HOURS stores the "up to 6 PM" portion,
        # OVERTIME_HOURS the "past 6 PM" portion, and OT_CHECK_IN /
        # OT_CHECK_OUT capture the OT session edges (6 PM → actual
        # punch-out) for payroll audit.
        _recompute_hours_with_ot(row, event_time)

    row.DEVICE_INFO = "ESSL_ADMS"


def _decide_action(row: Attendance, event_time: datetime) -> Optional[str]:
    """
    Return 'CHECK_IN' / 'CHECK_OUT' / None (ignore).

    Time-of-day rules (independent of device status code):
      • Punch before MORNING_CUTOFF (13:00)
          → CHECK_IN if none recorded yet.
          → Ignore if CHECK_IN already set within DEDUP_WINDOW_SECONDS
            (accidental verify + real punch).
          → Otherwise still ignore — a second morning punch after the
            debounce window is treated as duplicate, NOT as check-out.
            (This is what fixes the "09:25 → 09:41 EARLY_EXIT" bug.)
      • Punch at or after MORNING_CUTOFF
          → CHECK_OUT (latest wins; a later punch overrides an earlier
            check-out on the same day, so a 6:15 PM punch replaces a
            2:00 PM lunch-time punch).
          → Ignored if it arrives within DEDUP_WINDOW_SECONDS of the
            currently-recorded CHECK_OUT.
    """

    is_morning = event_time.time() < MORNING_CUTOFF

    if is_morning:

        if row.CHECK_IN is None:
            return "CHECK_IN"

        # Second morning punch — treat as duplicate scan, not exit.
        # (The user's bug: double-punch in the morning was recorded
        # as CHECK_OUT and marked EARLY_EXIT.)
        return None

    # Afternoon / evening punch — this is the departure.
    if row.CHECK_OUT is not None:
        delta = abs((event_time - row.CHECK_OUT).total_seconds())
        if delta < DEDUP_WINDOW_SECONDS:
            return None
        # Otherwise the LATEST evening punch wins — someone who
        # punched at 2 PM (lunch outing) and again at 6:30 PM should
        # have CHECK_OUT = 6:30 PM.

    return "CHECK_OUT"


def _recompute_hours_with_ot(row: Attendance, punch_out: datetime) -> None:
    """
    Fill WORKED_HOURS (regular) and OVERTIME_HOURS (past 6 PM) on the
    Attendance row from the current CHECK_IN → punch_out range.

    Payroll consumes both fields:
      WORKED_HOURS   → base salary calculation
      OVERTIME_HOURS → OT payout at the OT rate
      OT_CHECK_IN    → 6:00 PM anchor of the OT session
      OT_CHECK_OUT   → actual punch-out
    """
    if not row.CHECK_IN or not punch_out or punch_out <= row.CHECK_IN:
        return

    office_end_dt = datetime.combine(punch_out.date(), OFFICE_END)

    if punch_out <= office_end_dt:
        # Left at or before 6 PM — no OT this day.
        worked = (punch_out - row.CHECK_IN).total_seconds() / 3600.0
        row.WORKED_HOURS = round(max(0.0, worked), 2)
        row.OVERTIME_HOURS = 0.0
        row.OT_CHECK_IN = None
        row.OT_CHECK_OUT = None
        return

    # CHECK_OUT is past 6 PM: split into regular + OT.
    # Regular = CHECK_IN → 6 PM (or CHECK_IN → punch_out if arrived
    # after 6 PM, which we clamp to zero regular hours).
    if row.CHECK_IN < office_end_dt:
        regular = (office_end_dt - row.CHECK_IN).total_seconds() / 3600.0
        ot_start = office_end_dt
    else:
        # Arrived after 6 PM (unusual — maybe a night shift). Treat
        # the whole thing as OT so the regular column doesn't carry
        # a hostile negative.
        regular = 0.0
        ot_start = row.CHECK_IN

    ot_hours = (punch_out - ot_start).total_seconds() / 3600.0

    row.WORKED_HOURS = round(max(0.0, regular), 2)
    row.OVERTIME_HOURS = round(max(0.0, ot_hours), 2)
    row.OT_CHECK_IN = ot_start
    row.OT_CHECK_OUT = punch_out


# =====================================================================
# OPERLOG handler — user enrolments / operator actions.
#
# We store a BiometricEvent for audit but don't try to auto-create
# Employees. Enrolment stays a manual admin step in the ERP so we
# always have an approving user.
# =====================================================================
def _handle_operlog(db: Session, device_sn: str, body: str) -> int:

    lines = [ln.strip() for ln in body.replace("\r\n", "\n").split("\n") if ln.strip()]

    for line in lines:

        _write_biometric_event(
            db, device_sn, None, None,
            verify_mode="OPERLOG",
            result="INFO",
            raw=line[:1000],
            event_time=datetime.utcnow(),
        )

    if lines:
        db.commit()

    return len(lines)


# =====================================================================
# Helpers
# =====================================================================
def _verify_label(code: str) -> str:

    return {
        "0":  "PWD",
        "1":  "FP",
        "2":  "CARD",
        "3":  "FP_PWD",
        "4":  "FP_CARD",
        "15": "FACE",
    }.get(code.strip(), f"MODE_{code}")


def _write_biometric_event(
    db: Session,
    device_sn: str,
    pin: Optional[str],
    employee_id: Optional[str],
    *,
    verify_mode: str,
    result: str,
    raw: str,
    event_time: datetime,
) -> None:

    ev = BiometricEvent(
        DEVICE_ID=device_sn or "UNKNOWN",
        FINGERPRINT_ID=(pin or ""),
        EMPLOYEE_ID=employee_id,
        EVENT_TIME=event_time,
        VERIFY_MODE=verify_mode,
        RESULT=result,
        RAW_PAYLOAD=(raw or "")[:1000],
        VENDOR_ID=1,
    )
    db.add(ev)


# =====================================================================
# USB-import fallback — when the device can't push over network.
#
# Workflow the admin follows:
#   1. Insert a USB pen drive into the ESSL X2008.
#   2. Menu → USB Manager → Download → Attendance Data
#      (device writes something like `1_attlog.dat` or `attlog.txt`).
#   3. Remove USB, plug into PC, open ERP admin page.
#   4. Choose the file → Upload.
#
# We accept both the plain-text ATTLOG format the ADMS Push uses AND
# the alternative "|"-separated format some ESSL firmwares export.
# Same PIN → Employee lookup, same dedup, same Attendance rules.
# =====================================================================
@router.post("/import-attlog")
async def import_attlog_file(
    file: UploadFile = File(...),
    device_sn: str = Form("MANUAL_USB"),
    db: Session = Depends(get_db),
):
    """
    Upload an attendance log exported from the biometric device via
    USB. Two file formats accepted:

      1. Native ESSL/ZKTeco text export (.dat / .txt / .log / .csv).
         Passed straight through to the existing _handle_attlog parser
         which is what the ADMS-Push endpoint uses too.

      2. Excel workbook (.xlsx) — for cases where HR downloaded the
         attendance sheet from the device's web UI or converted it
         manually. We parse the sheet, extract PIN + timestamp per
         row, synthesise an ATTLOG text block in the exact tab-
         separated shape _handle_attlog expects, then hand it to the
         same parser. The biometric handler stays untouched — it
         still just sees canonical ATTLOG text.

    Returns the number of records processed vs. inserted vs. skipped
    as duplicates. To see the *calculated* per-employee summary
    (present/absent/OT/etc.) for the same month, call
    GET /iclock/import-summary?year=YYYY&month=M after the upload.
    """
    raw = await file.read()

    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    filename = (file.filename or "").lower()
    is_xlsx = (
        filename.endswith(".xlsx")
        or filename.endswith(".xlsm")
        or (file.content_type or "").endswith("sheetml.sheet")
    )

    if is_xlsx:
        try:
            normalised = _xlsx_to_attlog_lines(raw)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Could not parse Excel file: {exc}",
            )
        if not normalised.strip():
            raise HTTPException(
                status_code=400,
                detail=(
                    "Excel file had no recognisable rows. Expected columns "
                    "for PIN/EmployeeCode and a Date+Time (or single DateTime)."
                ),
            )
    else:
        # ESSL exports are usually latin-1; some newer firmwares are utf-8.
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1", errors="replace")

        # Normalise: some devices use "|" as separator, others use tab.
        # Convert any "|" runs to "\t" so _handle_attlog works uniformly.
        normalised = text.replace("|", "\t")

    # Count DB rows before we start so we can report a delta accurately.
    before = db.query(BiometricEvent).filter(
        BiometricEvent.DEVICE_ID == device_sn
    ).count()

    processed = _handle_attlog(db, device_sn, normalised)

    after = db.query(BiometricEvent).filter(
        BiometricEvent.DEVICE_ID == device_sn
    ).count()

    inserted = after - before

    return {
        "filename": file.filename,
        "device_sn": device_sn,
        "records_seen": processed,
        "rows_inserted": inserted,
        "rows_skipped_as_duplicate": max(0, processed - inserted),
        "file_kind": "xlsx" if is_xlsx else "text",
    }


# ---------------------------------------------------------------------
# Excel → ATTLOG text converter
# ---------------------------------------------------------------------
# The existing text parser _handle_attlog expects tab-separated lines
# like:
#     8\t2026-07-30 09:12:03\t0\t1\t0\t0
#     ^   ^                  ^   ^
#     PIN timestamp          status verify (unused)
#
# HR's Excel typically has one of these shapes:
#   A) One row per punch: EnrollNo/PIN | Name | Date | Time | InOut
#   B) One row per punch: PIN | DateTime            | InOut
#   C) Legacy: PIN | Name | Date & Time             | Status
# We detect PIN/timestamp columns tolerantly and skip anything else.
# =====================================================================

def _xlsx_to_attlog_lines(raw_bytes: bytes) -> str:

    from io import BytesIO
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl not installed. Run: pip install openpyxl"
        ) from e

    wb = load_workbook(BytesIO(raw_bytes), data_only=True, read_only=True)

    lines: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # First non-empty row = header; every subsequent row = a punch
        header: list[str] = []
        header_row_index = -1
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and any(c is not None and str(c).strip() for c in row):
                header = [
                    (str(c).strip().lower() if c is not None else "")
                    for c in row
                ]
                header_row_index = i
                break

        if not header:
            continue

        # Find columns tolerantly. First match wins.
        def _find_col(*candidates: str) -> int:
            for idx, h in enumerate(header):
                for cand in candidates:
                    if cand in h:
                        return idx
            return -1

        pin_col   = _find_col("pin", "enroll", "employee id", "empid", "emp_id", "employee code", "user id", "userid")
        dt_col    = _find_col("date time", "datetime", "date & time", "timestamp", "punch time", "log time")
        date_col  = _find_col("date")
        time_col  = _find_col("time")
        state_col = _find_col("in/out", "inout", "status", "verify state", "state")

        # HR's daily-summary shape: one row per (employee, day) with
        # multiple time columns for the check-in, check-out and OT edges.
        ci_col    = _find_col("check in", "check-in", "checkin", "in time", "punch in")
        co_col    = _find_col("check out", "check-out", "checkout", "out time", "punch out")
        ot_ci_col = _find_col("ot check in", "ot in", "overtime in", "ot-in")
        ot_co_col = _find_col("ot check out", "ot out", "overtime out", "ot-out")
        is_daily_summary = date_col >= 0 and ci_col >= 0

        if pin_col < 0:
            continue  # this sheet doesn't look like a punch log — skip

        # Single-punch-per-row shape needs a datetime source. Skip only
        # if we ALSO can't fall back to the daily-summary layout.
        if dt_col < 0 and (date_col < 0 or time_col < 0) and not is_daily_summary:
            continue

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row_index:
                continue
            if not row:
                continue

            def _cell(idx: int):
                return row[idx] if 0 <= idx < len(row) else None

            pin_raw = _cell(pin_col)
            if pin_raw is None or str(pin_raw).strip() == "":
                continue
            pin = str(pin_raw).strip()
            # Some Excel exports read the PIN as a float (e.g. 8.0) — normalize.
            if pin.endswith(".0"):
                pin = pin[:-2]

            if is_daily_summary:
                # Emit one ATTLOG line per non-empty time cell so the
                # existing punch handler can build the Attendance row.
                date_val = _cell(date_col)
                for tcol, state_flag in (
                    (ci_col,    "0"),
                    (co_col,    "1"),
                    (ot_ci_col, "0"),
                    (ot_co_col, "1"),
                ):
                    if tcol < 0:
                        continue
                    tval = _cell(tcol)
                    if tval is None or str(tval).strip() == "":
                        continue
                    ts = _extract_timestamp(None, date_val, tval)
                    if not ts:
                        continue
                    lines.append(f"{pin}\t{ts}\t{state_flag}\t1\t0\t0")
                continue

            timestamp_str = _extract_timestamp(_cell(dt_col), _cell(date_col), _cell(time_col))
            if not timestamp_str:
                continue

            state = _cell(state_col)
            state_flag = _normalise_state(state)

            # Layout: PIN <tab> timestamp <tab> status <tab> verify
            lines.append(f"{pin}\t{timestamp_str}\t{state_flag}\t1\t0\t0")

    return "\n".join(lines)


def _extract_timestamp(dt_val, date_val, time_val) -> str:
    """Return "YYYY-MM-DD HH:MM:SS" or "" if nothing parseable."""

    # Prefer a single combined datetime cell if present
    if isinstance(dt_val, datetime):
        return dt_val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(dt_val, str) and dt_val.strip():
        return _parse_dt_string(dt_val.strip())

    # Otherwise build from separate date + time cells
    dstr = ""
    tstr = "00:00:00"
    if isinstance(date_val, datetime):
        dstr = date_val.strftime("%Y-%m-%d")
    elif isinstance(date_val, date):
        dstr = date_val.strftime("%Y-%m-%d")
    elif isinstance(date_val, str) and date_val.strip():
        # Try common date-only formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                dstr = datetime.strptime(date_val.strip(), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

    if isinstance(time_val, datetime):
        tstr = time_val.strftime("%H:%M:%S")
    elif isinstance(time_val, time):
        tstr = time_val.strftime("%H:%M:%S")
    elif isinstance(time_val, str) and time_val.strip():
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                tstr = datetime.strptime(time_val.strip(), fmt).strftime("%H:%M:%S")
                break
            except ValueError:
                continue

    if not dstr:
        return ""
    return f"{dstr} {tstr}"


def _parse_dt_string(s: str) -> str:
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return ""


def _normalise_state(val) -> str:
    """Convert the In/Out column to the device's 0/1 status codes that
    _handle_attlog understands. 0 = check-in, 1 = check-out. Unknown
    values default to 0 — the time-of-day rules in _apply_to_attendance
    take over anyway."""

    if val is None:
        return "0"
    s = str(val).strip().lower()
    if s in {"1", "out", "check-out", "checkout", "co", "o", "check_out"}:
        return "1"
    if s in {"0", "in", "check-in", "checkin", "ci", "i", "check_in"}:
        return "0"
    # Numeric codes from newer devices (2/3/4/5 = overtime in/out etc.)
    if s.isdigit():
        return s
    return "0"


# =====================================================================
# GET /iclock/import-summary — per-employee calc for a given month
# ---------------------------------------------------------------------
# Read-only: never touches BiometricEvent, Attendance, PayrollSlip.
# Reads Attendance rows for the month and applies the payroll-side
# rules (OT@19:00, late-3x half-day, OT/late offset). See
# app/services/attendance_payroll_calc.py for the full rule set.
# =====================================================================

@router.get("/import-summary")
def import_summary(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    vendor_id: Optional[int] = Query(None),
    working_days: Optional[int] = Query(None, ge=1, le=31),
    db: Session = Depends(get_db),
):
    """Per-employee attendance + payroll calculation for the month.

    Query params
      year          required — YYYY
      month         required — 1..12
      vendor_id     optional — scope to one vendor (default: any)
      working_days  optional — override the auto-computed working
                    days (calendar days − Sundays − holidays)
    """

    from app.services.attendance_payroll_calc import compute_monthly_calculation

    rows = compute_monthly_calculation(
        db,
        year=year,
        month=month,
        vendor_id=vendor_id,
        working_days_override=working_days,
    )

    return {
        "year":       year,
        "month":      month,
        "vendor_id":  vendor_id,
        "employees":  rows,
        "count":      len(rows),
    }
