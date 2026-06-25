from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional, List
import io
import openpyxl

from app.database.database import get_db

from app.models.models import (
    Department,
    Designation,
    Role,
    Permission,
    RolePermission,
    TaskTemplate,
    Vendor,
    CustomField,
    CustomFieldTableValue,
)

from app.schemas.org_schema import (
    DepartmentCreate,
    DepartmentUpdate,
    DesignationCreate,
    DesignationUpdate,
    RoleCreate,
    RolePermissionsSet
)

from app.services.seed_data import (
    ORG_PRESETS,
    PERMISSIONS_CATALOG,
    STANDARD_ROLES
)

from pydantic import BaseModel

class OrgRoleCreate(BaseModel):
    NAME: str
    DEPARTMENT_ID: Optional[int] = None
    DESCRIPTION: Optional[str] = None
    VENDOR_ID: int = 1

class OrgRoleUpdate(BaseModel):
    NAME: Optional[str] = None
    DEPARTMENT_ID: Optional[int] = None
    DESCRIPTION: Optional[str] = None


router = APIRouter()


# =========================
# DEPARTMENTS
# =========================

@router.get("/departments")
def list_departments(
    vendor_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):

    q = db.query(Department)

    if vendor_id is not None:
        q = q.filter(Department.VENDOR_ID == vendor_id)

    if search:
        term = f"%{search}%"
        q = q.filter(
            Department.NAME.ilike(term) | Department.DEPARTMENT_CODE.ilike(term)
        )

    rows = q.order_by(Department.NAME).all()

    return [
        {
            "ID": d.ID,
            "NAME": d.NAME,
            "DEPARTMENT_CODE": d.DEPARTMENT_CODE,
            "DESCRIPTION": d.DESCRIPTION,
            "VENDOR_ID": d.VENDOR_ID,
            "CREATED_AT": d.CREATED_AT.isoformat() if d.CREATED_AT else None,
            "UPDATED_AT": d.UPDATED_AT.isoformat() if d.UPDATED_AT else None
        }
        for d in rows
    ]


@router.post("/departments")
def create_department(
    data: DepartmentCreate,
    db: Session = Depends(get_db)
):

    existing = db.query(Department).filter(
        Department.VENDOR_ID == data.VENDOR_ID,
        Department.DEPARTMENT_CODE == data.DEPARTMENT_CODE.upper()
    ).first()

    if existing:

        raise HTTPException(
            status_code=400,
            detail=f"Department code '{data.DEPARTMENT_CODE}' already exists for this vendor"
        )

    dept = Department(
        NAME=data.NAME,
        DEPARTMENT_CODE=data.DEPARTMENT_CODE.upper(),
        DESCRIPTION=data.DESCRIPTION,
        VENDOR_ID=data.VENDOR_ID
    )

    db.add(dept)

    db.commit()

    db.refresh(dept)

    return {"message": "Department created", "ID": dept.ID}


@router.put("/departments/{dept_id}")
def update_department(
    dept_id: int,
    data: DepartmentUpdate,
    db: Session = Depends(get_db)
):

    dept = db.query(Department).filter(
        Department.ID == dept_id
    ).first()

    if not dept:

        raise HTTPException(status_code=404, detail="Department not found")

    if data.NAME is not None:
        dept.NAME = data.NAME

    if data.DEPARTMENT_CODE is not None:
        dept.DEPARTMENT_CODE = data.DEPARTMENT_CODE.upper()

    if data.DESCRIPTION is not None:
        dept.DESCRIPTION = data.DESCRIPTION

    db.commit()

    return {"message": "Department updated"}


@router.delete("/departments/{dept_id}")
def delete_department(
    dept_id: int,
    db: Session = Depends(get_db)
):

    dept = db.query(Department).filter(
        Department.ID == dept_id
    ).first()

    if not dept:

        raise HTTPException(status_code=404, detail="Department not found")

    in_use = db.query(Designation).filter(
        Designation.DEPARTMENT_ID == dept_id
    ).first()

    if in_use:

        raise HTTPException(
            status_code=400,
            detail="Department has designations. Delete them first."
        )

    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "department",
        CustomFieldTableValue.TABLE_ROW_ID == str(dept_id),
    ).delete(synchronize_session=False)

    db.delete(dept)
    db.commit()
    return {"message": "Department deleted"}


# =========================
# DESIGNATIONS
# =========================

@router.get("/designations")
def list_designations(
    department_id: Optional[int] = Query(None),
    vendor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):

    q = db.query(Designation, Department).join(
        Department,
        Designation.DEPARTMENT_ID == Department.ID
    )

    if department_id is not None:

        q = q.filter(Designation.DEPARTMENT_ID == department_id)

    if vendor_id is not None:

        q = q.filter(Designation.VENDOR_ID == vendor_id)

    rows = q.order_by(Department.NAME, Designation.TITLE).all()

    return [
        {
            "ID": des.ID,
            "TITLE": des.TITLE,
            "DEPARTMENT_ID": des.DEPARTMENT_ID,
            "DEPARTMENT_NAME": dept.NAME,
            "BASE_SALARY": des.BASE_SALARY,
            "DESCRIPTION": des.DESCRIPTION,
            "VENDOR_ID": des.VENDOR_ID
        }
        for des, dept in rows
    ]


@router.post("/designations")
def create_designation(
    data: DesignationCreate,
    db: Session = Depends(get_db)
):

    dept = db.query(Department).filter(
        Department.ID == data.DEPARTMENT_ID
    ).first()

    if not dept:

        raise HTTPException(
            status_code=400,
            detail="Department not found"
        )

    des = Designation(
        TITLE=data.TITLE,
        DEPARTMENT_ID=data.DEPARTMENT_ID,
        BASE_SALARY=data.BASE_SALARY,
        DESCRIPTION=data.DESCRIPTION,
        VENDOR_ID=data.VENDOR_ID
    )

    db.add(des)

    db.commit()

    db.refresh(des)

    return {"message": "Designation created", "ID": des.ID}


@router.put("/designations/{des_id}")
def update_designation(
    des_id: int,
    data: DesignationUpdate,
    db: Session = Depends(get_db)
):

    des = db.query(Designation).filter(
        Designation.ID == des_id
    ).first()

    if not des:

        raise HTTPException(status_code=404, detail="Designation not found")

    if data.TITLE is not None:
        des.TITLE = data.TITLE

    if data.DEPARTMENT_ID is not None:
        des.DEPARTMENT_ID = data.DEPARTMENT_ID

    if data.BASE_SALARY is not None:
        des.BASE_SALARY = data.BASE_SALARY

    if data.DESCRIPTION is not None:
        des.DESCRIPTION = data.DESCRIPTION

    db.commit()

    return {"message": "Designation updated"}


@router.delete("/designations/{des_id}")
def delete_designation(
    des_id: int,
    db: Session = Depends(get_db)
):

    des = db.query(Designation).filter(
        Designation.ID == des_id
    ).first()

    if not des:

        raise HTTPException(status_code=404, detail="Designation not found")

    db.delete(des)

    db.commit()

    return {"message": "Designation deleted"}


# =========================
# PERMISSIONS (read-only)
# =========================

@router.get("/permissions")
def list_permissions(
    db: Session = Depends(get_db)
):

    rows = db.query(Permission).order_by(
        Permission.CATEGORY,
        Permission.CODE
    ).all()

    return [
        {
            "ID": p.ID,
            "CODE": p.CODE,
            "NAME": p.NAME,
            "CATEGORY": p.CATEGORY,
            "DESCRIPTION": p.DESCRIPTION
        }
        for p in rows
    ]


# =========================
# ROLES (per-vendor)
# =========================

@router.get("/roles")
def list_roles(
    vendor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):

    q = db.query(Role)

    if vendor_id is not None:

        q = q.filter(Role.VENDOR_ID == vendor_id)

    rows = q.order_by(Role.NAME).all()

    out = []

    for r in rows:

        perm_ids = [
            rp.PERMISSION_ID
            for rp in db.query(RolePermission).filter(
                RolePermission.ROLE_ID == r.ID
            ).all()
        ]

        out.append({
            "ID": r.ID,
            "NAME": r.NAME,
            "DESCRIPTION": r.DESCRIPTION,
            "VENDOR_ID": r.VENDOR_ID,
            "DEPARTMENT_ID": r.DEPARTMENT_ID,
            "PERMISSION_IDS": perm_ids
        })

    return out


@router.post("/roles")
def create_role(
    data: RoleCreate,
    db: Session = Depends(get_db)
):

    role = Role(
        NAME=data.NAME,
        DESCRIPTION=data.DESCRIPTION,
        VENDOR_ID=data.VENDOR_ID
    )

    db.add(role)

    db.commit()

    db.refresh(role)

    return {"message": "Role created", "ID": role.ID}


@router.delete("/roles/{role_id}")
def delete_role(
    role_id: int,
    db: Session = Depends(get_db)
):

    role = db.query(Role).filter(Role.ID == role_id).first()

    if not role:

        raise HTTPException(status_code=404, detail="Role not found")

    db.query(RolePermission).filter(
        RolePermission.ROLE_ID == role_id
    ).delete()

    db.delete(role)

    db.commit()

    return {"message": "Role deleted"}


@router.put("/roles/{role_id}/permissions")
def set_role_permissions(
    role_id: int,
    data: RolePermissionsSet,
    db: Session = Depends(get_db)
):

    role = db.query(Role).filter(Role.ID == role_id).first()

    if not role:

        raise HTTPException(status_code=404, detail="Role not found")

    # Wipe existing, replace with the new set
    db.query(RolePermission).filter(
        RolePermission.ROLE_ID == role_id
    ).delete()

    for pid in data.PERMISSION_IDS:

        db.add(RolePermission(
            ROLE_ID=role_id,
            PERMISSION_ID=pid
        ))

    db.commit()

    return {
        "message": "Role permissions updated",
        "count": len(data.PERMISSION_IDS)
    }


# =========================
# SEED — Org skeleton
# =========================

def do_seed_org(db: Session, preset_key: str, vendor_id: int) -> dict:
    """Idempotent core of org seeding — callable from both the
    /seed-org endpoint and from main.py startup."""

    preset_key = preset_key.upper()

    if preset_key not in ORG_PRESETS:

        raise ValueError(
            f"Unknown preset '{preset_key}'. "
            f"Available: {list(ORG_PRESETS.keys())}"
        )

    vendor = db.query(Vendor).filter(Vendor.ID == vendor_id).first()

    if not vendor:

        vendor = Vendor(ID=vendor_id, VENDOR_NAME="Bharath Vending Corporation")

        db.add(vendor)

        db.commit()

    # ---- 1. Permissions (global) ----
    perms_added = 0

    code_to_perm = {}

    for code, name, category, desc in PERMISSIONS_CATALOG:

        existing = db.query(Permission).filter(
            Permission.CODE == code
        ).first()

        if existing:

            code_to_perm[code] = existing

            continue

        p = Permission(
            CODE=code,
            NAME=name,
            CATEGORY=category,
            DESCRIPTION=desc
        )

        db.add(p)

        db.flush()

        code_to_perm[code] = p

        perms_added += 1

    db.commit()

    all_perm_ids = [p.ID for p in code_to_perm.values()]

    # ---- 2. Departments + designations ----
    depts_added = 0

    designations_added = 0

    for dept_name, dept_code, designations in ORG_PRESETS[preset_key]["departments"]:

        dept = db.query(Department).filter(
            Department.VENDOR_ID == vendor_id,
            Department.DEPARTMENT_CODE == dept_code
        ).first()

        if not dept:

            dept = Department(
                NAME=dept_name,
                DEPARTMENT_CODE=dept_code,
                VENDOR_ID=vendor_id
            )

            db.add(dept)

            db.flush()

            depts_added += 1

        for title, salary in designations:

            existing_des = db.query(Designation).filter(
                Designation.VENDOR_ID == vendor_id,
                Designation.DEPARTMENT_ID == dept.ID,
                Designation.TITLE == title
            ).first()

            if existing_des:

                continue

            db.add(Designation(
                TITLE=title,
                DEPARTMENT_ID=dept.ID,
                BASE_SALARY=salary,
                VENDOR_ID=vendor_id
            ))

            designations_added += 1

    db.commit()

    # ---- 3. Standard roles + permission assignments ----
    roles_added = 0

    for role_name, role_desc, perm_codes in STANDARD_ROLES:

        role = db.query(Role).filter(
            Role.VENDOR_ID == vendor_id,
            Role.NAME == role_name
        ).first()

        if not role:

            role = Role(
                NAME=role_name,
                DESCRIPTION=role_desc,
                VENDOR_ID=vendor_id
            )

            db.add(role)

            db.flush()

            roles_added += 1

        if perm_codes == "*":

            target_ids = set(all_perm_ids)

        else:

            target_ids = {
                code_to_perm[c].ID
                for c in perm_codes
                if c in code_to_perm
            }

        current_ids = {
            rp.PERMISSION_ID
            for rp in db.query(RolePermission).filter(
                RolePermission.ROLE_ID == role.ID
            ).all()
        }

        for pid in target_ids - current_ids:

            db.add(RolePermission(
                ROLE_ID=role.ID,
                PERMISSION_ID=pid
            ))

    db.commit()

    return {
        "preset": preset_key,
        "vendor_id": vendor_id,
        "permissions_added": perms_added,
        "departments_added": depts_added,
        "designations_added": designations_added,
        "roles_added": roles_added,
        "total_permissions": len(all_perm_ids)
    }


@router.post("/seed-org")
def seed_org(
    preset: str = Query("MANUFACTURING"),
    vendor_id: int = Query(1),
    db: Session = Depends(get_db)
):
    """
    Idempotent seed:
      1. Permissions catalog (global) — only adds missing codes.
      2. Departments + designations for `vendor_id` from the chosen preset.
      3. Standard roles for `vendor_id` with their permission assignments.

    Safe to re-run; existing rows are not overwritten.
    """

    try:

        result = do_seed_org(db, preset, vendor_id)

    except ValueError as exc:

        raise HTTPException(status_code=400, detail=str(exc))

    return {"message": f"Seeded {result['preset']} org structure", **result}


@router.get("/org-presets")
def list_presets():

    return [
        {"key": k, "label": v["label"]}
        for k, v in ORG_PRESETS.items()
    ]


# =========================
# Admin Module 2 — BVC24 9-role catalogue
# =========================

@router.post("/roles/seed-bvc24-catalogue")
def seed_bvc24_role_catalogue(
    vendor_id: int = Query(1),
    db: Session = Depends(get_db)
):
    """Idempotently seed/refresh the BVC24 9-role catalogue (Phase 2
    of Admin Module). Touches only permissions + roles + role_permission
    — leaves departments and designations alone.

    Safe to re-run at any time. Existing role-permission assignments
    are not removed; only missing ones are added (so admin can still
    customise a seeded role afterwards without losing the override on
    the next call)."""

    perms_added = 0

    code_to_perm = {}

    # ---- 1. Permission catalog (global, vendor_id ignored) ----
    for code, name, category, desc in PERMISSIONS_CATALOG:

        existing = db.query(Permission).filter(
            Permission.CODE == code
        ).first()

        if existing:

            code_to_perm[code] = existing

            continue

        p = Permission(
            CODE=code,
            NAME=name,
            CATEGORY=category,
            DESCRIPTION=desc
        )

        db.add(p)

        db.flush()

        code_to_perm[code] = p

        perms_added += 1

    all_perm_ids = [p.ID for p in code_to_perm.values()]

    # ---- 2. Roles + role-permission rows (per-vendor) ----
    roles_added = 0

    roles_touched = []

    for role_name, role_desc, perm_codes in STANDARD_ROLES:

        role = db.query(Role).filter(
            Role.VENDOR_ID == vendor_id,
            Role.NAME == role_name
        ).first()

        if not role:

            role = Role(
                NAME=role_name,
                DESCRIPTION=role_desc,
                VENDOR_ID=vendor_id
            )

            db.add(role)

            db.flush()

            roles_added += 1

        roles_touched.append(role.NAME)

        if perm_codes == "*":

            target_ids = set(all_perm_ids)

        else:

            target_ids = {
                code_to_perm[c].ID
                for c in perm_codes
                if c in code_to_perm
            }

        current_ids = {
            rp.PERMISSION_ID
            for rp in db.query(RolePermission).filter(
                RolePermission.ROLE_ID == role.ID
            ).all()
        }

        for pid in target_ids - current_ids:

            db.add(RolePermission(
                ROLE_ID=role.ID,
                PERMISSION_ID=pid
            ))

    db.commit()

    return {
        "message": (
            f"Catalogue synced. {roles_added} new role(s), "
            f"{perms_added} new permission code(s). "
            f"Total roles in catalogue: {len(roles_touched)}."
        ),
        "vendor_id": vendor_id,
        "permissions_added": perms_added,
        "roles_added": roles_added,
        "roles": roles_touched,
        "bvc24_target_roles": [
            "SUPER_ADMIN", "MANAGING_DIRECTOR", "HR_MANAGER",
            "SALES_MANAGER", "PURCHASE_MANAGER", "PRODUCTION_MANAGER",
            "INVENTORY_MANAGER", "ACCOUNTS_MANAGER", "EMPLOYEE",
        ]
    }


# =========================
# BULK UPLOAD — SHARED HELPERS (org module)
# =========================

def _org_cf_fields(table_name: str, vendor_id: int, db: Session):
    return (
        db.query(CustomField)
          .filter(CustomField.TABLE_NAME == table_name, CustomField.VENDOR_ID == vendor_id)
          .order_by(CustomField.SORT_ORDER, CustomField.FIELD_NAME)
          .all()
    )


def _org_upsert_cf(row_id: str, table_name: str, cf_id: str, value, db: Session):
    stored = value if value else None
    ex = (
        db.query(CustomFieldTableValue)
          .filter(
              CustomFieldTableValue.TABLE_NAME == table_name,
              CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
              CustomFieldTableValue.CUSTOM_FIELD_ID == cf_id,
          )
          .first()
    )
    if ex:
        ex.CUSTOM_FIELD_VALUE = stored
    elif stored is not None:
        db.add(CustomFieldTableValue(
            TABLE_NAME=table_name, TABLE_ROW_ID=str(row_id),
            CUSTOM_FIELD_ID=cf_id, CUSTOM_FIELD_VALUE=stored,
        ))


def _validate_cf_value(field, raw_val) -> Optional[str]:
    """Validate a raw bulk-upload value against the field's type and options.
    Returns an error message string if invalid, or None if valid/empty."""
    import re as _re
    from datetime import date as _d, datetime as _dt

    if raw_val is None or str(raw_val).strip() == "":
        return None

    val = str(raw_val).strip()
    ft  = field.FIELD_TYPE

    if ft == "NUMBER":
        try:
            float(val)
        except ValueError:
            return "Must be a number"

    elif ft == "EMAIL":
        if not _re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", val):
            return "Must be a valid email address (e.g. user@example.com)"

    elif ft == "PHONE":
        if not _re.fullmatch(r"\+?[\d\s\-().]{7,20}", val):
            return "Must be a valid phone number"

    elif ft == "DATE":
        try:
            _d.fromisoformat(val)
        except ValueError:
            return "Must be a valid date (YYYY-MM-DD)"

    elif ft == "DATETIME":
        try:
            _dt.fromisoformat(val)
        except ValueError:
            return "Must be a valid date/time (YYYY-MM-DDTHH:MM)"

    elif ft in ("SELECT", "RADIO"):
        allowed = field.OPTIONS or []
        if allowed and val not in allowed:
            return f'Invalid option "{val}". Allowed values: {", ".join(allowed)}'

    elif ft == "CHECKBOX":
        allowed = set(field.OPTIONS or [])
        if allowed:
            items = [v.strip() for v in val.split(",") if v.strip()] if isinstance(raw_val, str) else [val]
            bad = [v for v in items if v not in allowed]
            if bad:
                return f'Invalid option(s): {", ".join(bad)}. Allowed: {", ".join(field.OPTIONS or [])}'

    return None


def _org_cf_changed(row_id: str, table_name: str, cf_vals: dict, db: Session) -> bool:
    for cf_id, new_val in cf_vals.items():
        row = (
            db.query(CustomFieldTableValue)
              .filter(
                  CustomFieldTableValue.TABLE_NAME == table_name,
                  CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
                  CustomFieldTableValue.CUSTOM_FIELD_ID == cf_id,
              )
              .first()
        )
        old_s = str(row.CUSTOM_FIELD_VALUE) if (row and row.CUSTOM_FIELD_VALUE is not None) else ""
        new_s = str(new_val) if new_val else ""
        if old_s != new_s:
            return True
    return False


def _org_parse_xl(content: bytes, required_sheet: str):
    """Load workbook, validate sheet name (exact/case-sensitive), return (headers_upper, rows)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if required_sheet not in wb.sheetnames:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise HTTPException(
            status_code=400,
            detail=(
                f'Sheet "{required_sheet}" not found in the uploaded file. '
                f'Available sheets: {available}. '
                f'Please use the Template download to get a correctly named workbook.'
            ),
        )
    ws = wb[required_sheet]
    headers: Optional[List[str]] = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper() if c is not None else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(row)
    return headers, rows


def _org_cell(record: dict, *keys) -> str:
    for k in keys:
        v = record.get(k.upper())
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return ""


_DEPT_STD_COLS = {"NAME", "CODE", "DESCRIPTION", "S.NO", "S.N", "SN", ""}
_ROLE_STD_COLS = {"ROLE NAME", "DEPARTMENT NAME", "DESCRIPTION", "S.NO", "S.N", "SN", ""}


# =========================
# DEPARTMENT BULK UPLOAD
# =========================

@router.post("/departments/bulk-upload")
async def bulk_upload_departments(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    content = await file.read()
    headers, data_rows = _org_parse_xl(content, "Departments")

    cf_fields   = _org_cf_fields("department", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols     = [h for h in headers if h not in _DEPT_STD_COLS and h in cf_by_upper]

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        name = _org_cell(record, "NAME")
        code = _org_cell(record, "CODE")
        desc = _org_cell(record, "DESCRIPTION") or None

        if not name:
            errors.append({"row": row_num, "field": "Name", "message": "Name is required"})
            continue
        if not code:
            errors.append({"row": row_num, "field": "Code", "message": "Code is required"})
            continue

        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _org_cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({"row": row_num, "field": cf_f.FIELD_NAME,
                                "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing'})
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        existing = db.query(Department).filter(
            Department.VENDOR_ID == vendor_id,
            Department.DEPARTMENT_CODE == code.upper()
        ).first()

        if existing:
            row_changed = (
                (name or "") != (existing.NAME or "") or
                (desc or "") != (existing.DESCRIPTION or "")
            )
            if row_changed or _org_cf_changed(existing.ID, "department", cf_vals, db):
                if row_changed:
                    existing.NAME        = name
                    existing.DESCRIPTION = desc
                for cf_id, val in cf_vals.items():
                    _org_upsert_cf(existing.ID, "department", cf_id, val, db)
                updated += 1
            else:
                skipped += 1
        else:
            new_dept = Department(
                NAME=name, DEPARTMENT_CODE=code.upper(),
                DESCRIPTION=desc, VENDOR_ID=vendor_id,
            )
            db.add(new_dept)
            db.flush()
            for cf_id, val in cf_vals.items():
                _org_upsert_cf(new_dept.ID, "department", cf_id, val, db)
            inserted += 1

    db.commit()
    total = inserted + updated + skipped + len(errors)
    msg   = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {"message": msg, "inserted": inserted, "updated": updated,
            "skipped": skipped, "total_rows": total, "errors": errors}


# =========================
# ORG ROLES (job-function roles, separate from RBAC)
# =========================

@router.get("/org-roles")
def list_org_roles(
    vendor_id: Optional[int] = Query(None),
    dept_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(Role, Department).outerjoin(
        Department, Role.DEPARTMENT_ID == Department.ID
    )
    if vendor_id is not None:
        q = q.filter(Role.VENDOR_ID == vendor_id)
    if dept_id is not None:
        q = q.filter(Role.DEPARTMENT_ID == dept_id)
    if search:
        term = f"%{search}%"
        q = q.filter(Role.NAME.ilike(term))
    rows = q.order_by(Role.NAME).all()
    return [
        {
            "ID": r.ID,
            "NAME": r.NAME,
            "DESCRIPTION": r.DESCRIPTION,
            "DEPARTMENT_ID": r.DEPARTMENT_ID,
            "DEPARTMENT_NAME": d.NAME if d else None,
            "VENDOR_ID": r.VENDOR_ID
        }
        for r, d in rows
    ]


@router.post("/org-roles")
def create_org_role(
    data: OrgRoleCreate,
    db: Session = Depends(get_db)
):
    existing = db.query(Role).filter(
        Role.VENDOR_ID == data.VENDOR_ID,
        Role.NAME == data.NAME
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Role '{data.NAME}' already exists")
    role = Role(
        NAME=data.NAME,
        DESCRIPTION=data.DESCRIPTION,
        DEPARTMENT_ID=data.DEPARTMENT_ID,
        VENDOR_ID=data.VENDOR_ID
    )
    db.add(role)
    db.commit()
    db.refresh(role)
    return {"message": "Role created", "ID": role.ID}


@router.put("/org-roles/{role_id}")
def update_org_role(
    role_id: int,
    data: OrgRoleUpdate,
    db: Session = Depends(get_db)
):
    role = db.query(Role).filter(Role.ID == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if data.NAME is not None:
        role.NAME = data.NAME
    if data.DEPARTMENT_ID is not None:
        role.DEPARTMENT_ID = data.DEPARTMENT_ID
    if data.DESCRIPTION is not None:
        role.DESCRIPTION = data.DESCRIPTION
    db.commit()
    return {"message": "Role updated"}


@router.delete("/org-roles/{role_id}")
def delete_org_role(
    role_id: int,
    db: Session = Depends(get_db)
):
    role = db.query(Role).filter(Role.ID == role_id).first()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    in_use = db.query(TaskTemplate).filter(TaskTemplate.ROLE_ID == role_id).first()
    if in_use:
        raise HTTPException(
            status_code=400,
            detail="Role is referenced by task templates. Remove those references first."
        )
    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "role",
        CustomFieldTableValue.TABLE_ROW_ID == str(role_id),
    ).delete(synchronize_session=False)
    db.delete(role)
    db.commit()
    return {"message": "Role deleted"}


@router.post("/org-roles/bulk-upload")
async def bulk_upload_org_roles(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    content = await file.read()
    headers, data_rows = _org_parse_xl(content, "Roles")

    cf_fields   = _org_cf_fields("role", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols     = [h for h in headers if h not in _ROLE_STD_COLS and h in cf_by_upper]

    # Pre-build department lookup (case-insensitive name)
    dept_name_map = {
        d.NAME.strip().lower(): d
        for d in db.query(Department).filter(Department.VENDOR_ID == vendor_id).all()
    }

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        role_name = _org_cell(record, "ROLE NAME")
        dept_name = _org_cell(record, "DEPARTMENT NAME")
        desc      = _org_cell(record, "DESCRIPTION") or None

        if not role_name:
            errors.append({"row": row_num, "field": "Role Name", "message": "Role Name is required"})
            continue

        dept_id = None
        if dept_name:
            d = dept_name_map.get(dept_name.lower())
            if not d:
                errors.append({"row": row_num, "field": "Department Name",
                                "message": f'Department "{dept_name}" does not exist'})
                continue
            dept_id = d.ID

        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _org_cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({"row": row_num, "field": cf_f.FIELD_NAME,
                                "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing'})
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        existing = db.query(Role).filter(
            Role.VENDOR_ID == vendor_id,
            Role.NAME == role_name,
        ).first()

        if existing:
            row_changed = (
                (desc or "") != (existing.DESCRIPTION or "") or
                dept_id != existing.DEPARTMENT_ID
            )
            if row_changed or _org_cf_changed(existing.ID, "role", cf_vals, db):
                if row_changed:
                    existing.DESCRIPTION = desc
                    existing.DEPARTMENT_ID = dept_id
                for cf_id, val in cf_vals.items():
                    _org_upsert_cf(existing.ID, "role", cf_id, val, db)
                updated += 1
            else:
                skipped += 1
        else:
            new_role = Role(
                NAME=role_name, DESCRIPTION=desc,
                DEPARTMENT_ID=dept_id, VENDOR_ID=vendor_id,
            )
            db.add(new_role)
            db.flush()
            for cf_id, val in cf_vals.items():
                _org_upsert_cf(new_role.ID, "role", cf_id, val, db)
            inserted += 1

    db.commit()
    total = inserted + updated + skipped + len(errors)
    msg   = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {"message": msg, "inserted": inserted, "updated": updated,
            "skipped": skipped, "total_rows": total, "errors": errors}
