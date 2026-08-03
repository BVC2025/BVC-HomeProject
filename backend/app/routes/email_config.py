import io
import openpyxl

from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database.database import get_db
from app.models.email_models import VendorEmailConfig

router = APIRouter()

_PASSWORD_MASK = "●●●●●●●●"

# ── Excel column definitions ──────────────────────────────────────────────────
_TEMPLATE_HEADERS = [
    "SMTP Host", "SMTP Port", "SMTP Username", "SMTP Password",
    "From Name", "From Email", "BCC Name", "BCC Email", "Is Active",
]
_EXPORT_HEADERS = [
    "S.No", "SMTP Host", "SMTP Port", "SMTP Username",
    "From Name", "From Email", "BCC Name", "BCC Email",
    "Is Active", "Created At",
]


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class EmailConfigCreate(BaseModel):
    SMTP_HOST: str
    SMTP_PORT: int = 587
    SMTP_USERNAME: str
    SMTP_PASSWORD: str
    FROM_NAME: str
    FROM_EMAIL: str
    BCC_NAME: Optional[str] = None
    BCC_EMAIL: Optional[str] = None
    IS_ACTIVE: bool = False


class EmailConfigUpdate(BaseModel):
    SMTP_HOST: Optional[str] = None
    SMTP_PORT: Optional[int] = None
    SMTP_USERNAME: Optional[str] = None
    SMTP_PASSWORD: Optional[str] = None  # None = preserve existing password
    FROM_NAME: Optional[str] = None
    FROM_EMAIL: Optional[str] = None
    BCC_NAME: Optional[str] = None
    BCC_EMAIL: Optional[str] = None
    IS_ACTIVE: Optional[bool] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _serialize(cfg: VendorEmailConfig, mask: bool = True) -> dict:
    return {
        "ID":            cfg.ID,
        "VENDOR_ID":     cfg.VENDOR_ID,
        "SMTP_HOST":     cfg.SMTP_HOST,
        "SMTP_PORT":     cfg.SMTP_PORT,
        "SMTP_USERNAME": cfg.SMTP_USERNAME,
        "SMTP_PASSWORD": _PASSWORD_MASK if mask else cfg.SMTP_PASSWORD,
        "FROM_NAME":     cfg.FROM_NAME,
        "FROM_EMAIL":    cfg.FROM_EMAIL,
        "BCC_NAME":      cfg.BCC_NAME,
        "BCC_EMAIL":     cfg.BCC_EMAIL,
        "IS_ACTIVE":     cfg.IS_ACTIVE,
        "CREATED_AT":    cfg.CREATED_AT.isoformat() if cfg.CREATED_AT else None,
        "UPDATED_AT":    cfg.UPDATED_AT.isoformat() if cfg.UPDATED_AT else None,
    }


def _get_or_404(db: Session, config_id: str) -> VendorEmailConfig:
    cfg = db.query(VendorEmailConfig).filter(VendorEmailConfig.ID == config_id).first()
    if not cfg:
        raise HTTPException(status_code=404, detail="Email configuration not found")
    return cfg


def _parse_xl(content: bytes) -> tuple:
    """Parse the uploaded xlsx. Returns (headers, data_rows)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    sheet_name = "Email Configs"
    if sheet_name not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f'Sheet "{sheet_name}" not found. Expected a sheet named "{sheet_name}".'
        )

    ws = wb[sheet_name]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(row)

    return headers, rows


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/email-configs")
def list_email_configs(
    vendor_id: int = Query(1),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """List all email configurations for a vendor. Passwords are always masked."""
    q = db.query(VendorEmailConfig).filter(VendorEmailConfig.VENDOR_ID == vendor_id)
    if search:
        term = f"%{search.strip()}%"
        q = q.filter(
            VendorEmailConfig.SMTP_HOST.ilike(term)
            | VendorEmailConfig.FROM_EMAIL.ilike(term)
            | VendorEmailConfig.FROM_NAME.ilike(term)
        )
    rows = q.order_by(VendorEmailConfig.CREATED_AT.desc()).all()
    return [_serialize(r) for r in rows]


@router.post("/email-configs", status_code=201)
def create_email_config(
    data: EmailConfigCreate,
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    """Create a new email configuration. If IS_ACTIVE=True, all other configs are deactivated."""
    # Uniqueness check: (vendor_id, from_email)
    existing = db.query(VendorEmailConfig).filter(
        VendorEmailConfig.VENDOR_ID == vendor_id,
        VendorEmailConfig.FROM_EMAIL == data.FROM_EMAIL.strip().lower(),
    ).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"A configuration with From Email '{data.FROM_EMAIL}' already exists for this vendor."
        )

    cfg = VendorEmailConfig(
        VENDOR_ID     = vendor_id,
        SMTP_HOST     = data.SMTP_HOST.strip(),
        SMTP_PORT     = data.SMTP_PORT,
        SMTP_USERNAME = data.SMTP_USERNAME.strip(),
        SMTP_PASSWORD = data.SMTP_PASSWORD,
        FROM_NAME     = data.FROM_NAME.strip(),
        FROM_EMAIL    = data.FROM_EMAIL.strip().lower(),
        BCC_NAME      = data.BCC_NAME.strip() if data.BCC_NAME else None,
        BCC_EMAIL     = data.BCC_EMAIL.strip().lower() if data.BCC_EMAIL else None,
        IS_ACTIVE     = data.IS_ACTIVE,
    )
    db.add(cfg)
    db.commit()
    db.refresh(cfg)
    return {"message": "Email configuration created", "ID": cfg.ID, **_serialize(cfg)}


@router.get("/email-configs/active")
def get_active_email_config(
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    """Return all active email configurations with real passwords (server-side / future integration)."""
    cfgs = db.query(VendorEmailConfig).filter(
        VendorEmailConfig.VENDOR_ID == vendor_id,
        VendorEmailConfig.IS_ACTIVE == True,
    ).all()
    return {"configs": [_serialize(c, mask=False) for c in cfgs]}


@router.get("/email-configs/bulk-upload/template")
def download_email_config_template():
    """Download a blank xlsx template for bulk upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Configs"
    ws.append(_TEMPLATE_HEADERS)

    # Sample row so the user can see the expected format
    ws.append([
        "smtp.gmail.com", 587, "user@gmail.com", "your-app-password",
        "Company ERP", "erp@company.com", "", "", "FALSE",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="email_configs_template.xlsx"'},
    )


@router.post("/email-configs/bulk-upload")
def bulk_upload_email_configs(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upsert email configurations from an xlsx file (sheet: 'Email Configs').
    Upsert key: (VENDOR_ID, FROM_EMAIL). Passwords are required on insert; on update,
    an empty password cell leaves the existing password unchanged.
    """
    import asyncio
    content = asyncio.get_event_loop().run_until_complete(file.read()) if False else file.file.read()
    headers, data_rows = _parse_xl(content)

    if not headers:
        raise HTTPException(status_code=400, detail="Template has no header row.")

    # Normalise header names for lookup
    h = {name.strip().upper(): idx for idx, name in enumerate(headers)}

    def col(row, name: str, default=None):
        idx = h.get(name.upper())
        return row[idx] if idx is not None and idx < len(row) else default

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, row in enumerate(data_rows, start=2):
        smtp_host = str(col(row, "SMTP HOST") or "").strip()
        smtp_port_raw = col(row, "SMTP PORT")
        smtp_user = str(col(row, "SMTP USERNAME") or "").strip()
        smtp_pass = str(col(row, "SMTP PASSWORD") or "").strip()
        from_name = str(col(row, "FROM NAME") or "").strip()
        from_email = str(col(row, "FROM EMAIL") or "").strip().lower()
        bcc_name = str(col(row, "BCC NAME") or "").strip() or None
        bcc_email = str(col(row, "BCC EMAIL") or "").strip().lower() or None
        is_active_raw = str(col(row, "IS ACTIVE") or "").strip().upper()

        # Validate required fields
        missing = []
        if not smtp_host:   missing.append("SMTP Host")
        if not smtp_user:   missing.append("SMTP Username")
        if not from_name:   missing.append("From Name")
        if not from_email:  missing.append("From Email")
        if missing:
            errors.append({"row": row_num, "field": ", ".join(missing), "message": "Required field(s) missing"})
            continue

        try:
            smtp_port = int(smtp_port_raw) if smtp_port_raw is not None else 587
        except (ValueError, TypeError):
            errors.append({"row": row_num, "field": "SMTP Port", "message": "Must be an integer"})
            continue

        is_active = is_active_raw in ("TRUE", "1", "YES")

        existing = db.query(VendorEmailConfig).filter(
            VendorEmailConfig.VENDOR_ID == vendor_id,
            VendorEmailConfig.FROM_EMAIL == from_email,
        ).first()

        if existing:
            changed = False
            if smtp_host     and existing.SMTP_HOST     != smtp_host:   existing.SMTP_HOST     = smtp_host;  changed = True
            if smtp_port     and existing.SMTP_PORT     != smtp_port:   existing.SMTP_PORT     = smtp_port;  changed = True
            if smtp_user     and existing.SMTP_USERNAME != smtp_user:   existing.SMTP_USERNAME = smtp_user;  changed = True
            if smtp_pass:  # empty = keep existing
                if existing.SMTP_PASSWORD != smtp_pass: existing.SMTP_PASSWORD = smtp_pass; changed = True
            if from_name     and existing.FROM_NAME     != from_name:   existing.FROM_NAME     = from_name;  changed = True
            if bcc_name is not None and existing.BCC_NAME   != bcc_name:   existing.BCC_NAME   = bcc_name;  changed = True
            if bcc_email is not None and existing.BCC_EMAIL != bcc_email:  existing.BCC_EMAIL  = bcc_email; changed = True
            if existing.IS_ACTIVE != is_active:                             existing.IS_ACTIVE  = is_active; changed = True

            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            if not smtp_pass:
                errors.append({"row": row_num, "field": "SMTP Password", "message": "Required for new records"})
                continue

            new_cfg = VendorEmailConfig(
                VENDOR_ID     = vendor_id,
                SMTP_HOST     = smtp_host,
                SMTP_PORT     = smtp_port,
                SMTP_USERNAME = smtp_user,
                SMTP_PASSWORD = smtp_pass,
                FROM_NAME     = from_name,
                FROM_EMAIL    = from_email,
                BCC_NAME      = bcc_name,
                BCC_EMAIL     = bcc_email,
                IS_ACTIVE     = is_active,
            )
            db.add(new_cfg)
            inserted += 1

    db.commit()
    return {
        "message": f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped, {len(errors)} error(s)",
        "inserted":    inserted,
        "updated":     updated,
        "skipped":     skipped,
        "total_rows":  len(data_rows),
        "errors":      errors,
    }


@router.get("/email-configs/export/excel")
def export_email_configs(
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    """Export all email configurations for a vendor as xlsx. Password column is excluded."""
    rows = (
        db.query(VendorEmailConfig)
        .filter(VendorEmailConfig.VENDOR_ID == vendor_id)
        .order_by(VendorEmailConfig.CREATED_AT.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Configs"
    ws.append(_EXPORT_HEADERS)

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r.SMTP_HOST,
            r.SMTP_PORT,
            r.SMTP_USERNAME,
            r.FROM_NAME,
            r.FROM_EMAIL,
            r.BCC_NAME or "",
            r.BCC_EMAIL or "",
            "Active" if r.IS_ACTIVE else "Inactive",
            r.CREATED_AT.isoformat() if r.CREATED_AT else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="email_configs_export.xlsx"'},
    )


@router.get("/email-configs/{config_id}")
def get_email_config(
    config_id: str,
    db: Session = Depends(get_db),
):
    """Get a single email configuration by ID. Password is masked."""
    cfg = _get_or_404(db, config_id)
    return _serialize(cfg)


@router.put("/email-configs/{config_id}")
def update_email_config(
    config_id: str,
    data: EmailConfigUpdate,
    db: Session = Depends(get_db),
):
    """Update an email configuration. Omit SMTP_PASSWORD (or send null) to preserve the existing password."""
    cfg = _get_or_404(db, config_id)

    if data.SMTP_HOST     is not None: cfg.SMTP_HOST     = data.SMTP_HOST.strip()
    if data.SMTP_PORT     is not None: cfg.SMTP_PORT     = data.SMTP_PORT
    if data.SMTP_USERNAME is not None: cfg.SMTP_USERNAME = data.SMTP_USERNAME.strip()
    # Only update password when a non-empty value is provided
    if data.SMTP_PASSWORD and data.SMTP_PASSWORD.strip():
        cfg.SMTP_PASSWORD = data.SMTP_PASSWORD.strip()
    if data.FROM_NAME     is not None: cfg.FROM_NAME     = data.FROM_NAME.strip()
    if data.FROM_EMAIL    is not None:
        new_email = data.FROM_EMAIL.strip().lower()
        if new_email != cfg.FROM_EMAIL:
            conflict = db.query(VendorEmailConfig).filter(
                VendorEmailConfig.VENDOR_ID == cfg.VENDOR_ID,
                VendorEmailConfig.FROM_EMAIL == new_email,
                VendorEmailConfig.ID != config_id,
            ).first()
            if conflict:
                raise HTTPException(status_code=409, detail=f"From Email '{new_email}' is already used by another configuration.")
        cfg.FROM_EMAIL = new_email
    if data.BCC_NAME  is not None: cfg.BCC_NAME  = data.BCC_NAME.strip()  or None
    if data.BCC_EMAIL is not None: cfg.BCC_EMAIL = data.BCC_EMAIL.strip().lower() or None
    if data.IS_ACTIVE is not None:
        cfg.IS_ACTIVE = data.IS_ACTIVE

    db.commit()
    db.refresh(cfg)
    return {"message": "Email configuration updated", **_serialize(cfg)}


@router.delete("/email-configs/{config_id}")
def delete_email_config(
    config_id: str,
    db: Session = Depends(get_db),
):
    """Hard-delete an email configuration."""
    cfg = _get_or_404(db, config_id)
    db.delete(cfg)
    db.commit()
    return {"message": "Email configuration deleted"}


@router.post("/email-configs/{config_id}/activate")
def activate_email_config(
    config_id: str,
    db: Session = Depends(get_db),
):
    """Activate an email configuration."""
    cfg = _get_or_404(db, config_id)
    cfg.IS_ACTIVE = True
    db.commit()
    return {"message": "Configuration activated", **_serialize(cfg)}


@router.post("/email-configs/{config_id}/deactivate")
def deactivate_email_config(
    config_id: str,
    db: Session = Depends(get_db),
):
    """Deactivate an email configuration."""
    cfg = _get_or_404(db, config_id)
    cfg.IS_ACTIVE = False
    db.commit()
    return {"message": "Configuration deactivated", **_serialize(cfg)}
