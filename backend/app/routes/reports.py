import os
from io import BytesIO
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage
)

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.drawing.image import Image as XLImage

from app.database.database import get_db

from app.models.models import (
    Employee,
    Customer,
    Project,
    Task,
    Inventory,
    Attendance,
    Machine
)

router = APIRouter()


COMPANY_NAME = "Bharath Vending Corporation"
COMPANY_TAGLINE = "Manufacturing Management System"

LOGO_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "assets",
    "bharath-logo.png"
)

HEADER_COLOR = colors.HexColor("#0f172a")
ACCENT_COLOR = colors.HexColor("#dc2626")
SUBTLE_COLOR = colors.HexColor("#64748b")


# =========================
# DATA EXTRACTORS
# =========================

def employees_data(db):

    rows = db.query(Employee).all()

    return (
        ["Code", "Name", "Email", "Phone", "Status", "Role ID"],
        [
            [
                r.EMPLOYEE_CODE or "",
                r.NAME or "",
                r.EMAIL or "",
                r.PHONE or "",
                r.STATUS or "",
                r.ROLE_ID or ""
            ]
            for r in rows
        ]
    )


def customers_data(db):

    rows = db.query(Customer).all()

    return (
        ["ID", "Name", "Phone", "Email", "Address"],
        [
            [
                r.ID,
                r.CUSTOMER_NAME or "",
                r.PHONE or "",
                r.EMAIL or "",
                r.ADDRESS or ""
            ]
            for r in rows
        ]
    )


def projects_data(db):

    rows = db.query(Project).all()

    return (
        [
            "ID",
            "Project Name",
            "Description",
            "Status",
            "Customer ID"
        ],
        [
            [
                r.ID,
                r.PROJECT_NAME or "",
                r.DESCRIPTION or "",
                r.STATUS or "",
                r.CUSTOMER_ID or ""
            ]
            for r in rows
        ]
    )


def tasks_data(db):

    rows = db.query(Task).all()

    return (
        [
            "ID",
            "Task Name",
            "Description",
            "Status",
            "Priority",
            "Project ID"
        ],
        [
            [
                r.ID,
                r.TASK_NAME or "",
                r.DESCRIPTION or "",
                r.STATUS or "",
                r.PRIORITY or "",
                r.PROJECT_ID or ""
            ]
            for r in rows
        ]
    )


def inventory_data(db):

    rows = db.query(Inventory).all()

    return (
        [
            "ID",
            "Material Name",
            "Quantity",
            "Unit Price",
            "Total Value"
        ],
        [
            [
                r.ID,
                r.MATERIAL_NAME or "",
                r.QUANTITY or 0,
                r.UNIT_PRICE or 0,
                round(
                    (r.QUANTITY or 0)
                    * (r.UNIT_PRICE or 0),
                    2
                )
            ]
            for r in rows
        ]
    )


def attendance_data(db):

    rows = db.query(
        Attendance,
        Employee.NAME
    ).outerjoin(
        Employee,
        Attendance.EMPLOYEE_ID == Employee.ID
    ).order_by(
        Attendance.DATE.desc()
    ).all()

    return (
        [
            "ID",
            "Date",
            "Employee",
            "Check In",
            "Check Out",
            "Status",
            "Hours"
        ],
        [
            [
                r.ID,
                r.DATE.isoformat() if r.DATE else "",
                name or r.EMPLOYEE_ID or "",
                (
                    r.CHECK_IN.strftime("%H:%M")
                    if r.CHECK_IN else "—"
                ),
                (
                    r.CHECK_OUT.strftime("%H:%M")
                    if r.CHECK_OUT else "—"
                ),
                r.STATUS or "",
                (
                    round(
                        (
                            r.CHECK_OUT - r.CHECK_IN
                        ).total_seconds() / 3600,
                        2
                    )
                    if r.CHECK_IN and r.CHECK_OUT
                    else "—"
                )
            ]
            for r, name in rows
        ]
    )


def machines_data(db):

    rows = db.query(Machine).all()

    return (
        [
            "ID",
            "Name",
            "Type",
            "Location",
            "Status",
            "Last Updated"
        ],
        [
            [
                r.ID,
                r.MACHINE_NAME or "",
                r.MACHINE_TYPE or "",
                r.LOCATION or "—",
                r.STATUS or "",
                (
                    r.LAST_UPDATED.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if r.LAST_UPDATED else ""
                )
            ]
            for r in rows
        ]
    )


MODULES = {
    "employees": ("Employees", employees_data),
    "customers": ("Customers", customers_data),
    "projects": ("Projects", projects_data),
    "tasks": ("Tasks", tasks_data),
    "inventory": ("Inventory", inventory_data),
    "attendance": ("Attendance", attendance_data),
    "machines": ("Machines", machines_data)
}


# =========================
# PDF BUILDER
# =========================

def _resolve_company():
    """Fetch company branding from CompanyMaster at PDF-build time.
    Returns a tuple (legal_name, tagline, logo_disk_path)."""

    try:

        from app.database.database import SessionLocal

        from app.services.company_settings_service import get_company_settings

        db = SessionLocal()

        try:

            c = get_company_settings(db, 1)

            name = c.LEGAL_NAME or COMPANY_NAME

            tag = c.TAGLINE or COMPANY_TAGLINE

            logo_path = LOGO_PATH

            if c.LOGO_URL:

                # Map /static/company/<file> → backend/static/company/<file>
                rel = c.LOGO_URL.split("/static/", 1)[-1]

                disk = os.path.join(
                    os.path.dirname(__file__),
                    "..", "..", "static", rel
                )

                if os.path.exists(disk):

                    logo_path = disk

            return name, tag, logo_path

        finally:

            db.close()

    except Exception:

        return COMPANY_NAME, COMPANY_TAGLINE, LOGO_PATH


def build_pdf(title, headers, rows):

    company_name, company_tagline, logo_path = _resolve_company()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(
        "Company",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=HEADER_COLOR,
        spaceAfter=2,
        leading=22
    )

    tagline_style = ParagraphStyle(
        "Tagline",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=SUBTLE_COLOR,
        spaceAfter=0,
        leading=12
    )

    report_title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=ACCENT_COLOR,
        alignment=TA_RIGHT,
        spaceAfter=2,
        leading=18
    )

    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=SUBTLE_COLOR,
        alignment=TA_RIGHT,
        leading=11
    )

    elements = []

    left_cell = []

    if os.path.exists(logo_path):

        left_cell.append(
            RLImage(logo_path, width=22 * mm, height=22 * mm)
        )

    brand_text = [
        Paragraph(company_name, company_style),
        Paragraph(company_tagline, tagline_style)
    ]

    right_text = [
        Paragraph(f"{title} Report", report_title_style),
        Paragraph(
            "Generated "
            + datetime.now().strftime("%d %b %Y, %H:%M")
            + f"  |  {len(rows)} records",
            meta_style
        )
    ]

    header_table = Table(
        [[left_cell, brand_text, right_text]],
        colWidths=[26 * mm, 120 * mm, 120 * mm]
    )

    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0)
    ]))

    elements.append(header_table)

    elements.append(Spacer(1, 4 * mm))

    divider = Table(
        [[""]],
        colWidths=[266 * mm],
        rowHeights=[2]
    )

    divider.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), ACCENT_COLOR),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0)
    ]))

    elements.append(divider)

    elements.append(Spacer(1, 6 * mm))

    data = [headers] + [
        [str(cell) for cell in row]
        for row in rows
    ]

    if not rows:

        data.append(
            ["No records found"]
            + [""] * (len(headers) - 1)
        )

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor("#0f172a")
        ),
        (
            "TEXTCOLOR",
            (0, 0),
            (-1, 0),
            colors.whitesmoke
        ),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [
                colors.white,
                colors.HexColor("#f1f5f9")
            ]
        ),
        (
            "GRID",
            (0, 0),
            (-1, -1),
            0.5,
            colors.HexColor("#cbd5e1")
        )
    ]))

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return buffer


# =========================
# EXCEL BUILDER
# =========================

def build_excel(title, headers, rows):

    company_name, _company_tagline, _logo_path = _resolve_company()

    wb = Workbook()

    ws = wb.active

    ws.title = title[:30]

    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 60

    ws.row_dimensions[3].height = 20

    ws.row_dimensions[4].height = 18

    ws["B2"] = company_name

    ws["B2"].font = Font(size=18, bold=True, color="0F172A")

    ws["B2"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    ws.merge_cells(
        start_row=2,
        start_column=2,
        end_row=2,
        end_column=max(len(headers), 2)
    )

    ws["B3"] = f"{title} Report"

    ws["B3"].font = Font(size=13, bold=True, color="DC2626")

    ws["B3"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    ws.merge_cells(
        start_row=3,
        start_column=2,
        end_row=3,
        end_column=max(len(headers), 2)
    )

    ws["B4"] = (
        "Generated "
        + datetime.now().strftime("%d %b %Y, %H:%M")
        + f"  |  {len(rows)} records"
    )

    ws["B4"].font = Font(size=10, italic=True, color="64748B")

    ws["B4"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    ws.merge_cells(
        start_row=4,
        start_column=2,
        end_row=4,
        end_column=max(len(headers), 2)
    )

    if os.path.exists(_logo_path):

        try:

            img = XLImage(_logo_path)

            img.width = 60

            img.height = 60

            ws.column_dimensions["A"].width = 11

            ws.add_image(img, "A2")

        except Exception:

            pass

    header_row = 6

    header_fill = PatternFill(
        start_color="0F172A",
        end_color="0F172A",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin = Side(border_style="thin", color="CBD5E1")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for col_idx, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=header
        )

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        cell.border = border

    for r_idx, row in enumerate(
        rows,
        start=header_row + 1
    ):

        for c_idx, value in enumerate(row, start=1):

            cell = ws.cell(
                row=r_idx,
                column=c_idx,
                value=value
            )

            cell.border = border

            cell.alignment = Alignment(
                vertical="center"
            )

    # auto-size columns
    for col_idx, header in enumerate(headers, start=1):

        column_letter = ws.cell(
            row=header_row,
            column=col_idx
        ).column_letter

        max_len = len(str(header))

        for row in rows:

            if col_idx - 1 < len(row):

                val = str(row[col_idx - 1])

                if len(val) > max_len:

                    max_len = len(val)

        ws.column_dimensions[
            column_letter
        ].width = min(max(max_len + 4, 12), 50)

    buffer = BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer


# =========================
# ENDPOINTS
# =========================

def _get_module(module: str):

    key = module.lower()

    if key not in MODULES:

        raise HTTPException(
            status_code=404,
            detail=(
                "Unknown module. Valid: "
                + ", ".join(MODULES.keys())
            )
        )

    return MODULES[key]


@router.get("/report/{module}.pdf")
def report_pdf(
    module: str,
    db: Session = Depends(get_db)
):

    title, extractor = _get_module(module)

    headers, rows = extractor(db)

    buffer = build_pdf(title, headers, rows)

    filename = (
        f"{module}_report_"
        f"{date.today().isoformat()}.pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        }
    )


@router.get("/report/{module}.xlsx")
def report_excel(
    module: str,
    db: Session = Depends(get_db)
):

    title, extractor = _get_module(module)

    headers, rows = extractor(db)

    buffer = build_excel(title, headers, rows)

    filename = (
        f"{module}_report_"
        f"{date.today().isoformat()}.xlsx"
    )

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"'
            )
        }
    )
