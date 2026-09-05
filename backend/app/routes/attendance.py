from fastapi import APIRouter, Depends, HTTPException, Request, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, time
from typing import Optional
from pydantic import BaseModel

from app.database.database import get_db

from app.models.models import (
    Attendance,
    Employee,
    TaskAssignment,
    Department,
    GeofenceSettings,
    AttendanceSecurityLog
)

from app.schemas.attendance_schema import (
    CheckInRequest,
    CheckOutRequest,
    MarkAbsentRequest
)

from app.routes.geofence import (
    haversine_meters,
    _get_or_create_settings as get_geofence_settings
)

from app.utils.employee_resolver import resolve_employee_uuid

from app.auth.auth_bearer import (
    get_current_admin,
    get_current_user,
    assert_self_or_admin,
    require,
    ADMIN_ROLES,
)

router = APIRouter()


# =========================
# Geofence helpers (used by both check-in and check-out)
# =========================

def _check_geofence(
    db: Session,
    vendor_id: int,
    lat: float | None,
    lng: float | None
) -> dict:
    """Return {allowed, distance_m, status, settings}.

    - allowed: True if inside radius OR enforcement is OFF OR no
      coordinates were sent (back-compat with legacy callers).
    - distance_m: metres from the configured office (None if no coords).
    - status: 'INSIDE' / 'OUTSIDE' / 'UNKNOWN'.
    """

    settings = get_geofence_settings(db, vendor_id)

    if lat is None or lng is None:

        return {
            "allowed": True,                  # back-compat
            "distance_m": None,
            "status": "UNKNOWN",
            "settings": settings
        }

    distance = haversine_meters(
        lat, lng, settings.LATITUDE, settings.LONGITUDE
    )

    inside = distance <= settings.RADIUS_METERS

    enforced = bool(settings.IS_ACTIVE)

    return {
        "allowed": inside or (not enforced),
        "distance_m": round(distance, 2),
        "status": "INSIDE" if inside else "OUTSIDE",
        "settings": settings
    }


def _log_failure(
    db: Session,
    employee_id: str | None,
    reason: str,
    lat: float | None,
    lng: float | None,
    distance: float | None,
    detail: str | None,
    device_info: str | None,
    ip: str | None,
    vendor_id: int
):
    """Record a blocked attempt in the security log table."""

    row = AttendanceSecurityLog(
        EMPLOYEE_ID=employee_id,
        LATITUDE=lat,
        LONGITUDE=lng,
        DISTANCE=distance,
        REASON=reason[:80],
        DETAIL=(detail or "")[:500] or None,
        DEVICE_INFO=(device_info or "")[:255] or None,
        IP_ADDRESS=ip,
        VENDOR_ID=vendor_id
    )

    db.add(row)

    db.commit()


# Fallback office-start used only when the Setting rows are missing.
# Source of truth is attendance_settings_service.get_office_hours(db)
# — that reads from the `setting` table (configurable from the UI).
WORK_START_HOUR = 9
WORK_START_MINUTE = 20


# =========================
# HELPERS
# =========================

def compute_status(check_in_time: datetime) -> str:

    if not check_in_time:

        return "PRESENT"

    # Policy: 9:20 AM is on-time (PRESENT). Only punches from 9:21 AM
    # onwards are LATE. `late_cutoff` = start + 1 minute; anyone at or
    # after that instant is late.
    from datetime import timedelta
    start_dt = check_in_time.replace(
        hour=WORK_START_HOUR, minute=WORK_START_MINUTE,
        second=0, microsecond=0,
    )
    late_cutoff = start_dt + timedelta(minutes=1)

    return "LATE" if check_in_time >= late_cutoff else "PRESENT"


# =========================
# CHECK IN
# =========================

@router.post("/check-in")
def check_in(
    data: CheckInRequest,
    request: Request,
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):

    # An employee can only check IN as themselves; admins (e.g. kiosk
    # operators or HR via the live floor board) may check in anyone.
    assert_self_or_admin(data.EMPLOYEE_ID, payload)

    # Normalise either UUID or EMPLOYEE_CODE (e.g. "EMP101") to the
    # canonical UUID. The /employee-login flow returns the CODE under
    # the EMPLOYEE_ID key, so self-service callers need this bridge.
    # Raises 404 here if the employee genuinely doesn't exist.
    data.EMPLOYEE_ID = resolve_employee_uuid(db, data.EMPLOYEE_ID)

    emp = db.query(Employee).filter(
        Employee.ID == data.EMPLOYEE_ID
    ).first()

    if not emp:

        raise HTTPException(
            status_code=404,
            detail="Employee not found"
        )

    ip = request.client.host if request.client else None

    # ---- Geofence gate (server-side re-validation) ----
    geo = _check_geofence(db, data.VENDOR_ID, data.LATITUDE, data.LONGITUDE)

    if not geo["allowed"] and not data.BYPASS_GEOFENCE:

        _log_failure(
            db, data.EMPLOYEE_ID, "OUTSIDE_GEOFENCE",
            data.LATITUDE, data.LONGITUDE, geo["distance_m"],
            f"Check-in blocked: {geo['distance_m']}m from office (max {geo['settings'].RADIUS_METERS}m)",
            data.DEVICE_INFO, ip, data.VENDOR_ID
        )

        raise HTTPException(
            status_code=403,
            detail=(
                f"Outside office geofence — you are {round(geo['distance_m'])}m "
                f"from {geo['settings'].OFFICE_NAME or 'the office'} "
                f"(max allowed {geo['settings'].RADIUS_METERS}m). "
                f"Move closer and try again."
            )
        )

    # If the caller explicitly bypassed the gate, log the override
    # for audit + tag the status so HR can filter on it later.
    if not geo["allowed"] and data.BYPASS_GEOFENCE:
        _log_failure(
            db, data.EMPLOYEE_ID, "GEOFENCE_BYPASSED",
            data.LATITUDE, data.LONGITUDE, geo["distance_m"],
            "Admin/employee bypassed geofence gate at check-in.",
            data.DEVICE_INFO, ip, data.VENDOR_ID
        )
        geo["status"] = "BYPASSED"

    today = date.today()

    record = db.query(Attendance).filter(
        Attendance.EMPLOYEE_ID == data.EMPLOYEE_ID,
        Attendance.DATE == today
    ).first()

    now = datetime.now()

    if record:

        if record.CHECK_IN:

            raise HTTPException(
                status_code=400,
                detail="Employee has already checked in today"
            )

        record.CHECK_IN = now

        record.STATUS = compute_status(now)

    else:

        record = Attendance(
            EMPLOYEE_ID=data.EMPLOYEE_ID,
            DATE=today,
            CHECK_IN=now,
            STATUS=compute_status(now),
            VENDOR_ID=data.VENDOR_ID
        )

        db.add(record)

    # ---- Persist geo + device info ----
    record.CHECKIN_LATITUDE  = data.LATITUDE
    record.CHECKIN_LONGITUDE = data.LONGITUDE
    record.CHECKIN_DISTANCE  = geo["distance_m"]
    record.GEOFENCE_STATUS   = geo["status"]
    record.DEVICE_INFO       = (data.DEVICE_INFO or "")[:255] or record.DEVICE_INFO
    record.BROWSER_INFO      = (data.BROWSER_INFO or "")[:255] or record.BROWSER_INFO
    record.IP_ADDRESS        = ip or record.IP_ADDRESS

    db.commit()

    db.refresh(record)

    return {
        "message": "Checked in",
        "attendance_id": record.ID,
        "status": record.STATUS,
        "geofence_status": record.GEOFENCE_STATUS,
        "distance_meters": record.CHECKIN_DISTANCE
    }


# =========================
# CHECK OUT
# =========================

@router.post("/check-out")
def check_out(
    data: CheckOutRequest,
    request: Request,
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):

    # An employee can only check OUT as themselves; admins may
    # check out anyone (e.g. shift supervisor closing the floor).
    assert_self_or_admin(data.EMPLOYEE_ID, payload)

    # Accept either UUID or EMPLOYEE_CODE — see check-in route comment.
    data.EMPLOYEE_ID = resolve_employee_uuid(db, data.EMPLOYEE_ID)

    today = date.today()

    record = db.query(Attendance).filter(
        Attendance.EMPLOYEE_ID == data.EMPLOYEE_ID,
        Attendance.DATE == today
    ).first()

    if not record:

        raise HTTPException(
            status_code=404,
            detail="No check-in record for today"
        )

    if not record.CHECK_IN:

        raise HTTPException(
            status_code=400,
            detail="Employee has not checked in today"
        )

    if record.CHECK_OUT:

        raise HTTPException(
            status_code=400,
            detail="Employee has already checked out today"
        )

    ip = request.client.host if request.client else None

    # ---- Geofence gate on check-out as well ----
    vendor_id = record.VENDOR_ID or 1

    geo = _check_geofence(db, vendor_id, data.LATITUDE, data.LONGITUDE)

    if not geo["allowed"] and not data.BYPASS_GEOFENCE:

        _log_failure(
            db, data.EMPLOYEE_ID, "OUTSIDE_GEOFENCE",
            data.LATITUDE, data.LONGITUDE, geo["distance_m"],
            f"Check-out blocked: {geo['distance_m']}m from office",
            data.DEVICE_INFO, ip, vendor_id
        )

        raise HTTPException(
            status_code=403,
            detail=(
                f"Outside office geofence — you are {round(geo['distance_m'])}m "
                f"from the office (max {geo['settings'].RADIUS_METERS}m). "
                f"Move closer to check out."
            )
        )

    if not geo["allowed"] and data.BYPASS_GEOFENCE:
        _log_failure(
            db, data.EMPLOYEE_ID, "GEOFENCE_BYPASSED",
            data.LATITUDE, data.LONGITUDE, geo["distance_m"],
            "Admin/employee bypassed geofence gate at check-out.",
            data.DEVICE_INFO, ip, vendor_id
        )
        # Keep the existing GEOFENCE_STATUS on the record (set at check-in
        # time) — checkout-side bypass doesn't overwrite the morning's
        # geofence verdict. Distance below is still computed + stored.

    now = datetime.now()

    record.CHECK_OUT = now

    # Compute worked hours + overtime
    delta = now - record.CHECK_IN

    hours = round(delta.total_seconds() / 3600, 2)

    record.WORKED_HOURS = hours

    # OT is no longer auto-derived from regular hours. Employees must
    # explicitly start a separate OT session (POST /attendance/ot-check-in)
    # for any time after their regular check-out to count as overtime.

    # ---- Persist check-out geo ----
    record.CHECKOUT_LATITUDE  = data.LATITUDE
    record.CHECKOUT_LONGITUDE = data.LONGITUDE
    record.CHECKOUT_DISTANCE  = geo["distance_m"]

    db.commit()

    return {
        "message": "Checked out",
        "attendance_id": record.ID,
        "worked_hours": record.WORKED_HOURS,
        "overtime_hours": record.OVERTIME_HOURS,
        "checkout_distance_meters": record.CHECKOUT_DISTANCE
    }


# =========================
# OT CHECK-IN / CHECK-OUT
# =========================
# Overtime is a SECOND session within the same day. The employee
# completes their regular check-out first, then starts an OT session
# when they continue working. OT hours are computed strictly from
# (OT_CHECK_OUT - OT_CHECK_IN).

class _OtRequest(BaseModel):
    EMPLOYEE_ID: str


@router.post("/ot-check-in")
def ot_check_in(
    data: _OtRequest,
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Start the OT session for today. Requires regular check-out to
    be done first. Idempotent — re-calling while OT is in progress
    returns the existing OT_CHECK_IN."""

    assert_self_or_admin(data.EMPLOYEE_ID, payload)
    data.EMPLOYEE_ID = resolve_employee_uuid(db, data.EMPLOYEE_ID)

    today = date.today()
    record = db.query(Attendance).filter(
        Attendance.EMPLOYEE_ID == data.EMPLOYEE_ID,
        Attendance.DATE == today
    ).first()

    if not record:
        raise HTTPException(
            status_code=404,
            detail="No check-in record for today. Check in first."
        )

    if not record.CHECK_OUT:
        raise HTTPException(
            status_code=400,
            detail="Complete the regular check-out before starting OT."
        )

    if record.OT_CHECK_OUT:
        raise HTTPException(
            status_code=400,
            detail="OT session for today is already closed."
        )

    if record.OT_CHECK_IN:
        # Idempotent — return the existing timestamp
        return {
            "message": "OT already in progress",
            "ot_check_in": record.OT_CHECK_IN.isoformat()
        }

    # OT clock only starts from 7:00 PM. Employees who click the OT
    # button between 6:00 and 7:00 PM are still logged, but the OT
    # anchor is clamped to 7:00 PM so the 6-7 grace window doesn't
    # count as paid overtime.
    now_ = datetime.now()
    ot_floor = now_.replace(hour=19, minute=0, second=0, microsecond=0)
    record.OT_CHECK_IN = max(now_, ot_floor)
    db.commit()

    return {
        "message": "OT started",
        "attendance_id": record.ID,
        "ot_check_in": record.OT_CHECK_IN.isoformat()
    }


@router.post("/ot-check-out")
def ot_check_out(
    data: _OtRequest,
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Close the OT session — computes OVERTIME_HOURS from the delta."""

    assert_self_or_admin(data.EMPLOYEE_ID, payload)
    data.EMPLOYEE_ID = resolve_employee_uuid(db, data.EMPLOYEE_ID)

    today = date.today()
    record = db.query(Attendance).filter(
        Attendance.EMPLOYEE_ID == data.EMPLOYEE_ID,
        Attendance.DATE == today
    ).first()

    if not record or not record.OT_CHECK_IN:
        raise HTTPException(
            status_code=400,
            detail="No OT session in progress. Start OT first."
        )

    if record.OT_CHECK_OUT:
        raise HTTPException(
            status_code=400,
            detail="OT session for today is already closed."
        )

    now = datetime.now()
    record.OT_CHECK_OUT = now
    delta = now - record.OT_CHECK_IN
    record.OVERTIME_HOURS = max(0.0, round(delta.total_seconds() / 3600, 2))
    db.commit()

    return {
        "message": "OT closed",
        "attendance_id": record.ID,
        "ot_check_in":  record.OT_CHECK_IN.isoformat(),
        "ot_check_out": record.OT_CHECK_OUT.isoformat(),
        "overtime_hours": record.OVERTIME_HOURS
    }


# =========================
# MARK ABSENT
# =========================

@router.post("/mark-absent", dependencies=[Depends(require("attendance.mark.others"))])
def mark_absent(
    data: MarkAbsentRequest,
    db: Session = Depends(get_db)
):

    # Accept either UUID or EMPLOYEE_CODE — see check-in route comment.
    data.EMPLOYEE_ID = resolve_employee_uuid(db, data.EMPLOYEE_ID)

    today = date.today()

    existing = db.query(Attendance).filter(
        Attendance.EMPLOYEE_ID == data.EMPLOYEE_ID,
        Attendance.DATE == today
    ).first()

    if existing:

        existing.STATUS = "ABSENT"

        existing.CHECK_IN = None

        existing.CHECK_OUT = None

        existing.WORKED_HOURS = None

        existing.OVERTIME_HOURS = 0

        if data.NOTE:

            existing.REMARKS = data.NOTE

        db.commit()

        return {"message": "Marked absent"}

    record = Attendance(
        EMPLOYEE_ID=data.EMPLOYEE_ID,
        DATE=today,
        STATUS="ABSENT",
        REMARKS=data.NOTE,
        VENDOR_ID=data.VENDOR_ID
    )

    db.add(record)

    db.commit()

    return {"message": "Marked absent"}


# =========================
# LIST
# =========================

@router.get("/attendance", dependencies=[Depends(require("attendance.view.all"))])
def get_attendance(
    start_date:  Optional[date] = Query(None, description="Inclusive lower bound on DATE"),
    end_date:    Optional[date] = Query(None, description="Inclusive upper bound on DATE"),
    employee_id: Optional[str]  = Query(None, description="Employee UUID or EMPLOYEE_CODE"),
    status:      Optional[str]  = Query(None, description="PRESENT / LATE / ABSENT / HALF_DAY"),
    limit:       int            = Query(100, ge=1, le=1000),
    offset:      int            = Query(0,   ge=0),
    db: Session = Depends(get_db),
):
    """
    Filterable attendance history.

    Returns an envelope `{ total, rows }` so the UI can paginate. With no
    filters, returns the most-recent `limit` rows.
    """
    q = db.query(
        Attendance, Employee.NAME, Employee.EMPLOYEE_CODE
    ).outerjoin(
        Employee, Attendance.EMPLOYEE_ID == Employee.ID
    )

    if start_date:
        q = q.filter(Attendance.DATE >= start_date)
    if end_date:
        q = q.filter(Attendance.DATE <= end_date)
    if status:
        q = q.filter(Attendance.STATUS == status.upper().strip())
    if employee_id:
        # Accept either UUID or EMPLOYEE_CODE
        q = q.filter(
            (Attendance.EMPLOYEE_ID == employee_id) |
            (Employee.EMPLOYEE_CODE == employee_id)
        )

    total = q.count()

    rows = q.order_by(
        Attendance.DATE.desc(),
        Attendance.CHECK_IN.desc()
    ).offset(offset).limit(limit).all()

    out = []
    for record, name, code in rows:
        out.append({
            "ID": record.ID,
            "EMPLOYEE_ID": record.EMPLOYEE_ID,
            "EMPLOYEE_CODE": code,
            "EMPLOYEE_NAME": name,
            "DATE": record.DATE.isoformat() if record.DATE else None,
            "CHECK_IN":  record.CHECK_IN.isoformat()  if record.CHECK_IN  else None,
            "CHECK_OUT": record.CHECK_OUT.isoformat() if record.CHECK_OUT else None,
            "STATUS": record.STATUS,
            "WORKED_HOURS": record.WORKED_HOURS,
            "OVERTIME_HOURS": record.OVERTIME_HOURS,
            "OT_CHECK_IN":  record.OT_CHECK_IN.isoformat()  if record.OT_CHECK_IN  else None,
            "OT_CHECK_OUT": record.OT_CHECK_OUT.isoformat() if record.OT_CHECK_OUT else None,
            "REMARKS": record.REMARKS,
            "VENDOR_ID": record.VENDOR_ID,
            # ---- Geofence ----
            "CHECKIN_LATITUDE":   record.CHECKIN_LATITUDE,
            "CHECKIN_LONGITUDE":  record.CHECKIN_LONGITUDE,
            "CHECKIN_DISTANCE":   record.CHECKIN_DISTANCE,
            "CHECKOUT_LATITUDE":  record.CHECKOUT_LATITUDE,
            "CHECKOUT_LONGITUDE": record.CHECKOUT_LONGITUDE,
            "CHECKOUT_DISTANCE":  record.CHECKOUT_DISTANCE,
            "GEOFENCE_STATUS":    record.GEOFENCE_STATUS,
            "DEVICE_INFO":        record.DEVICE_INFO,
            "BROWSER_INFO":       record.BROWSER_INFO,
            "IP_ADDRESS":         record.IP_ADDRESS,
        })

    return {
        "total":  total,
        "limit":  limit,
        "offset": offset,
        "rows":   out,
    }


@router.get("/attendance/today")
def get_today_attendance(
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Today's attendance rows.

    - Admins see every employee's row.
    - Employees see only their own row (so the Employee Portal's
      attendance widget can render without a separate endpoint).
    """

    today = date.today()

    q = db.query(
        Attendance,
        Employee.NAME,
        Employee.EMPLOYEE_CODE
    ).outerjoin(
        Employee,
        Attendance.EMPLOYEE_ID == Employee.ID
    ).filter(
        Attendance.DATE == today
    )

    if payload.get("role") not in ADMIN_ROLES:
        # Scope to caller's own row only.
        q = q.filter(Attendance.EMPLOYEE_ID == payload.get("employee_id"))

    rows = q.all()

    return [
        {
            "ID": rec.ID,
            "EMPLOYEE_ID": rec.EMPLOYEE_ID,
            "EMPLOYEE_CODE": code,
            "EMPLOYEE_NAME": name,
            "DATE": today.isoformat(),
            "CHECK_IN": (
                rec.CHECK_IN.isoformat()
                if rec.CHECK_IN else None
            ),
            "CHECK_OUT": (
                rec.CHECK_OUT.isoformat()
                if rec.CHECK_OUT else None
            ),
            "STATUS": rec.STATUS,
            "WORKED_HOURS": rec.WORKED_HOURS,
            "OVERTIME_HOURS": rec.OVERTIME_HOURS,
            "OT_CHECK_IN":  rec.OT_CHECK_IN.isoformat()  if rec.OT_CHECK_IN  else None,
            "OT_CHECK_OUT": rec.OT_CHECK_OUT.isoformat() if rec.OT_CHECK_OUT else None,
            # ---- Geofence (must match /attendance shape so the Today
            # ----  tab renders the same columns as All Records)
            "CHECKIN_LATITUDE":   rec.CHECKIN_LATITUDE,
            "CHECKIN_LONGITUDE":  rec.CHECKIN_LONGITUDE,
            "CHECKIN_DISTANCE":   rec.CHECKIN_DISTANCE,
            "CHECKOUT_LATITUDE":  rec.CHECKOUT_LATITUDE,
            "CHECKOUT_LONGITUDE": rec.CHECKOUT_LONGITUDE,
            "CHECKOUT_DISTANCE":  rec.CHECKOUT_DISTANCE,
            "GEOFENCE_STATUS":    rec.GEOFENCE_STATUS,
            "DEVICE_INFO":        rec.DEVICE_INFO,
            "BROWSER_INFO":       rec.BROWSER_INFO,
            "IP_ADDRESS":         rec.IP_ADDRESS
        }
        for rec, name, code in rows
    ]


@router.get("/attendance/live-board", dependencies=[Depends(require("attendance.view.all"))])
def live_floor_board(
    db: Session = Depends(get_db)
):
    """
    Live shop-floor display: one tile per ACTIVE employee with
    today's CHECK_IN / CHECK_OUT, current task, status. Powers
    the "Floor Board" view in the Attendance page — refreshes
    every 10 seconds for a live wall display.
    """

    today = date.today()

    employees = (
        db.query(Employee, Department)
        .outerjoin(Department, Employee.DEPARTMENT_ID == Department.ID)
        .filter(Employee.STATUS == "ACTIVE")
        .order_by(Employee.EMPLOYEE_CODE)
        .all()
    )

    # Pre-load today's attendance keyed by employee
    attendance_rows = (
        db.query(Attendance)
        .filter(Attendance.DATE == today)
        .all()
    )

    att_map = {a.EMPLOYEE_ID: a for a in attendance_rows}

    # Pre-load active tasks
    active_tasks = (
        db.query(TaskAssignment)
        .filter(
            TaskAssignment.ASSIGNED_DATE == today,
            TaskAssignment.TASK_STATUS.in_(
                ["PENDING", "IN_PROGRESS"]
            )
        )
        .all()
    )

    task_map = {}

    for task in active_tasks:

        # last wins, but each employee really only has one
        task_map[task.EMPLOYEE_ID] = task

    # Completed-task counts today (matches MD Performance logic)
    completed_today = (
        db.query(
            TaskAssignment.EMPLOYEE_ID,
            func.count(TaskAssignment.TASK_ID)
        )
        .filter(
            TaskAssignment.ASSIGNED_DATE == today,
            TaskAssignment.TASK_STATUS.in_(["DONE", "COMPLETED"])
        )
        .group_by(TaskAssignment.EMPLOYEE_ID)
        .all()
    )

    completed_map = {emp_id: cnt for emp_id, cnt in completed_today}

    out = []

    for emp, dept in employees:

        att = att_map.get(emp.ID)

        current_task = task_map.get(emp.ID)

        out.append({
            "EMPLOYEE_ID": emp.ID,
            "EMPLOYEE_CODE": emp.EMPLOYEE_CODE,
            "NAME": emp.NAME,
            "DEPARTMENT": dept.NAME if dept else None,
            "DEPARTMENT_CODE": dept.DEPARTMENT_CODE if dept else None,
            "SKILLS": emp.SKILLS,
            "CHECK_IN": (
                att.CHECK_IN.isoformat()
                if (att and att.CHECK_IN) else None
            ),
            "CHECK_OUT": (
                att.CHECK_OUT.isoformat()
                if (att and att.CHECK_OUT) else None
            ),
            "STATUS": att.STATUS if att else "NOT_CHECKED_IN",
            "WORKED_HOURS": att.WORKED_HOURS if att else None,
            "OVERTIME_HOURS": (
                att.OVERTIME_HOURS if att else None
            ),
            "CURRENT_TASK_ID": (
                current_task.TASK_ID if current_task else None
            ),
            "CURRENT_TASK_NAME": (
                current_task.TASK_NAME if current_task else None
            ),
            "CURRENT_TASK_STATUS": (
                current_task.TASK_STATUS if current_task else None
            ),
            "CURRENT_PROJECT": None,  # project_legacy FK removed
            "TASKS_COMPLETED_TODAY": completed_map.get(emp.ID, 0)
        })

    # Sort: checked-in first, then by check-in time desc
    out.sort(
        key=lambda r: (
            0 if r["CHECK_IN"] else 1,
            -(
                int(r["CHECK_IN"].replace(":", "").replace("-", "")
                    .replace("T", "")[8:14])
                if r["CHECK_IN"] else 0
            )
        )
    )

    # Summary tile data
    total = len(out)

    in_office = sum(
        1 for r in out
        if r["CHECK_IN"] and not r["CHECK_OUT"]
    )

    checked_out = sum(1 for r in out if r["CHECK_OUT"])

    not_in = total - in_office - checked_out

    return {
        "summary": {
            "total_active": total,
            "in_office": in_office,
            "checked_out": checked_out,
            "not_checked_in": not_in
        },
        "employees": out,
        "as_of": datetime.now().isoformat()
    }


@router.delete("/attendance/{attendance_id}", dependencies=[Depends(require("attendance.delete"))])
def delete_attendance(
    attendance_id: int,
    db: Session = Depends(get_db)
):

    record = db.query(Attendance).filter(
        Attendance.ID == attendance_id
    ).first()

    if not record:

        raise HTTPException(
            status_code=404,
            detail="Attendance record not found"
        )

    db.delete(record)

    db.commit()

    return {"message": "Attendance deleted"}


# =====================================================================
# ATTENDANCE REPORT — per-employee aggregates over a date range
# =====================================================================

@router.get("/attendance/report",
            dependencies=[Depends(require("attendance.view.all"))])
def attendance_report(
    start_date: date = Query(..., description="Inclusive lower bound"),
    end_date:   date = Query(..., description="Inclusive upper bound"),
    db: Session = Depends(get_db),
):
    """
    Returns per-employee summary across the date range:
      - working_days (date range total)
      - present, absent, late, half_day counts
      - worked_hours, overtime_hours
      - attendance_pct
    Plus a totals block for the whole company.
    """
    if end_date < start_date:
        raise HTTPException(400, "end_date must be >= start_date")

    # Pre-load ACTIVE employees so we surface zeros for those with no
    # attendance rows in the range.
    employees = (db.query(Employee)
                 .filter(Employee.STATUS == "ACTIVE")
                 .order_by(Employee.NAME.asc()).all())

    # Aggregate attendance per employee in one query
    from collections import defaultdict
    buckets = defaultdict(lambda: {
        "present": 0, "absent": 0, "late": 0, "half_day": 0,
        "worked_hours": 0.0, "overtime_hours": 0.0,
    })
    rows = (db.query(Attendance)
            .filter(Attendance.DATE >= start_date,
                    Attendance.DATE <= end_date).all())
    for r in rows:
        b = buckets[r.EMPLOYEE_ID]
        s = (r.STATUS or "").upper()
        if   s == "PRESENT":  b["present"]  += 1
        elif s == "LATE":     b["late"]     += 1; b["present"] += 1
        elif s == "ABSENT":   b["absent"]   += 1
        elif s == "HALF_DAY": b["half_day"] += 1; b["present"] += 0.5
        b["worked_hours"]   += float(r.WORKED_HOURS or 0)
        b["overtime_hours"] += float(r.OVERTIME_HOURS or 0)

    # Working-day count = total days in range minus Sundays
    working_days = 0
    d = start_date
    while d <= end_date:
        if d.weekday() != 6:   # 6 = Sunday
            working_days += 1
        d = date.fromordinal(d.toordinal() + 1)

    items = []
    for emp in employees:
        b = buckets[emp.ID]
        present = b["present"]
        attendance_pct = (
            round(present / working_days * 100, 1)
            if working_days else 0.0
        )
        items.append({
            "employee_id":   emp.ID,
            "employee_code": emp.EMPLOYEE_CODE,
            "employee_name": emp.NAME,
            "working_days":  working_days,
            "present":       present,
            "absent":        b["absent"],
            "late":          b["late"],
            "half_day":      b["half_day"],
            "worked_hours":  round(b["worked_hours"], 1),
            "overtime_hours":round(b["overtime_hours"], 1),
            "attendance_pct": attendance_pct,
        })

    # Sort by attendance_pct desc so top performers show first
    items.sort(key=lambda r: (-r["attendance_pct"], r["employee_name"]))

    totals = {
        "employees":       len(items),
        "working_days":    working_days,
        "total_present":   sum(i["present"]  for i in items),
        "total_absent":    sum(i["absent"]   for i in items),
        "total_late":      sum(i["late"]     for i in items),
        "total_hours":     round(sum(i["worked_hours"]   for i in items), 1),
        "total_overtime":  round(sum(i["overtime_hours"] for i in items), 1),
        "avg_attendance_pct": (
            round(sum(i["attendance_pct"] for i in items) / len(items), 1)
            if items else 0.0
        ),
    }

    return {
        "start_date": start_date.isoformat(),
        "end_date":   end_date.isoformat(),
        "totals":     totals,
        "rows":       items,
    }


# =====================================================================
# EMPLOYEE TRACKING — per-employee daily attendance over last N days
# =====================================================================

@router.get("/attendance/employee/{employee_id}/tracking")
def employee_attendance_tracking(
    employee_id: str,
    days: int = Query(90, ge=7, le=365),
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    """Per-employee tracking view used by the Attendance > Tracking tab.

      - summary KPIs (present, absent, late, hours, attendance %)
      - daily timeline (date → status) for the last N days, used by the
        UI to render a calendar heatmap
    """
    from datetime import timedelta as _td

    emp = (db.query(Employee)
           .filter((Employee.ID == employee_id) |
                   (Employee.EMPLOYEE_CODE == employee_id)).first())
    if not emp:
        raise HTTPException(404, "Employee not found")

    end_d   = date.today()
    start_d = end_d - _td(days=days - 1)

    rows = (db.query(Attendance)
            .filter(Attendance.EMPLOYEE_ID == emp.ID,
                    Attendance.DATE >= start_d,
                    Attendance.DATE <= end_d)
            .order_by(Attendance.DATE.asc()).all())

    by_date = {r.DATE: r for r in rows}

    timeline = []
    present, absent, late, half = 0, 0, 0, 0
    worked_hours, ot_hours = 0.0, 0.0

    d = start_d
    while d <= end_d:
        rec = by_date.get(d)
        if rec:
            s = (rec.STATUS or "").upper()
            if s == "PRESENT":  present += 1
            elif s == "LATE":   late    += 1; present += 1
            elif s == "ABSENT": absent  += 1
            elif s == "HALF_DAY": half += 1; present += 0.5
            worked_hours += float(rec.WORKED_HOURS or 0)
            ot_hours     += float(rec.OVERTIME_HOURS or 0)
            check_in_str = (
                rec.CHECK_IN.strftime("%H:%M") if rec.CHECK_IN else None
            )
            check_out_str = (
                rec.CHECK_OUT.strftime("%H:%M") if rec.CHECK_OUT else None
            )
            timeline.append({
                "date":          d.isoformat(),
                "weekday":       d.strftime("%a"),
                "status":        s,
                "check_in":      check_in_str,
                "check_out":     check_out_str,
                "worked_hours":  float(rec.WORKED_HOURS or 0),
                "overtime_hours":float(rec.OVERTIME_HOURS or 0),
            })
        else:
            # No row for that day — treat Sundays as WEEKLY_OFF, else NO_DATA
            timeline.append({
                "date":     d.isoformat(),
                "weekday":  d.strftime("%a"),
                "status":   "WEEKLY_OFF" if d.weekday() == 6 else "NO_DATA",
                "check_in": None, "check_out": None,
                "worked_hours": 0, "overtime_hours": 0,
            })
        d = date.fromordinal(d.toordinal() + 1)

    # Working days = days in range minus Sundays
    working_days = sum(1 for t in timeline if t["status"] != "WEEKLY_OFF")
    attendance_pct = round(present / working_days * 100, 1) if working_days else 0.0

    return {
        "employee": {
            "id":   emp.ID,
            "code": emp.EMPLOYEE_CODE,
            "name": emp.NAME,
            "department_id":  emp.DEPARTMENT_ID,
            "designation_id": emp.DESIGNATION_ID,
        },
        "window": {
            "start_date":   start_d.isoformat(),
            "end_date":     end_d.isoformat(),
            "days":         days,
            "working_days": working_days,
        },
        "summary": {
            "present":         present,
            "absent":          absent,
            "late":            late,
            "half_day":        half,
            "worked_hours":    round(worked_hours, 1),
            "overtime_hours":  round(ot_hours, 1),
            "attendance_pct":  attendance_pct,
        },
        "timeline": timeline,
    }


# =====================================================================
# GET /attendance/export.xlsx
#
# Downloads a month's Attendance rows as an .xlsx file. Columns include
# both edges of the day, worked/OT hours, status, late-by, and the
# source (biometric device vs. web check-in). Optional filters:
#
#   month=YYYY-MM       (required) — e.g. 2026-08
#   employee_id=<uuid>  optional   — restrict to one employee
#   department_id=<int> optional   — restrict to one department
#
# Access is gated the same way as the /attendance list view (view.all).
# =====================================================================
from io import BytesIO


@router.get(
    "/attendance/download/xlsx",
    dependencies=[Depends(require("attendance.view.all"))],
)
def export_attendance_excel(
    month: str = Query(..., description="YYYY-MM, e.g. 2026-08"),
    employee_id: Optional[str] = Query(None),
    employee_ids: Optional[str] = Query(
        None,
        description="Comma-separated employee UUIDs. Wins over "
                    "employee_id when both are set.",
    ),
    department_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """Stream an .xlsx of the requested month's attendance rows."""

    # Lazy imports so the module stays cheap when the endpoint isn't hit.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    # ---- Parse month + compute range -------------------------------
    try:
        year_str, mon_str = month.split("-")
        year, mon = int(year_str), int(mon_str)
        start_date = date(year, mon, 1)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    # Last day of the month, computed without dateutil
    if mon == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, mon + 1, 1)

    # ---- Query attendance rows for the range -----------------------
    q = (
        db.query(Attendance, Employee)
          .join(Employee, Employee.ID == Attendance.EMPLOYEE_ID)
          .filter(Attendance.DATE >= start_date, Attendance.DATE < end_date)
          # Don't surface auto-stamped ABSENT rows from before the
          # employee had actually joined. Prior to this filter, someone
          # like Deepthi R (joined 10-08-2026) was showing ABSENT for
          # 01..09 August in the export because the month-close job
          # back-filled every calendar date. Rows whose JOINING_DATE
          # is null (legacy data) fall through unfiltered so we don't
          # accidentally hide anything real.
          .filter(
              (Employee.JOINING_DATE.is_(None))
              | (Attendance.DATE >= Employee.JOINING_DATE)
          )
    )

    # Multi-select wins when provided. Falls back to the single-id
    # param so old callers still work.
    if employee_ids:
        ids = [i.strip() for i in employee_ids.split(",") if i.strip()]
        if ids:
            q = q.filter(Attendance.EMPLOYEE_ID.in_(ids))
    elif employee_id:
        q = q.filter(Attendance.EMPLOYEE_ID == employee_id)

    if department_id:
        q = q.filter(Employee.DEPARTMENT_ID == department_id)

    rows = q.order_by(Employee.EMPLOYEE_CODE, Attendance.DATE).all()

    # ---- Load holidays that fall in this month.
    #
    # Source: the Announcement table (Admin → Announcements → type
    # HOLIDAY). Whatever the admin publishes there is the single source
    # of truth for what counts as a company holiday. If no HOLIDAY
    # announcement exists for a date, that date is a regular working
    # day — even if it's a well-known public holiday.
    #
    # `holiday_by_date`  — every announced holiday in this month
    #                      (drives STATUS label + Remarks on daily rows).
    # `countable_dates`  — subset that rolls up into the per-employee
    #                      summary's "Holidays" column. India's three
    #                      mandatory national holidays (Aug 15,
    #                      Jan 26, Oct 2) are excluded from that
    #                      rollup — they still show as HOLIDAY on the
    #                      daily rows but don't inflate the summary.
    MANDATORY_NATIONAL_DATES = {
        (1, 26),   # Republic Day
        (8, 15),   # Independence Day
        (10, 2),   # Gandhi Jayanti
    }
    holiday_by_date: dict = {}
    countable_dates: set = set()
    try:
        from app.models.models import Announcement
        aq = db.query(Announcement).filter(
            Announcement.TYPE == "HOLIDAY",
            Announcement.IS_ACTIVE == 1,
            Announcement.EVENT_DATE >= start_date,
            Announcement.EVENT_DATE <  end_date,
        )
        for a in aq.all():
            if a.EVENT_DATE is not None:
                holiday_by_date[a.EVENT_DATE] = a.TITLE or "Holiday"
                key = (a.EVENT_DATE.month, a.EVENT_DATE.day)
                if key not in MANDATORY_NATIONAL_DATES:
                    countable_dates.add(a.EVENT_DATE)
    except Exception:
        holiday_by_date = {}
        countable_dates = set()

    # ---- Build the workbook ----------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {month}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    alt_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    headers = [
        "Employee Code",
        "Name",
        "Date",
        "Day",
        "Check In",
        "Check Out",
        "OT Check In",
        "OT Check Out",
        "Worked Hours",
        "OT Hours",
        "Status",
        "Late By (min)",
        "Source",
        "Remarks",
    ]

    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ---- Row-by-row data -------------------------------------------
    WORK_START = time(9, 20)  # match attendance.py's cutoff

    def fmt_time(dt):
        return dt.strftime("%H:%M") if dt else ""

    def compute_late_by(check_in):
        """Minutes late past 09:20. 0 if on-time or missing."""
        if not check_in:
            return 0
        cutoff = datetime.combine(check_in.date(), WORK_START)
        delta = (check_in - cutoff).total_seconds() / 60.0
        return max(0, round(delta))

    def hours_to_hhmm(dec_hours):
        """Convert decimal hours (e.g. 4.92) to 'H:MM' string ('4:55').

        Empty / zero / None → '0:00' so the column reads cleanly for
        absent or non-OT days. Uses floor (truncates the seconds) so
        the HH:MM matches the OT Check In / Check Out columns, which
        are already displayed with seconds stripped (%H:%M).
        """
        try:
            v = float(dec_hours or 0)
        except (TypeError, ValueError):
            v = 0.0
        if v <= 0:
            return "0:00"
        total_min = int(v * 60)  # floor — matches visible HH:MM punches
        h, m = divmod(total_min, 60)
        return f"{h}:{m:02d}"

    # Per-employee accumulators — written as a summary block at the
    # bottom of the sheet so the salary team can read the month
    # totals for every employee without summing daily rows by hand.
    emp_totals: dict = {}

    for i, (att, emp) in enumerate(rows, start=2):

        # If this DATE is on the company holiday calendar and the
        # employee didn't punch in, replace the raw STATUS (usually
        # ABSENT, stamped by the month-close job) with HOLIDAY. Also
        # surface the holiday name in the Remarks column so the export
        # is self-explanatory.
        holiday_name = holiday_by_date.get(att.DATE)
        if holiday_name and att.CHECK_IN is None:
            status_val = "HOLIDAY"
            remarks_val = holiday_name
        else:
            status_val = att.STATUS or ""
            remarks_val = att.REMARKS or ""

        # OT columns must only show a value on days that genuinely
        # earned OT — i.e. CHECK_OUT was past 19:00 and the row's
        # OVERTIME_HOURS is > 0. Legacy rows written before the 19:00
        # OT floor rule sometimes carry stale OT_CHECK_OUT = CHECK_OUT
        # even though no OT was earned. Hiding them here keeps the
        # export clean without needing a data-migration.
        has_ot = (
            (att.OVERTIME_HOURS or 0) > 0
            and att.CHECK_OUT is not None
            and att.CHECK_OUT.time() > time(19, 0)
        )
        ot_in_display  = fmt_time(att.OT_CHECK_IN)  if has_ot else ""
        ot_out_display = fmt_time(att.OT_CHECK_OUT) if has_ot else ""

        row_data = [
            emp.EMPLOYEE_CODE or "",
            emp.NAME or "",
            att.DATE.strftime("%d-%m-%Y") if att.DATE else "",
            att.DATE.strftime("%A") if att.DATE else "",
            fmt_time(att.CHECK_IN),
            fmt_time(att.CHECK_OUT),
            ot_in_display,
            ot_out_display,
            round(att.WORKED_HOURS or 0, 2),
            hours_to_hhmm(att.OVERTIME_HOURS) if has_ot else "0:00",
            status_val,
            compute_late_by(att.CHECK_IN),
            att.DEVICE_INFO or "",
            remarks_val,
        ]

        ws.append(row_data)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill

        # Update this employee's month totals for the summary block.
        # Late Minutes uses the same 09:20 cutoff as the daily "Late By"
        # column so both figures reconcile.
        key = emp.EMPLOYEE_CODE or emp.ID
        bucket = emp_totals.setdefault(key, {
            "code":            emp.EMPLOYEE_CODE or "",
            "name":            emp.NAME or "",
            "employee":        emp,        # kept for post-loop holiday counting
            "physical_days":   0,          # rows with CHECK_IN (paid via punch)
            "punched_dates":   set(),      # dates the employee actually punched
            "days_present":    0,          # filled in post-loop
            "days_late":       0,
            "days_absent":     0,
            "days_holiday":    0,          # filled in post-loop
            "worked_hours":    0.0,
            "ot_hours":        0.0,
            "late_minutes":    0,
        })

        if att.CHECK_IN is not None:
            bucket["physical_days"] += 1
            bucket["punched_dates"].add(att.DATE)
            if (att.STATUS or "").upper() == "LATE":
                bucket["days_late"] += 1
        elif (att.STATUS or "").upper() == "ABSENT" and not holiday_by_date.get(att.DATE):
            # Real absent: not on a holiday, and no punch. This is a
            # money-lost day (unless covered by CL in payroll math).
            bucket["days_absent"] += 1

        bucket["worked_hours"] += float(att.WORKED_HOURS or 0)
        if has_ot:
            bucket["ot_hours"] += float(att.OVERTIME_HOURS or 0)
        bucket["late_minutes"] += compute_late_by(att.CHECK_IN)

    # If no rows, drop a friendly note in row 2 so the file isn't empty.
    if not rows:
        ws.cell(row=2, column=1, value=f"No attendance rows for {month}.")

    # ---- Month totals block per employee ---------------------------
    # Written after the daily rows so the salary team can consume the
    # month at a glance without pivoting. Sorted by employee code.
    if emp_totals:
        # ---- Recompute Days Present + Holidays from the employee's
        # active date range, not per-row. This makes both columns
        # correct even when the Attendance table is missing rows for
        # some holiday dates (e.g. month-close never wrote an Aug 26
        # row for a specific employee).
        #
        # Days Present = physical punches + holidays in range NOT
        #                already covered by a punch. That matches the
        #                payslip's "Paid Days" (days the employee gets
        #                paid for).
        # Holidays     = non-mandatory-national holidays in the range,
        #                not already covered by a punch (subset for
        #                informational rollup; overlaps with Present).
        from datetime import timedelta
        month_last_day = end_date - timedelta(days=1)

        for bucket in emp_totals.values():
            emp = bucket.get("employee")
            if emp is None:
                continue
            join_date = getattr(emp, "JOINING_DATE", None) or start_date
            range_start = max(start_date, join_date)
            range_end   = month_last_day
            punched     = bucket.get("punched_dates", set())

            paid_holiday_credit = 0
            countable_holiday_credit = 0
            for d in holiday_by_date.keys():
                if range_start <= d <= range_end and d not in punched:
                    paid_holiday_credit += 1
                    if d in countable_dates:
                        countable_holiday_credit += 1

            bucket["days_present"] = bucket["physical_days"] + paid_holiday_credit
            bucket["days_holiday"] = countable_holiday_credit

        totals_sorted = sorted(emp_totals.values(), key=lambda x: x["code"])

        # Spacer row
        blank_row_idx = ws.max_row + 2

        # Section title spanning columns A..N so it reads as a header.
        title_cell = ws.cell(row=blank_row_idx, column=1,
                             value=f"MONTH TOTALS PER EMPLOYEE — {month}")
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=blank_row_idx, start_column=1,
            end_row=blank_row_idx,   end_column=len(headers),
        )

        # Column headers for the summary. Holidays column removed
        # per admin request — Days Present already accounts for paid
        # holidays (see the post-loop counter), so a separate rollup
        # column was redundant.
        summary_headers = [
            "Employee Code",
            "Name",
            "Days Present",
            "Late Arrivals",
            "Days Absent",
            "Total Worked Hours",
            "Total OT Hours",
            "Total Late Minutes",
        ]
        header_row_idx = blank_row_idx + 1
        ws.append(summary_headers)
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # One data row per employee
        for row_offset, t in enumerate(totals_sorted, start=header_row_idx + 1):
            ws.append([
                t["code"],
                t["name"],
                t["days_present"],
                t["days_late"],
                t["days_absent"],
                round(t["worked_hours"], 2),
                hours_to_hhmm(t["ot_hours"]),
                int(t["late_minutes"]),
            ])
            for col_idx in range(1, len(summary_headers) + 1):
                cell = ws.cell(row=row_offset, column=col_idx)
                cell.border = thin_border
                if row_offset % 2 == 0:
                    cell.fill = alt_fill

        # Overall grand-total row across all employees
        grand_row_idx = ws.max_row + 1
        gt_worked = sum(t["worked_hours"] for t in totals_sorted)
        gt_ot     = sum(t["ot_hours"]     for t in totals_sorted)
        gt_late   = sum(t["late_minutes"] for t in totals_sorted)
        gt_days_p = sum(t["days_present"] for t in totals_sorted)
        gt_days_a = sum(t["days_absent"]  for t in totals_sorted)
        gt_days_l = sum(t["days_late"]    for t in totals_sorted)

        ws.append([
            "GRAND TOTAL",
            f"{len(totals_sorted)} employees",
            gt_days_p,
            gt_days_l,
            gt_days_a,
            round(gt_worked, 2),
            hours_to_hhmm(gt_ot),
            gt_late,
        ])
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws.cell(row=grand_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2",
                                    fill_type="solid")
            cell.border = thin_border

    # ---- Column widths ---------------------------------------------
    widths = [15, 22, 12, 11, 11, 11, 12, 13, 13, 10, 12, 13, 15, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ---- Stream as download ----------------------------------------
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance-{month}.xlsx"
    if employee_id:
        filename = f"attendance-{month}-{employee_id[:8]}.xlsx"

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =====================================================================
# GET /attendance/monthly-summary
#
# HR admin: returns one row per employee for the requested month, with
# every metric the Monthly Summary tab renders (present/late/absent
# counts, late minutes, OT hours, memo eligibility, star-score
# breakdown, etc.). Gated by attendance.view.all so employees can't
# read the whole company.
# =====================================================================
@router.get(
    "/attendance/summary/monthly",
    dependencies=[Depends(require("attendance.view.all"))],
)
def monthly_summary_all(
    month: str = Query(..., description="YYYY-MM"),
    department_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    from app.services.monthly_attendance import compute_monthly_summary

    try:
        year_str, mon_str = month.split("-")
        year, mon = int(year_str), int(mon_str)
        # Validate the month is a real calendar month
        date(year, mon, 1)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    q = db.query(Employee).filter(Employee.STATUS == "ACTIVE")
    if department_id:
        q = q.filter(Employee.DEPARTMENT_ID == department_id)

    employees = q.order_by(Employee.EMPLOYEE_CODE).all()

    rows = [
        compute_monthly_summary(db, emp, year, mon, include_days=False)
        for emp in employees
    ]

    # Aggregate roll-up so the HR view can show company-wide numbers
    # (e.g. "18 late arrivals across the org this month") without a
    # second call.
    totals = {
        "employees": len(rows),
        "days_present": sum(r["days_present"] for r in rows),
        "days_absent": sum(r["days_absent"] for r in rows),
        "unpaid_absences": sum(r["unpaid_absences"] for r in rows),
        "late_arrivals": sum(r["late_arrivals"] for r in rows),
        "missed_checkouts": sum(r["missed_checkouts"] for r in rows),
        "total_ot_hours": round(sum(r["total_ot_hours"] for r in rows), 2),
        "will_get_warning": sum(1 for r in rows if r["memo_flags"]["will_get_warning"]),
        "will_get_appreciation": sum(1 for r in rows if r["memo_flags"]["will_get_appreciation"]),
    }

    return {
        "month": f"{year}-{mon:02d}",
        "month_label": date(year, mon, 1).strftime("%B %Y"),
        "totals": totals,
        "employees": rows,
    }


# =====================================================================
# GET /attendance/summary/my
#
# Employee self-service. Returns the same summary but for the caller
# only, plus a `days` array so the portal can render a calendar grid.
# No admin role needed — the endpoint identifies the caller from the
# JWT and returns their own data only. Employees passing someone
# else's employee_id get a 403.
# =====================================================================
@router.get("/attendance/summary/my")
def monthly_summary_mine(
    month: str = Query(..., description="YYYY-MM"),
    employee_id: Optional[str] = Query(
        None,
        description="Optional — admin can pass another employee's ID"
    ),
    db: Session = Depends(get_db),
    payload: dict = Depends(get_current_user),
):
    from app.services.monthly_attendance import compute_monthly_summary

    try:
        year_str, mon_str = month.split("-")
        year, mon = int(year_str), int(mon_str)
        date(year, mon, 1)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    # Resolve who we're querying. Employees can only fetch themselves.
    caller_id = payload.get("sub") or payload.get("employee_id") or ""
    role = (payload.get("role") or "").upper()

    target_id = employee_id or caller_id

    # Allow admins/HR to pass another employee's id; block regular staff
    if employee_id and employee_id != caller_id and role not in ADMIN_ROLES:
        raise HTTPException(
            status_code=403,
            detail="You can only view your own attendance summary."
        )

    # Normalise UUID vs EMPLOYEE_CODE
    target_id = resolve_employee_uuid(db, target_id)

    emp = db.query(Employee).filter(Employee.ID == target_id).first()

    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    return compute_monthly_summary(db, emp, year, mon, include_days=True)
