"""
Inventory Items — product stock-threshold management.

InventoryStock now sits directly under ProductMaster (one row per
(VENDOR_ID, PRODUCT_ID) — the old InventoryItem location-scoped
intermediate table has been removed, see inventory_models.py's NOTE
block). Every stock-level change is logged in InventoryMovement via
inventory_automation_service.record_movement().
"""

import io
from typing import Optional, List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from sqlalchemy import exists
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.utils.db_error_handler import raise_db_error

from app.database.database import get_db
from app.models.models import CustomField, CustomFieldTableValue
from app.models.inventory_models import ProductMaster, InventoryStock
from app.models.supplier_models import SupplierProduct
from app.schemas.inventory_item_schema import (
    StockThresholdCreate,
    StockThresholdUpdate,
    StockMovementRequest,
    BatchCreate,
    BatchUpdate,
)
from app.services.inventory_automation_service import record_movement
from app.auth.auth_bearer import require

router = APIRouter(prefix="/inventory-items", tags=["Inventory Items"])


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def _cf_fields_for_table(table_name: str, vendor_id: int, db: Session):
    return (
        db.query(CustomField)
        .filter(CustomField.TABLE_NAME == table_name, CustomField.VENDOR_ID == vendor_id)
        .order_by(CustomField.SORT_ORDER, CustomField.FIELD_NAME)
        .all()
    )


def _upsert_cf_bulk(row_id: str, table_name: str, cf_field_id: str, value, db: Session):
    existing = (
        db.query(CustomFieldTableValue)
        .filter(
            CustomFieldTableValue.TABLE_NAME == table_name,
            CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
            CustomFieldTableValue.CUSTOM_FIELD_ID == cf_field_id,
        )
        .first()
    )
    if existing:
        existing.CUSTOM_FIELD_VALUE = value if value else None
    elif value is not None:
        db.add(CustomFieldTableValue(
            TABLE_NAME=table_name,
            TABLE_ROW_ID=str(row_id),
            CUSTOM_FIELD_ID=cf_field_id,
            CUSTOM_FIELD_VALUE=value,
        ))


def _serialize_stock(stock: InventoryStock, product: Optional[ProductMaster] = None) -> dict:
    product = product or stock.product
    return {
        "ID": stock.ID,
        "VENDOR_ID": stock.VENDOR_ID,
        "PRODUCT_ID": stock.PRODUCT_ID,
        "PRODUCT_CODE": product.PRODUCT_CODE if product else None,
        "PRODUCT_NAME": product.PRODUCT_NAME if product else None,
        "UNIT": product.UNIT if product else None,
        "CATEGORY_ID": product.CATEGORY_ID if product else None,
        "MIN_QTY": stock.MIN_QTY,
        "MAX_QTY": stock.MAX_QTY,
        "CURRENT_QTY": stock.CURRENT_QTY,
        "STATUS": stock.STATUS,
        "UPDATED_AT": stock.UPDATED_AT.isoformat() if stock.UPDATED_AT else None,
    }


# ─────────────────────────────────────────────────────────────────────
# Stock-threshold CRUD (one row per product)
# ─────────────────────────────────────────────────────────────────────

@router.get("", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def list_items(
    vendor_id: int = Query(1),
    product_id: Optional[str] = Query(None),
    category_id: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    q = (
        db.query(InventoryStock)
        .join(ProductMaster, ProductMaster.ID == InventoryStock.PRODUCT_ID)
        .filter(InventoryStock.VENDOR_ID == vendor_id)
    )
    if product_id:
        q = q.filter(InventoryStock.PRODUCT_ID == product_id)
    if category_id:
        q = q.filter(ProductMaster.CATEGORY_ID == category_id)
    if supplier_id:
        # InventoryStock has no supplier column — a product's supplier(s)
        # live only in SupplierProduct, so filtering "stock for products
        # supplied by X" is an EXISTS semi-join through that relationship
        # rather than duplicating supplier data onto InventoryStock.
        q = q.filter(exists().where(
            SupplierProduct.VENDOR_ID == vendor_id,
            SupplierProduct.SUPPLIER_ID == supplier_id,
            SupplierProduct.PRODUCT_ID == InventoryStock.PRODUCT_ID,
        ))
    if search:
        term = f"%{search}%"
        q = q.filter(
            ProductMaster.PRODUCT_NAME.ilike(term) |
            ProductMaster.PRODUCT_CODE.ilike(term)
        )
    if status:
        q = q.filter(InventoryStock.STATUS == status.upper())

    total = q.count()
    rows = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [_serialize_stock(r) for r in rows],
    }


@router.post("", dependencies=[Depends(require("inventory.purchase", "inventory.items.create"))])
def create_item(payload: StockThresholdCreate, db: Session = Depends(get_db)):
    product = db.query(ProductMaster).filter(
        ProductMaster.ID == payload.PRODUCT_ID,
        ProductMaster.VENDOR_ID == payload.VENDOR_ID,
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    existing = db.query(InventoryStock).filter(
        InventoryStock.VENDOR_ID == payload.VENDOR_ID,
        InventoryStock.PRODUCT_ID == payload.PRODUCT_ID,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="This product is already tracked in inventory")

    try:
        stock = InventoryStock(
            VENDOR_ID=payload.VENDOR_ID,
            PRODUCT_ID=payload.PRODUCT_ID,
            MIN_QTY=payload.MIN_QTY or 0.0,
            MAX_QTY=payload.MAX_QTY,
            CURRENT_QTY=0.0,
            STATUS="OUT_OF_STOCK",
        )
        db.add(stock)
        db.commit()
        db.refresh(stock)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create inventory stock row")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create inventory stock row")
    return {"message": "Inventory stock row created", "ID": stock.ID}


@router.get("/low-stock", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def get_low_stock(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    rows = db.query(InventoryStock).filter(
        InventoryStock.VENDOR_ID == vendor_id,
        InventoryStock.STATUS == "LOW_STOCK",
    ).all()
    return [_serialize_stock(r) for r in rows]


@router.get("/out-of-stock", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def get_out_of_stock(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    rows = db.query(InventoryStock).filter(
        InventoryStock.VENDOR_ID == vendor_id,
        InventoryStock.STATUS == "OUT_OF_STOCK",
    ).all()
    return [_serialize_stock(r) for r in rows]


@router.get("/{item_id}", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def get_item(item_id: str, db: Session = Depends(get_db)):
    stock = db.query(InventoryStock).filter(InventoryStock.ID == item_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Inventory stock row not found")
    result = _serialize_stock(stock)
    result["custom_fields"] = [
        {"ID": f.ID, "FIELD_NAME": f.FIELD_NAME, "FIELD_TYPE": f.FIELD_TYPE,
         "IS_REQUIRED": f.IS_REQUIRED, "SORT_ORDER": f.SORT_ORDER, "OPTIONS": f.OPTIONS}
        for f in _cf_fields_for_table("inventory_stock", stock.VENDOR_ID, db)
    ]
    result["custom_field_values"] = [
        {"CUSTOM_FIELD_ID": v.CUSTOM_FIELD_ID, "VALUE": v.CUSTOM_FIELD_VALUE}
        for v in db.query(CustomFieldTableValue).filter(
            CustomFieldTableValue.TABLE_NAME == "inventory_stock",
            CustomFieldTableValue.TABLE_ROW_ID == str(item_id),
        ).all()
    ]
    return result


@router.put("/{item_id}", dependencies=[Depends(require("inventory.purchase", "inventory.items.update"))])
def update_item(item_id: str, payload: StockThresholdUpdate, db: Session = Depends(get_db)):
    stock = db.query(InventoryStock).filter(InventoryStock.ID == item_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Inventory stock row not found")
    for k, v in payload.dict(exclude_none=True).items():
        setattr(stock, k, v)
    from app.services.inventory_automation_service import _compute_status
    stock.STATUS = _compute_status(stock.CURRENT_QTY, stock.MIN_QTY, stock.MAX_QTY)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update inventory stock row")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update inventory stock row")
    return {"message": "Inventory thresholds updated"}


@router.delete("/{item_id}", dependencies=[Depends(require("inventory.purchase", "inventory.items.delete"))])
def delete_item(item_id: str, db: Session = Depends(get_db)):
    stock = db.query(InventoryStock).filter(InventoryStock.ID == item_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Inventory stock row not found")
    if stock.CURRENT_QTY and stock.CURRENT_QTY > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete a stock row with stock ({stock.CURRENT_QTY} units remaining). "
                   "Adjust stock to zero first."
        )
    try:
        db.delete(stock)
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete inventory stock row")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete inventory stock row")
    return {"message": "Inventory stock row deleted"}


@router.get("/{item_id}/stock", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def get_stock(item_id: str, db: Session = Depends(get_db)):
    stock = db.query(InventoryStock).filter(InventoryStock.ID == item_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="No stock record found")
    return _serialize_stock(stock)


# ─────────────────────────────────────────────────────────────────────
# Stock Operations — all go through record_movement(), keyed by PRODUCT_ID
# ─────────────────────────────────────────────────────────────────────

@router.post("/stock-in", dependencies=[Depends(require("inventory.purchase", "inventory.items.update"))])
def stock_in(payload: StockMovementRequest, db: Session = Depends(get_db)):
    """Receive stock into inventory (manual — GRN receiving uses its own
    path in purchase_order.py's _apply_grn_to_inventory, which also
    creates a proper InventoryBatch; this endpoint is for ad-hoc receipts
    with no PO/GRN behind them)."""
    movement = record_movement(
        db=db,
        vendor_id=payload.VENDOR_ID,
        product_id=payload.PRODUCT_ID,
        movement_type="STOCK_IN",
        qty=payload.QTY,
        performed_by_id=payload.PERFORMED_BY_ID,
        reference_type=payload.REFERENCE_TYPE,
        reference_id=payload.REFERENCE_ID,
        batch_id=payload.BATCH_ID,
        reason=payload.REASON,
        notes=payload.NOTES,
        unit_cost=payload.UNIT_COST,
    )
    db.commit()
    return {"message": "Stock in recorded", "MOVEMENT_ID": movement.ID, "QTY_AFTER": movement.QTY_AFTER}


@router.post("/stock-out", dependencies=[Depends(require("inventory.consume", "inventory.items.update"))])
def stock_out(payload: StockMovementRequest, db: Session = Depends(get_db)):
    """Issue stock from inventory."""
    movement = record_movement(
        db=db,
        vendor_id=payload.VENDOR_ID,
        product_id=payload.PRODUCT_ID,
        movement_type="STOCK_OUT",
        qty=payload.QTY,
        performed_by_id=payload.PERFORMED_BY_ID,
        reference_type=payload.REFERENCE_TYPE,
        reference_id=payload.REFERENCE_ID,
        batch_id=payload.BATCH_ID,
        reason=payload.REASON,
        notes=payload.NOTES,
    )
    db.commit()
    return {"message": "Stock out recorded", "MOVEMENT_ID": movement.ID, "QTY_AFTER": movement.QTY_AFTER}


@router.post("/stock-adjust", dependencies=[Depends(require("inventory.purchase", "inventory.items.update"))])
def stock_adjust(payload: StockMovementRequest, db: Session = Depends(get_db)):
    """Manual stock adjustment — sets qty to an absolute value. REASON is
    required here (enforced by the schema) — the audit trail (who/when/
    previous/new/difference) is InventoryMovement itself: QTY_BEFORE,
    QTY_AFTER, PERFORMED_BY_ID, CREATED_AT are all already recorded by
    record_movement(); the difference is QTY_AFTER - QTY_BEFORE."""
    movement = record_movement(
        db=db,
        vendor_id=payload.VENDOR_ID,
        product_id=payload.PRODUCT_ID,
        movement_type="ADJUSTMENT",
        qty=payload.QTY,
        performed_by_id=payload.PERFORMED_BY_ID,
        reason=payload.REASON,
        notes=payload.NOTES,
        unit_cost=payload.UNIT_COST,
    )
    db.commit()
    return {
        "message": "Stock adjusted",
        "MOVEMENT_ID": movement.ID,
        "QTY_BEFORE": movement.QTY_BEFORE,
        "QTY_AFTER": movement.QTY_AFTER,
        "DIFFERENCE": movement.QTY_AFTER - movement.QTY_BEFORE,
    }


# ─────────────────────────────────────────────────────────────────────
# Bulk Upload / Template / Export
# ─────────────────────────────────────────────────────────────────────

def _parse_bulk_xl(content: bytes, required_sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if required_sheet not in wb.sheetnames:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise HTTPException(
            status_code=400,
            detail=f'Sheet "{required_sheet}" not found. Available: {available}.',
        )
    ws = wb[required_sheet]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper() if c is not None else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(row)
    return headers, rows


def _cell(record: dict, *keys) -> str:
    for k in keys:
        v = record.get(k.upper())
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return ""


_ITEM_STD_COLS = {"PRODUCT CODE", "MIN QTY", "MAX QTY", "S.NO", "SN", ""}


@router.get("/bulk-template", dependencies=[Depends(require("inventory.view", "inventory.items.view"))])
def download_item_template(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InventoryItems"
    std_cols = ["PRODUCT CODE", "MIN QTY", "MAX QTY"]
    cf_fields = _cf_fields_for_table("inventory_stock", vendor_id, db)
    cf_cols = [f.FIELD_NAME for f in cf_fields]
    ws.append(std_cols + cf_cols)
    from openpyxl.styles import Font, PatternFill
    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_items_template.xlsx"},
    )


@router.post("/bulk-upload", dependencies=[Depends(require("inventory.purchase", "inventory.items.create", "inventory.items.import"))])
async def bulk_upload_items(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "InventoryItems")

    cf_fields = _cf_fields_for_table("inventory_stock", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols = [h for h in headers if h not in _ITEM_STD_COLS and h in cf_by_upper]

    products_by_code = {
        p.PRODUCT_CODE.upper(): p
        for p in db.query(ProductMaster).filter(ProductMaster.VENDOR_ID == vendor_id).all()
    }

    inserted = updated = skipped = 0
    errors: List[dict] = []

    def safe_float(val, default=0.0):
        try:
            return float(val) if val not in (None, "") else default
        except (ValueError, TypeError):
            return default

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}
        prod_code = _cell(record, "PRODUCT CODE")
        if not prod_code:
            errors.append({"row": row_num, "field": "PRODUCT CODE", "message": "Required"})
            continue

        product = products_by_code.get(prod_code.upper())
        if not product:
            errors.append({"row": row_num, "field": "PRODUCT CODE", "message": f"Product '{prod_code}' not found"})
            continue

        existing = db.query(InventoryStock).filter(
            InventoryStock.VENDOR_ID == vendor_id,
            InventoryStock.PRODUCT_ID == product.ID,
        ).first()

        min_qty = safe_float(_cell(record, "MIN QTY"), None)
        max_qty = safe_float(_cell(record, "MAX QTY"), None)

        if existing:
            if min_qty is not None:
                existing.MIN_QTY = min_qty
            if max_qty is not None:
                existing.MAX_QTY = max_qty
            for cf_id in cf_cols:
                _upsert_cf_bulk(existing.ID, "inventory_stock", cf_by_upper[cf_id].ID, record.get(cf_id), db)
            updated += 1
        else:
            stock = InventoryStock(
                VENDOR_ID=vendor_id, PRODUCT_ID=product.ID,
                MIN_QTY=min_qty or 0.0, MAX_QTY=max_qty,
                CURRENT_QTY=0.0, STATUS="OUT_OF_STOCK",
            )
            db.add(stock)
            db.flush()
            for col in cf_cols:
                _upsert_cf_bulk(stock.ID, "inventory_stock", cf_by_upper[col].ID, record.get(col), db)
            inserted += 1

    db.commit()
    return {
        "message": f"Upload: {inserted} inserted, {updated} updated, {skipped} skipped",
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": len(data_rows), "errors": errors,
    }


@router.get("/export/excel", dependencies=[Depends(require("inventory.view", "inventory.items.view", "inventory.items.export"))])
def export_items(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    rows = (
        db.query(InventoryStock)
        .join(ProductMaster, ProductMaster.ID == InventoryStock.PRODUCT_ID, isouter=True)
        .filter(InventoryStock.VENDOR_ID == vendor_id)
        .order_by(ProductMaster.PRODUCT_NAME)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InventoryItems"
    ws.append(["PRODUCT CODE", "PRODUCT NAME", "UNIT", "CURRENT QTY", "STATUS", "MIN QTY", "MAX QTY"])
    for r in rows:
        ws.append([
            r.product.PRODUCT_CODE if r.product else "",
            r.product.PRODUCT_NAME if r.product else "",
            r.product.UNIT if r.product else "",
            r.CURRENT_QTY, r.STATUS, r.MIN_QTY, r.MAX_QTY,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_items_export.xlsx"},
    )
