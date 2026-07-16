"""
Supplier master endpoints.

Models the companies BVC24 buys from. Mirrors the Employee
master pattern: full CRUD with code lookup, status filtering,
and category-based grouping for the purchase workflow.
"""

import io
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.database.database import get_db
from app.models.models import Supplier, Vendor, BOMItem
from app.models.supplier_models import (
    SupplierPerformanceMetrics,
    SupplierRanking,
    SupplierProduct as SupplierProductModel,
)
from app.models.inventory_models import ProductMaster, InventoryCategory
from app.schemas.supplier_schema import (
    SupplierCreate,
    SupplierUpdate
)


router = APIRouter(prefix="/suppliers", tags=["Suppliers"])


def _resolve_vendor_id(db: Session, requested: Optional[int]) -> int:

    if requested:

        has_data = (
            db.query(Supplier)
            .filter(Supplier.VENDOR_ID == requested)
            .first()
            is not None
        )

        if has_data:

            return requested

    bvc = db.query(Vendor).filter(
        Vendor.VENDOR_NAME == "Bharath Vending Corporation"
    ).first()

    if bvc:

        return bvc.ID

    any_v = db.query(Vendor).first()

    return any_v.ID if any_v else (requested or 1)


def _serialize_supplier(s: Supplier) -> dict:

    return {
        "ID": s.ID,
        "SUPPLIER_CODE": s.SUPPLIER_CODE,
        "COMPANY_NAME": s.COMPANY_NAME,
        "CONTACT_PERSON": s.CONTACT_PERSON,
        "PHONE": s.PHONE,
        "EMAIL": s.EMAIL,
        "ADDRESS_LINE1": s.ADDRESS_LINE1,
        "ADDRESS_LINE2": s.ADDRESS_LINE2,
        "CITY": s.CITY,
        "STATE": s.STATE,
        "PINCODE": s.PINCODE,
        "GST_NUMBER": s.GST_NUMBER,
        "PAN_NUMBER": s.PAN_NUMBER,
        "BANK_NAME": s.BANK_NAME,
        "ACCOUNT_NUMBER": s.ACCOUNT_NUMBER,
        "IFSC_CODE": s.IFSC_CODE,
        "CATEGORY": s.CATEGORY,
        "PAYMENT_TERMS": s.PAYMENT_TERMS,
        "STATUS": s.STATUS,
        "NOTES": s.NOTES,
        "VENDOR_ID": s.VENDOR_ID,
        "CREATED_AT": (
            s.CREATED_AT.isoformat() if s.CREATED_AT else None
        ),
        "REGISTRATION_NO": s.REGISTRATION_NO,
        "COMPANY_TYPE": s.COMPANY_TYPE,
        "WEBSITE": s.WEBSITE,
        "ALTERNATE_EMAIL": s.ALTERNATE_EMAIL,
        "ALTERNATE_PHONE": s.ALTERNATE_PHONE,
        "YEARS_IN_BUSINESS": s.YEARS_IN_BUSINESS,
        "ANNUAL_TURNOVER": float(s.ANNUAL_TURNOVER) if s.ANNUAL_TURNOVER is not None else None,
        "EMPLOYEE_COUNT": s.EMPLOYEE_COUNT,
        "CERTIFICATIONS": s.CERTIFICATIONS,
        "ADVANCE_PERCENT": float(s.ADVANCE_PERCENT) if s.ADVANCE_PERCENT is not None else None,
        "CREDIT_DAYS": s.CREDIT_DAYS,
        "MINIMUM_ORDER_VALUE": float(s.MINIMUM_ORDER_VALUE) if s.MINIMUM_ORDER_VALUE is not None else None,
        "LEAD_TIME_DAYS": s.LEAD_TIME_DAYS,
        "DELIVERY_MODES": s.DELIVERY_MODES,
    }


@router.post("")
def create_supplier(
    data: SupplierCreate,
    db: Session = Depends(get_db)
):

    clash = db.query(Supplier).filter(
        Supplier.VENDOR_ID == data.VENDOR_ID,
        Supplier.SUPPLIER_CODE == data.SUPPLIER_CODE
    ).first()

    if clash:

        raise HTTPException(
            status_code=409,
            detail=(
                f"SUPPLIER_CODE {data.SUPPLIER_CODE} already "
                f"exists for this vendor."
            )
        )

    supplier = Supplier(**data.dict())

    db.add(supplier)

    db.commit()

    db.refresh(supplier)

    return {
        "message": "Supplier created",
        "supplier": _serialize_supplier(supplier)
    }


@router.get("")
def list_suppliers(
    vendor_id: int = 1,
    status: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):

    vendor_id = _resolve_vendor_id(db, vendor_id)

    q = db.query(Supplier).filter(Supplier.VENDOR_ID == vendor_id)

    if status:

        q = q.filter(Supplier.STATUS == status)

    if category:

        q = q.filter(Supplier.CATEGORY == category)

    if search:

        pattern = f"%{search.strip()}%"

        q = q.filter(
            or_(
                Supplier.COMPANY_NAME.ilike(pattern),
                Supplier.SUPPLIER_CODE.ilike(pattern),
                Supplier.CONTACT_PERSON.ilike(pattern),
                Supplier.GST_NUMBER.ilike(pattern)
            )
        )

    rows = q.order_by(Supplier.COMPANY_NAME).all()

    return [_serialize_supplier(s) for s in rows]


@router.get("/categories")
def supplier_categories(
    vendor_id: int = 1,
    db: Session = Depends(get_db)
):
    """Distinct categories in use — useful for the purchase
    dropdown to filter suppliers by what they sell."""

    vendor_id = _resolve_vendor_id(db, vendor_id)

    rows = (
        db.query(Supplier.CATEGORY)
        .filter(
            Supplier.VENDOR_ID == vendor_id,
            Supplier.CATEGORY.isnot(None)
        )
        .distinct()
        .all()
    )

    return sorted([r[0] for r in rows if r[0]])


_SUPPLIER_XL_COLUMNS = [
    "SUPPLIER_CODE", "COMPANY_NAME", "CONTACT_PERSON", "PHONE", "EMAIL",
    "ADDRESS_LINE1", "ADDRESS_LINE2", "CITY", "STATE", "PINCODE",
    "GST_NUMBER", "PAN_NUMBER", "BANK_NAME", "ACCOUNT_NUMBER", "IFSC_CODE",
    "CATEGORY", "PAYMENT_TERMS", "STATUS", "NOTES",
]


@router.get("/export/excel")
def export_suppliers(
    vendor_id: int = Query(1),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    vendor_id = _resolve_vendor_id(db, vendor_id)
    q = db.query(Supplier).filter(Supplier.VENDOR_ID == vendor_id)
    if status:
        q = q.filter(Supplier.STATUS == status)
    rows = q.order_by(Supplier.COMPANY_NAME).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    ws.append(_SUPPLIER_XL_COLUMNS)
    for s in rows:
        ws.append([getattr(s, col, None) or "" for col in _SUPPLIER_XL_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suppliers_export.xlsx"},
    )


@router.get("/bulk-template")
def supplier_bulk_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    ws.append(_SUPPLIER_XL_COLUMNS)
    ws.append([
        "SUP-001", "Acme Supplies Pvt Ltd", "John Doe", "9876543210",
        "john@acme.com", "123 MG Road", "", "Bengaluru", "Karnataka", "560001",
        "29ABCDE1234F1Z5", "ABCDE1234F", "State Bank of India",
        "1234567890", "SBIN0001234", "Electronics", "Net 30", "ACTIVE", "",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=supplier_bulk_template.xlsx"},
    )


@router.get("/{supplier_id}")
def get_supplier(
    supplier_id: int,
    db: Session = Depends(get_db)
):

    supplier = db.query(Supplier).filter(
        Supplier.ID == supplier_id
    ).first()

    if not supplier:

        raise HTTPException(status_code=404, detail="Supplier not found")

    # Also surface which BOM items use this supplier — handy for
    # the supplier detail view
    using_bom_count = db.query(BOMItem).filter(
        BOMItem.PREFERRED_SUPPLIER_ID == supplier_id
    ).count()

    return {
        "supplier": _serialize_supplier(supplier),
        "bom_items_linked": using_bom_count
    }


@router.patch("/{supplier_id}")
def update_supplier(
    supplier_id: int,
    data: SupplierUpdate,
    db: Session = Depends(get_db)
):

    supplier = db.query(Supplier).filter(
        Supplier.ID == supplier_id
    ).first()

    if not supplier:

        raise HTTPException(status_code=404, detail="Supplier not found")

    for field, value in data.dict(exclude_unset=True).items():

        setattr(supplier, field, value)

    db.commit()

    db.refresh(supplier)

    return {
        "message": "Supplier updated",
        "supplier": _serialize_supplier(supplier)
    }


@router.delete("/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db)
):
    """Permanently delete a BLACKLISTED supplier.

    Only suppliers with STATUS == BLACKLISTED may be deleted.
    BOMItem.PREFERRED_SUPPLIER_ID references are nulled before deletion;
    all other child records (SupplierProduct, rankings, metrics,
    price history, purchase recommendations) cascade via DB-level ON DELETE CASCADE.
    """
    from sqlalchemy.exc import IntegrityError as _IntegrityError

    supplier = db.query(Supplier).filter(
        Supplier.ID == supplier_id
    ).first()

    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    if supplier.STATUS != "BLACKLISTED":
        raise HTTPException(
            status_code=400,
            detail=(
                "Supplier can only be permanently deleted when status is BLACKLISTED. "
                "Change the supplier status to BLACKLISTED first."
            ),
        )

    # BOMItem has no ON DELETE CASCADE/SET NULL on PREFERRED_SUPPLIER_ID — null it manually
    db.query(BOMItem).filter(
        BOMItem.PREFERRED_SUPPLIER_ID == supplier_id
    ).update({"PREFERRED_SUPPLIER_ID": None}, synchronize_session=False)

    db.delete(supplier)

    try:
        db.commit()
    except _IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=(
                "Cannot delete supplier — it is still referenced by other records "
                "(e.g. purchase orders or GRN entries). Remove those references first."
            ),
        )
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Delete failed: {exc}")

    return {"message": "Supplier permanently deleted"}


# ── Supplier extended profile & bulk operations ────────────────────────────


@router.post("/bulk-upload")
async def supplier_bulk_upload(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Bulk-create suppliers from Excel. Rows with existing SUPPLIER_CODE are skipped."""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Cannot parse file — upload a valid .xlsx file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows (need header + at least 1 data row)")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    created, skipped, errors = 0, 0, []

    for idx, raw in enumerate(rows[1:], start=2):
        row = dict(zip(headers, raw))
        code = str(row.get("SUPPLIER_CODE") or "").strip()
        if not code:
            errors.append(f"Row {idx}: SUPPLIER_CODE required — skipped")
            continue

        if db.query(Supplier).filter(
            Supplier.VENDOR_ID == vendor_id,
            Supplier.SUPPLIER_CODE == code,
        ).first():
            skipped += 1
            continue

        def _s(key: str) -> Optional[str]:
            v = row.get(key)
            return str(v).strip() or None if v is not None else None

        db.add(Supplier(
            VENDOR_ID=vendor_id,
            SUPPLIER_CODE=code,
            COMPANY_NAME=_s("COMPANY_NAME") or code,
            CONTACT_PERSON=_s("CONTACT_PERSON"),
            PHONE=_s("PHONE"),
            EMAIL=_s("EMAIL"),
            ADDRESS_LINE1=_s("ADDRESS_LINE1"),
            ADDRESS_LINE2=_s("ADDRESS_LINE2"),
            CITY=_s("CITY"),
            STATE=_s("STATE"),
            PINCODE=_s("PINCODE"),
            GST_NUMBER=_s("GST_NUMBER"),
            PAN_NUMBER=_s("PAN_NUMBER"),
            BANK_NAME=_s("BANK_NAME"),
            ACCOUNT_NUMBER=_s("ACCOUNT_NUMBER"),
            IFSC_CODE=_s("IFSC_CODE"),
            CATEGORY=_s("CATEGORY"),
            PAYMENT_TERMS=_s("PAYMENT_TERMS"),
            STATUS=(_s("STATUS") or "ACTIVE").upper(),
            NOTES=_s("NOTES"),
        ))
        created += 1

    db.commit()
    return {"created": created, "skipped_existing": skipped, "errors": errors}



@router.get("/{supplier_id}/performance")
def get_supplier_performance(supplier_id: int, db: Session = Depends(get_db)):
    supplier = db.query(Supplier).filter(Supplier.ID == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    metrics = db.query(SupplierPerformanceMetrics).filter(
        SupplierPerformanceMetrics.SUPPLIER_ID == supplier_id
    ).first()

    return {
        "supplier_id": supplier_id,
        "metrics": _serialize_performance(metrics) if metrics else None,
    }


@router.get("/{supplier_id}/ranking")
def get_supplier_ranking(
    supplier_id: int,
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    supplier = db.query(Supplier).filter(Supplier.ID == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    rows = (
        db.query(SupplierRanking)
        .filter(
            SupplierRanking.VENDOR_ID == vendor_id,
            SupplierRanking.SUPPLIER_ID == supplier_id,
        )
        .order_by(SupplierRanking.RANK.asc())
        .all()
    )

    return {
        "supplier_id": supplier_id,
        "rankings": [_serialize_ranking(r) for r in rows],
    }


@router.get("/{supplier_id}/products")
def get_supplier_products(
    supplier_id: int,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(2000, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    """Return all products registered for a supplier, with optional search and status filter."""
    supplier = db.query(Supplier).filter(Supplier.ID == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    q = (
        db.query(SupplierProductModel, ProductMaster, InventoryCategory)
        .join(ProductMaster, SupplierProductModel.PRODUCT_ID == ProductMaster.ID)
        .outerjoin(InventoryCategory, ProductMaster.CATEGORY_ID == InventoryCategory.ID)
        .filter(SupplierProductModel.SUPPLIER_ID == supplier_id)
    )

    if search:
        term = f"%{search.strip()}%"
        q = q.filter(
            or_(
                ProductMaster.PRODUCT_NAME.ilike(term),
                ProductMaster.PRODUCT_CODE.ilike(term),
            )
        )
    if status:
        q = q.filter(SupplierProductModel.STATUS == status.upper())

    total = q.count()
    rows = (
        q.order_by(ProductMaster.PRODUCT_NAME)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [
        {
            "ID": sp.ID,
            "PRODUCT_ID": sp.PRODUCT_ID,
            "PRODUCT_CODE": pm.PRODUCT_CODE,
            "PRODUCT_NAME": pm.PRODUCT_NAME,
            # Use the direct CATEGORY_ID on SupplierProduct when set;
            # fall back to the category derived via ProductMaster.
            "CATEGORY_NAME": cat.NAME if cat else None,
            "CATEGORY_ID": sp.CATEGORY_ID or (pm.CATEGORY_ID if pm else None),
            "UNIT": pm.UNIT,
            "UNIT_PRICE": float(sp.UNIT_PRICE) if sp.UNIT_PRICE is not None else 0.0,
            "CURRENCY": sp.CURRENCY or "INR",
            "MOQ": sp.MOQ,
            "LEAD_TIME_DAYS": sp.LEAD_TIME_DAYS,
            "AVAILABLE_QTY": sp.AVAILABLE_QTY,
            "IS_PREFERRED": sp.IS_PREFERRED,
            "STATUS": sp.STATUS,
            "NOTES": sp.NOTES,
            "LAST_PRICE_UPDATED_AT": sp.LAST_PRICE_UPDATED_AT.isoformat() if sp.LAST_PRICE_UPDATED_AT else None,
            "CREATED_AT": sp.CREATED_AT.isoformat() if sp.CREATED_AT else None,
        }
        for sp, pm, cat in rows
    ]

    return {"total": total, "page": page, "page_size": page_size, "items": items}



def _serialize_performance(m: SupplierPerformanceMetrics) -> dict:
    return {
        "ID": m.ID,
        "SUPPLIER_ID": m.SUPPLIER_ID,
        "VENDOR_ID": m.VENDOR_ID,
        "TOTAL_ORDERS": m.TOTAL_ORDERS,
        "COMPLETED_ORDERS": m.COMPLETED_ORDERS,
        "ON_TIME_DELIVERIES": m.ON_TIME_DELIVERIES,
        "DELAYED_DELIVERIES": m.DELAYED_DELIVERIES,
        "REJECTED_DELIVERIES": m.REJECTED_DELIVERIES,
        "TOTAL_QTY_ORDERED": m.TOTAL_QTY_ORDERED,
        "TOTAL_QTY_RECEIVED": m.TOTAL_QTY_RECEIVED,
        "TOTAL_QTY_REJECTED": m.TOTAL_QTY_REJECTED,
        "QUALITY_SCORE": m.QUALITY_SCORE,
        "PRICE_COMPETITIVENESS_SCORE": m.PRICE_COMPETITIVENESS_SCORE,
        "ON_TIME_RATE": m.ON_TIME_RATE,
        "OVERALL_SCORE": m.OVERALL_SCORE,
        "LAST_RECALCULATED_AT": m.LAST_RECALCULATED_AT.isoformat() if m.LAST_RECALCULATED_AT else None,
        "UPDATED_AT": m.UPDATED_AT.isoformat() if m.UPDATED_AT else None,
    }


def _serialize_ranking(r: SupplierRanking) -> dict:
    return {
        "ID": r.ID,
        "VENDOR_ID": r.VENDOR_ID,
        "PRODUCT_ID": r.PRODUCT_ID,
        "SUPPLIER_ID": r.SUPPLIER_ID,
        "SUPPLIER_PRODUCT_ID": r.SUPPLIER_PRODUCT_ID,
        "RANK": r.RANK,
        "PRICE_SCORE": r.PRICE_SCORE,
        "AVAILABILITY_SCORE": r.AVAILABILITY_SCORE,
        "PERFORMANCE_SCORE": r.PERFORMANCE_SCORE,
        "SENIORITY_SCORE": r.SENIORITY_SCORE,
        "COMPOSITE_SCORE": r.COMPOSITE_SCORE,
        "UNIT_PRICE_AT_RANK": float(r.UNIT_PRICE_AT_RANK) if r.UNIT_PRICE_AT_RANK else None,
        "RECALCULATED_AT": r.RECALCULATED_AT.isoformat() if r.RECALCULATED_AT else None,
    }
