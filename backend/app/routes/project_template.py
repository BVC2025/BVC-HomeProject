from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.utils.db_error_handler import raise_db_error
from typing import Optional, List
from pydantic import BaseModel, Field
import io
import csv
import json
import logging
import openpyxl

log = logging.getLogger(__name__)

from app.database.database import get_db
from app.utils.datetime_utils import now_ist
from app.auth.auth_bearer import require, get_current_admin

from app.models.models import (
    ProjectCategory,
    Project,
    TaskTemplate,
    TaskTemplateRequirement,
    TaskGroup,
    ProjectPricing,
    ProjectProductRequirement,
    Department,
    Role,
    CustomField,
    CustomFieldTableValue,
    ProjectQuotationTemplate,
    CustomerProjectTask,
)
from app.models.inventory_models import ProductMaster
from app.services.company_settings_service import get_company_settings
from app.services.project_quotation_service import (
    build_default_quotation_content,
    default_quotation_number,
    render_quotation_html,
    sync_final_price_into_quotation,
)


router = APIRouter()


# =========================
# SCHEMAS
# =========================

class CategoryCreate(BaseModel):
    NAME: str
    DESCRIPTION: Optional[str] = None
    VENDOR_ID: int = 1


class CategoryUpdate(BaseModel):
    NAME: Optional[str] = None
    DESCRIPTION: Optional[str] = None


class TaskTemplateRequirementIn(BaseModel):
    DEPARTMENT_ID: int
    ROLE_ID: int
    EXPERIENCE_LEVEL: str
    REQUIRED_COUNT: int = Field(default=1, ge=1)


class TaskGroupIn(BaseModel):
    """Wizard path only (ProjectCreate/ProjectUpdate.task_groups) —
    `task_indexes` and `DEPENDS_ON_TASK_INDEX` are 0-based positions within
    the SAME `tasks` array being submitted, since brand-new tasks don't
    have real IDs yet at payload-construction time. DEPENDS_ON_TASK_INDEX
    is only meaningful (and required) when DEPENDENCY_RULE == "ONE", and
    must itself be one of `task_indexes` — validated server-side."""
    NAME: Optional[str] = None
    task_indexes: List[int] = []
    DEPENDENCY_RULE: str = "ALL"
    DEPENDS_ON_TASK_INDEX: Optional[int] = None


class TaskGroupCreate(BaseModel):
    """Standalone /projects/{id}/task-groups path — real, already-
    persisted task IDs. DEPENDS_ON_TASK_TEMPLATE_ID is only meaningful
    (and required) when DEPENDENCY_RULE == "ONE", and must itself be one
    of `task_template_ids` — validated server-side."""
    NAME: Optional[str] = None
    task_template_ids: List[str] = []
    DEPENDENCY_RULE: str = "ALL"
    DEPENDS_ON_TASK_TEMPLATE_ID: Optional[str] = None


class TaskGroupUpdate(BaseModel):
    """Full-replace-when-provided, matching the existing convention used
    for TaskTemplateRequirement lists: omit a field to leave it untouched,
    provide it (including an empty list/None) to replace it entirely."""
    NAME: Optional[str] = None
    task_template_ids: Optional[List[str]] = None
    DEPENDENCY_RULE: Optional[str] = None
    DEPENDS_ON_TASK_TEMPLATE_ID: Optional[str] = None


class TaskTemplateIn(BaseModel):
    NAME: str
    DESCRIPTION: Optional[str] = None
    DURATION_VALUE: float = 1.0
    DURATION_UNIT: str = "DAYS"
    SEQUENCE_NUMBER: int = 0
    TASK_SCOPE: Optional[str] = "PROJECT"
    requirements: List[TaskTemplateRequirementIn] = []


class ProjectProductRequirementIn(BaseModel):
    PRODUCT_ID: str
    REQUIRED_QTY: float = 1.0


class ProjectCreate(BaseModel):
    CATEGORY_ID: str
    NAME: str
    DESCRIPTION: Optional[str] = None
    BOM_MODE: Optional[str] = None
    ASSIGNMENT_MODE: Optional[str] = None
    VENDOR_ID: int = 1
    tasks: Optional[List[TaskTemplateIn]] = []
    task_groups: Optional[List[TaskGroupIn]] = []
    product_requirements: Optional[List[ProjectProductRequirementIn]] = []


class ProjectUpdate(BaseModel):
    NAME: Optional[str] = None
    DESCRIPTION: Optional[str] = None
    BOM_MODE: Optional[str] = None
    ASSIGNMENT_MODE: Optional[str] = None
    CATEGORY_ID: Optional[str] = None
    tasks: Optional[List[TaskTemplateIn]] = None
    task_groups: Optional[List[TaskGroupIn]] = None
    product_requirements: Optional[List[ProjectProductRequirementIn]] = None
    VENDOR_ID: int = 1


class TaskTemplateCreate(BaseModel):
    PROJECT_ID: str
    NAME: str
    DESCRIPTION: Optional[str] = None
    DURATION_VALUE: float = 1.0
    DURATION_UNIT: str = "DAYS"
    SEQUENCE_NUMBER: int = 0
    TASK_SCOPE: Optional[str] = "PROJECT"
    requirements: List[TaskTemplateRequirementIn] = []
    VENDOR_ID: int = 1


class TaskTemplateUpdate(BaseModel):
    NAME: Optional[str] = None
    DESCRIPTION: Optional[str] = None
    DURATION_VALUE: Optional[float] = None
    DURATION_UNIT: Optional[str] = None
    SEQUENCE_NUMBER: Optional[int] = None
    TASK_SCOPE: Optional[str] = None
    requirements: Optional[List[TaskTemplateRequirementIn]] = None


class ReorderItem(BaseModel):
    id: str
    sequence_number: int


class ProjectPricingCreate(BaseModel):
    PROJECT_ID: str
    CURRENCY: str = "INR"
    ORIGINAL_PRICE: float
    MINIMUM_NEGOTIATION_PRICE: Optional[float] = None
    NEGOTIATION_PERCENT: Optional[float] = None
    PACKING_CHARGE: float = 0
    TRANSPORTATION_CHARGE: float = 0
    INSTALLATION_CHARGE: float = 0
    SERVICE_CHARGE: float = 0
    ADDITIONAL_CHARGES: float = 0
    TAX_AMOUNT: float = 0
    DISCOUNT_AMOUNT: float = 0
    REMARKS: Optional[str] = None
    IS_ACTIVE: bool = True
    VENDOR_ID: int = 1


class ProjectPricingUpdate(BaseModel):
    CURRENCY: Optional[str] = None
    ORIGINAL_PRICE: Optional[float] = None
    MINIMUM_NEGOTIATION_PRICE: Optional[float] = None
    NEGOTIATION_PERCENT: Optional[float] = None
    PACKING_CHARGE: Optional[float] = None
    TRANSPORTATION_CHARGE: Optional[float] = None
    INSTALLATION_CHARGE: Optional[float] = None
    SERVICE_CHARGE: Optional[float] = None
    ADDITIONAL_CHARGES: Optional[float] = None
    TAX_AMOUNT: Optional[float] = None
    DISCOUNT_AMOUNT: Optional[float] = None
    REMARKS: Optional[str] = None
    IS_ACTIVE: Optional[bool] = None


# =========================
# DURATION HELPERS
# =========================

_DEFAULT_WORK_HOURS = 8.0  # used whenever the vendor has no working schedule configured

_UNIT_TO_DAYS = {
    "DAYS":   1.0,
    "WEEKS":  5.0,
    "MONTHS": 22.0,
    "YEARS":  260.0,
}


def _resolve_work_hours(db: Session, project: Project) -> float:
    company = get_company_settings(db, project.VENDOR_ID)
    if company.WORK_HOURS and float(company.WORK_HOURS) > 0:
        return float(company.WORK_HOURS)
    return _DEFAULT_WORK_HOURS


def _to_hours(value: float, unit: str, work_hours: float) -> float:
    """Normalizes a task's DURATION_VALUE/DURATION_UNIT to hours using the
    given company working-hours-per-day figure."""
    unit = (unit or "DAYS").upper()
    if unit == "HOURS":
        return float(value)
    return float(value) * _UNIT_TO_DAYS.get(unit, 1.0) * work_hours


def calculate_project_estimated_duration(db: Session, project: Project) -> dict:
    """Centralized, single source of truth for Project.ESTIMATED_TOTAL_DAYS
    — the only place this calculation happens; every mutation that can
    affect it (task create/update/delete, project create/update, Task
    Group create/update/delete) calls this instead of duplicating the math.

    Every task's DURATION_VALUE/DURATION_UNIT is normalized to hours using
    the vendor's configured company working hours (falling back to the
    existing 8h/day default when unconfigured — see _resolve_work_hours).
    Tasks sharing a TASK_GROUP_ID run in parallel: a group's duration is
    the MAX of its members' hours, not their sum. Top-level items (each
    group, each ungrouped/standalone task) then combine according to
    Project.ASSIGNMENT_MODE: SEQUENTIAL sums them (matches the existing
    sequence/dependency-driven flow); PARALLEL takes their MAX (all
    eligible top-level items start together and progress independently, so
    the project's expected completion is the longest one, not their sum).

    Writes project.ESTIMATED_TOTAL_DAYS and returns a breakdown dict for
    the Review UI. Callers commit alongside their own transaction — this
    function only flushes, so a failed sibling write in the same request
    rolls the duration change back with it."""

    tasks = (
        db.query(TaskTemplate)
          .filter(TaskTemplate.PROJECT_ID == project.ID)
          .order_by(TaskTemplate.SEQUENCE_NUMBER)
          .all()
    )
    work_hours = _resolve_work_hours(db, project)

    grouped: "dict[str, list]" = {}
    standalone = []
    for t in tasks:
        hours = _to_hours(float(t.DURATION_VALUE), t.DURATION_UNIT, work_hours)
        if t.TASK_GROUP_ID:
            grouped.setdefault(t.TASK_GROUP_ID, []).append((t, hours))
        else:
            standalone.append((t, hours))

    group_name_by_id = {}
    if grouped:
        group_rows = db.query(TaskGroup).filter(TaskGroup.ID.in_(grouped.keys())).all()
        group_name_by_id = {g.ID: g.NAME for g in group_rows}

    group_breakdown = []
    top_level_hours = []
    for i, (group_id, members) in enumerate(grouped.items()):
        duration_hours = max(h for _, h in members)
        group_breakdown.append({
            "TASK_GROUP_ID": group_id,
            "NAME": group_name_by_id.get(group_id) or f"Group {i + 1}",
            "task_count": len(members),
            "duration_hours": round(duration_hours, 2),
            "duration_days": round(duration_hours / work_hours, 2),
        })
        top_level_hours.append(duration_hours)

    standalone_breakdown = []
    for t, hours in standalone:
        standalone_breakdown.append({
            "TASK_TEMPLATE_ID": t.ID,
            "NAME": t.NAME,
            "duration_hours": round(hours, 2),
            "duration_days": round(hours / work_hours, 2),
        })
        top_level_hours.append(hours)

    if not top_level_hours:
        total_hours = 0.0
    elif project.ASSIGNMENT_MODE == "PARALLEL":
        total_hours = max(top_level_hours)
    else:
        total_hours = sum(top_level_hours)

    total_days = round(total_hours / work_hours, 2) if work_hours else 0.0

    project.ESTIMATED_TOTAL_DAYS = total_days
    db.flush()

    return {
        "total_days": total_days,
        "total_hours": round(total_hours, 2),
        "work_hours": work_hours,
        "assignment_mode": project.ASSIGNMENT_MODE,
        "groups": group_breakdown,
        "standalone_tasks": standalone_breakdown,
    }


# =========================
# TASK HELPERS
# =========================

def _requirement_to_dict(r: TaskTemplateRequirement, dept_name=None, role_name=None):
    return {
        "ID": r.ID,
        "DEPARTMENT_ID": r.DEPARTMENT_ID,
        "DEPARTMENT_NAME": dept_name,
        "ROLE_ID": r.ROLE_ID,
        "ROLE_NAME": role_name,
        "EXPERIENCE_LEVEL": r.EXPERIENCE_LEVEL,
        "REQUIRED_COUNT": r.REQUIRED_COUNT,
    }


def _task_to_dict(t: TaskTemplate, requirements: list, group_name_by_id: dict = None):
    group_name_by_id = group_name_by_id or {}
    return {
        "ID": t.ID,
        "PROJECT_ID": t.PROJECT_ID,
        "NAME": t.NAME,
        "DESCRIPTION": t.DESCRIPTION,
        "DURATION_VALUE": float(t.DURATION_VALUE) if t.DURATION_VALUE is not None else 1.0,
        "DURATION_UNIT": t.DURATION_UNIT,
        "SEQUENCE_NUMBER": t.SEQUENCE_NUMBER,
        "TASK_SCOPE": t.TASK_SCOPE,
        # Read-only, informational — group membership is configured
        # exclusively through the Task Group endpoints, never here.
        "TASK_GROUP_ID": t.TASK_GROUP_ID,
        "TASK_GROUP_NAME": group_name_by_id.get(t.TASK_GROUP_ID) if t.TASK_GROUP_ID else None,
        "requirements": requirements,
        "TOTAL_REQUIRED_COUNT": sum(r["REQUIRED_COUNT"] for r in requirements),
        "VENDOR_ID": t.VENDOR_ID,
        "CREATED_AT": t.CREATED_AT.isoformat() if t.CREATED_AT else None,
        "UPDATED_AT": t.UPDATED_AT.isoformat() if t.UPDATED_AT else None
    }


def _enrich_tasks(tasks, db):
    """Batch-fetches every referenced Department/Role/group name once
    (instead of per row) and assembles each task's requirements list."""
    task_ids = [t.ID for t in tasks]
    reqs = (
        db.query(TaskTemplateRequirement)
          .filter(TaskTemplateRequirement.TASK_TEMPLATE_ID.in_(task_ids))
          .all()
        if task_ids else []
    )
    dept_ids = {r.DEPARTMENT_ID for r in reqs if r.DEPARTMENT_ID}
    role_ids = {r.ROLE_ID for r in reqs if r.ROLE_ID}
    dept_map = {
        d.ID: d.NAME for d in db.query(Department).filter(Department.ID.in_(dept_ids)).all()
    } if dept_ids else {}
    role_map = {
        r.ID: r.NAME for r in db.query(Role).filter(Role.ID.in_(role_ids)).all()
    } if role_ids else {}

    reqs_by_task = {}
    for r in reqs:
        reqs_by_task.setdefault(r.TASK_TEMPLATE_ID, []).append(
            _requirement_to_dict(r, dept_map.get(r.DEPARTMENT_ID), role_map.get(r.ROLE_ID))
        )

    group_ids = {t.TASK_GROUP_ID for t in tasks if t.TASK_GROUP_ID}
    group_name_by_id = {
        g.ID: g.NAME for g in db.query(TaskGroup).filter(TaskGroup.ID.in_(group_ids)).all()
    } if group_ids else {}

    return [
        _task_to_dict(t, reqs_by_task.get(t.ID, []), group_name_by_id)
        for t in tasks
    ]


def _task_group_to_dict(g: TaskGroup, member_dicts: list, depends_on_name: str = None):
    return {
        "ID": g.ID,
        "PROJECT_ID": g.PROJECT_ID,
        "NAME": g.NAME,
        "SEQUENCE_NUMBER": g.SEQUENCE_NUMBER,
        "DEPENDENCY_RULE": g.DEPENDENCY_RULE,
        # Only meaningful when DEPENDENCY_RULE == "ONE" — always one of
        # this group's own `task_templates` members (see
        # DEPENDS_ON_TASK_NAME, resolved from that same member list, no
        # extra query needed).
        "DEPENDS_ON_TASK_TEMPLATE_ID": g.DEPENDS_ON_TASK_TEMPLATE_ID,
        "DEPENDS_ON_TASK_NAME": depends_on_name,
        "task_templates": member_dicts,
        "VENDOR_ID": g.VENDOR_ID,
        "CREATED_AT": g.CREATED_AT.isoformat() if g.CREATED_AT else None,
        "UPDATED_AT": g.UPDATED_AT.isoformat() if g.UPDATED_AT else None,
    }


def _enrich_task_groups(groups, db):
    """Batch-fetches every group's member tasks (with requirements),
    mirroring _enrich_tasks' batching approach. The dependency target (if
    any) is always one of a group's own members, so its name is resolved
    directly from that same batch — no separate query needed."""
    group_ids = [g.ID for g in groups]
    if not group_ids:
        return []

    members = (
        db.query(TaskTemplate)
          .filter(TaskTemplate.TASK_GROUP_ID.in_(group_ids))
          .order_by(TaskTemplate.SEQUENCE_NUMBER)
          .all()
    )
    member_task_dicts = _enrich_tasks(members, db) if members else []
    members_by_group = {}
    name_by_task_id = {}
    for d in member_task_dicts:
        members_by_group.setdefault(d["TASK_GROUP_ID"], []).append(d)
        name_by_task_id[d["ID"]] = d["NAME"]

    return [
        _task_group_to_dict(
            g,
            members_by_group.get(g.ID, []),
            name_by_task_id.get(g.DEPENDS_ON_TASK_TEMPLATE_ID),
        )
        for g in groups
    ]


def _validate_no_duplicate_requirements(requirements: list):
    """A Department + Role + Experience Level combination appearing twice
    in the same task isn't meaningful — "2 Intermediate Technicians" is one
    requirement with REQUIRED_COUNT=2, not two rows of 1."""
    seen = set()
    for req in requirements:
        key = (req.DEPARTMENT_ID, req.ROLE_ID, req.EXPERIENCE_LEVEL)
        if key in seen:
            raise HTTPException(
                status_code=400,
                detail="This Department + Role + Experience Level combination is already added — "
                       "increase its Required Count instead of adding a duplicate row.",
            )
        seen.add(key)


def _create_requirement_rows(task_id: str, requirements: list, db: Session):
    _validate_no_duplicate_requirements(requirements)
    for req in requirements:
        db.add(TaskTemplateRequirement(
            TASK_TEMPLATE_ID=task_id,
            DEPARTMENT_ID=req.DEPARTMENT_ID,
            ROLE_ID=req.ROLE_ID,
            EXPERIENCE_LEVEL=req.EXPERIENCE_LEVEL,
            REQUIRED_COUNT=req.REQUIRED_COUNT,
        ))


def _validate_unique_sequence(db: Session, project_id: str, sequence_number: int, exclude_task_id: str = None):
    """Two tasks in the same project must not share a SEQUENCE_NUMBER —
    only relevant where the Sequence # field is directly user-editable
    (the standalone /task-templates Add/Edit modal); the wizard's embedded
    tasks auto-assign via array index and reorder always reassigns a
    contiguous 0..N-1 range, so neither can produce a collision."""
    q = db.query(TaskTemplate).filter(
        TaskTemplate.PROJECT_ID == project_id,
        TaskTemplate.SEQUENCE_NUMBER == sequence_number,
    )
    if exclude_task_id:
        q = q.filter(TaskTemplate.ID != exclude_task_id)
    clash = q.first()
    if clash:
        raise HTTPException(
            status_code=400,
            detail=f"Sequence Number {sequence_number} is already used by task \"{clash.NAME}\" in this project.",
        )


_VALID_DEPENDENCY_RULES = ("ALL", "ANY", "ONE")


def _normalize_dependency_rule(rule: Optional[str]) -> str:
    """Validates rule membership only."""
    rule = (rule or "ALL").upper()
    if rule not in _VALID_DEPENDENCY_RULES:
        raise HTTPException(status_code=400, detail=f"Invalid dependency rule '{rule}'. Must be ALL, ANY, or ONE.")
    return rule


def _resolve_group_dependency(rule: str, depends_on_id: Optional[str], member_task_ids: list) -> Optional[str]:
    """A TaskGroup's dependency target is always one of its own members —
    there is no external "depends on another group/task" concept. For
    ALL/ANY the rule is evaluated over every/any member directly, so no
    explicit target is stored (any supplied value is ignored, not
    errored — forgiving of a stray client value left over from switching
    rules). For ONE, exactly one target must be supplied and it must be
    one of `member_task_ids`."""
    if rule != "ONE":
        return None
    if not depends_on_id:
        raise HTTPException(
            status_code=400,
            detail="Dependency Rule 'ONE' requires selecting one task to depend on.",
        )
    if depends_on_id not in member_task_ids:
        raise HTTPException(
            status_code=400,
            detail="The dependency task must be one of this group's own selected tasks.",
        )
    return depends_on_id


def _project_has_live_production_tasks(db: Session, project_id: str) -> bool:
    """True once at least one real CustomerProjectTask row exists for any
    of this project's TaskTemplate rows — i.e. a customer's production
    schedule has actually been approved and tasks generated/assigned
    against the CURRENT task list (see production_scheduling_service.
    approve_schedule()/reject_and_reschedule() -> task_generation_
    service.generate_tasks_for_schedule()). CustomerProjectTask.
    TASK_TEMPLATE_ID has an ON DELETE RESTRICT FK into task_template, so
    the wizard's own delete-then-recreate replacement of tasks/groups
    would otherwise crash with an unhandled IntegrityError (500) the
    moment this becomes true for a project — this check exists so
    update_project() can detect that BEFORE attempting the delete, and
    skip the task/group replacement instead of crashing or silently
    destroying live production data."""
    return db.query(CustomerProjectTask.ID).join(
        TaskTemplate, TaskTemplate.ID == CustomerProjectTask.TASK_TEMPLATE_ID
    ).filter(TaskTemplate.PROJECT_ID == project_id).first() is not None


def _create_tasks_and_groups(db: Session, project: Project, tasks_in: list, groups_in: list, vendor_id: int) -> list:
    """Shared 3-pass task + group creation, used by both create_project()
    and update_project() whenever a `tasks` array is supplied:
      1. create every TaskTemplate row (brand new, no group/dependency
         fields — those are configured exclusively via Task Groups now).
      2. create every TaskGroup, resolving each group's `task_indexes`
         (0-based positions in `tasks_in`) into the real TaskTemplate IDs
         from pass 1, assigning membership, and resolving its own
         DEPENDS_ON_TASK_INDEX (only meaningful for rule ONE — always one
         of that SAME group's own indexes) into a real TaskTemplate ID.
    Returns the created TaskTemplate IDs in submitted order."""

    created_task_ids = []
    for i, t in enumerate(tasks_in):
        task = TaskTemplate(
            PROJECT_ID=project.ID,
            NAME=t.NAME,
            DESCRIPTION=t.DESCRIPTION,
            DURATION_VALUE=t.DURATION_VALUE,
            DURATION_UNIT=t.DURATION_UNIT,
            SEQUENCE_NUMBER=t.SEQUENCE_NUMBER if t.SEQUENCE_NUMBER else i,
            TASK_SCOPE=t.TASK_SCOPE or "PROJECT",
            VENDOR_ID=vendor_id,
        )
        db.add(task)
        db.flush()
        created_task_ids.append(task.ID)
        _create_requirement_rows(task.ID, t.requirements, db)

    groups_in = groups_in or []
    created_group_ids = []
    for g in groups_in:
        member_ids = []
        for idx in g.task_indexes:
            if idx < 0 or idx >= len(created_task_ids):
                raise HTTPException(status_code=400, detail=f"Invalid task reference in group \"{g.NAME or ''}\".")
            member_ids.append(created_task_ids[idx])
        if len(member_ids) != len(set(member_ids)):
            raise HTTPException(status_code=400, detail=f"A task can only appear once in group \"{g.NAME or ''}\".")

        rule = _normalize_dependency_rule(g.DEPENDENCY_RULE)

        depends_on_id = None
        if g.DEPENDS_ON_TASK_INDEX is not None:
            if g.DEPENDS_ON_TASK_INDEX < 0 or g.DEPENDS_ON_TASK_INDEX >= len(created_task_ids):
                raise HTTPException(status_code=400, detail=f"Invalid dependency reference in group \"{g.NAME or ''}\".")
            depends_on_id = created_task_ids[g.DEPENDS_ON_TASK_INDEX]
        depends_on_id = _resolve_group_dependency(rule, depends_on_id, member_ids)

        group = TaskGroup(
            PROJECT_ID=project.ID,
            VENDOR_ID=vendor_id,
            NAME=g.NAME or None,
            DEPENDENCY_RULE=rule,
            DEPENDS_ON_TASK_TEMPLATE_ID=depends_on_id,
            SEQUENCE_NUMBER=len(created_group_ids),
        )
        db.add(group)
        db.flush()
        created_group_ids.append(group.ID)

        if member_ids:
            db.query(TaskTemplate).filter(TaskTemplate.ID.in_(member_ids)).update(
                {"TASK_GROUP_ID": group.ID}, synchronize_session=False
            )

    return created_task_ids


# =========================
# PROJECT CATEGORIES
# =========================

@router.get("/project-categories", dependencies=[Depends(require("project.view", "project.categories.view"))])
def list_categories(
    vendor_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(ProjectCategory)
    if vendor_id is not None:
        q = q.filter(ProjectCategory.VENDOR_ID == vendor_id)
    if search:
        term = f"%{search}%"
        q = q.filter(ProjectCategory.NAME.ilike(term))
    rows = q.order_by(ProjectCategory.NAME).all()
    return [
        {
            "ID": c.ID,
            "NAME": c.NAME,
            "DESCRIPTION": c.DESCRIPTION,
            "VENDOR_ID": c.VENDOR_ID,
            "PROJECT_COUNT": len(c.projects),
            "CREATED_AT": c.CREATED_AT.isoformat() if c.CREATED_AT else None,
            "UPDATED_AT": c.UPDATED_AT.isoformat() if c.UPDATED_AT else None
        }
        for c in rows
    ]


@router.get("/project-categories/{category_id}", dependencies=[Depends(require("project.view", "project.categories.view"))])
def get_category(category_id: str, db: Session = Depends(get_db)):
    c = db.query(ProjectCategory).filter(ProjectCategory.ID == category_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Category not found")
    return {
        "ID": c.ID,
        "NAME": c.NAME,
        "DESCRIPTION": c.DESCRIPTION,
        "VENDOR_ID": c.VENDOR_ID,
        "PROJECT_COUNT": len(c.projects),
        "CREATED_AT": c.CREATED_AT.isoformat() if c.CREATED_AT else None,
        "UPDATED_AT": c.UPDATED_AT.isoformat() if c.UPDATED_AT else None
    }


@router.post("/project-categories", dependencies=[Depends(require("project.create", "project.categories.create"))])
def create_category(data: CategoryCreate, db: Session = Depends(get_db)):
    existing = db.query(ProjectCategory).filter(
        ProjectCategory.VENDOR_ID == data.VENDOR_ID,
        ProjectCategory.NAME == data.NAME
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Category '{data.NAME}' already exists")
    cat = ProjectCategory(
        NAME=data.NAME,
        DESCRIPTION=data.DESCRIPTION,
        VENDOR_ID=data.VENDOR_ID
    )
    try:
        db.add(cat)
        db.commit()
        db.refresh(cat)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create project category")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create project category")
    return {"message": "Category created", "ID": cat.ID}


@router.put("/project-categories/{category_id}", dependencies=[Depends(require("project.update", "project.categories.update"))])
def update_category(category_id: str, data: CategoryUpdate, db: Session = Depends(get_db)):
    cat = db.query(ProjectCategory).filter(ProjectCategory.ID == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    if data.NAME is not None:
        cat.NAME = data.NAME
    if data.DESCRIPTION is not None:
        cat.DESCRIPTION = data.DESCRIPTION
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update project category")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update project category")
    return {"message": "Category updated"}


@router.delete("/project-categories/{category_id}", dependencies=[Depends(require("project.delete", "project.categories.delete"))])
def delete_category(category_id: str, db: Session = Depends(get_db)):
    cat = db.query(ProjectCategory).filter(ProjectCategory.ID == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    if cat.projects:
        raise HTTPException(
            status_code=400,
            detail="Category has projects. Delete them first."
        )
    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "project_category",
        CustomFieldTableValue.TABLE_ROW_ID == str(category_id),
    ).delete(synchronize_session=False)
    try:
        db.delete(cat)
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete project category")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete project category")
    return {"message": "Category deleted"}


# =========================
# PROJECTS (formerly SubProjectTemplates)
# =========================

@router.get("/projects", dependencies=[Depends(require("project.view"))])
def list_projects(
    category_id: Optional[str] = Query(None),
    vendor_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(Project, ProjectCategory).join(
        ProjectCategory, Project.CATEGORY_ID == ProjectCategory.ID
    )
    if vendor_id is not None:
        q = q.filter(Project.VENDOR_ID == vendor_id)
    if category_id:
        q = q.filter(Project.CATEGORY_ID == category_id)
    if search:
        term = f"%{search}%"
        q = q.filter(Project.NAME.ilike(term))
    rows = q.order_by(Project.NAME).all()
    return [
        {
            "ID": p.ID,
            "NAME": p.NAME,
            "DESCRIPTION": p.DESCRIPTION,
            "CATEGORY_ID": p.CATEGORY_ID,
            "CATEGORY_NAME": c.NAME,
            "BOM_MODE": p.BOM_MODE,
            "ASSIGNMENT_MODE": p.ASSIGNMENT_MODE,
            "ESTIMATED_TOTAL_DAYS": float(p.ESTIMATED_TOTAL_DAYS) if p.ESTIMATED_TOTAL_DAYS else 0.0,
            "TASK_COUNT": len(p.task_templates),
            "VENDOR_ID": p.VENDOR_ID,
            "CREATED_AT": p.CREATED_AT.isoformat() if p.CREATED_AT else None,
            "UPDATED_AT": p.UPDATED_AT.isoformat() if p.UPDATED_AT else None
        }
        for p, c in rows
    ]


@router.get("/projects/{project_id}", dependencies=[Depends(require("project.view"))])
def get_project(project_id: str, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.ID == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    cat = db.query(ProjectCategory).filter(ProjectCategory.ID == p.CATEGORY_ID).first()
    tasks = _enrich_tasks(
        db.query(TaskTemplate)
            .filter(TaskTemplate.PROJECT_ID == project_id)
            .order_by(TaskTemplate.SEQUENCE_NUMBER)
            .all(),
        db
    )
    return {
        "ID": p.ID,
        "NAME": p.NAME,
        "DESCRIPTION": p.DESCRIPTION,
        "CATEGORY_ID": p.CATEGORY_ID,
        "CATEGORY_NAME": cat.NAME if cat else None,
        "BOM_MODE": p.BOM_MODE,
        "ASSIGNMENT_MODE": p.ASSIGNMENT_MODE,
        "ESTIMATED_TOTAL_DAYS": float(p.ESTIMATED_TOTAL_DAYS) if p.ESTIMATED_TOTAL_DAYS else 0.0,
        "VENDOR_ID": p.VENDOR_ID,
        "CREATED_AT": p.CREATED_AT.isoformat() if p.CREATED_AT else None,
        "UPDATED_AT": p.UPDATED_AT.isoformat() if p.UPDATED_AT else None,
        "tasks": tasks
    }


def _create_product_requirements(db: Session, project: Project, requirements_in: list, vendor_id: int) -> None:
    """Bulk-creates ProjectProductRequirement rows for a project — same
    "caller deletes existing rows first on update, this only ever
    inserts" convention as _create_tasks_and_groups. Validates every
    PRODUCT_ID exists (for this vendor) and rejects a duplicate
    PRODUCT_ID within the same submitted list (the DB unique constraint
    on PROJECT_ID+PRODUCT_ID is the final backstop, but failing fast
    here gives a much clearer error message than a raw IntegrityError)."""
    if not requirements_in:
        return
    seen_product_ids = set()
    for req in requirements_in:
        if req.PRODUCT_ID in seen_product_ids:
            raise HTTPException(status_code=400, detail="The same product cannot be added twice to a project's inventory requirements.")
        seen_product_ids.add(req.PRODUCT_ID)

    products = {
        p.ID: p for p in db.query(ProductMaster).filter(
            ProductMaster.ID.in_(seen_product_ids), ProductMaster.VENDOR_ID == vendor_id,
        ).all()
    }
    missing = seen_product_ids - set(products.keys())
    if missing:
        raise HTTPException(status_code=404, detail=f"Product(s) not found: {', '.join(missing)}")

    for req in requirements_in:
        db.add(ProjectProductRequirement(
            PROJECT_ID=project.ID, VENDOR_ID=vendor_id,
            PRODUCT_ID=req.PRODUCT_ID, REQUIRED_QTY=req.REQUIRED_QTY or 1.0,
        ))


def _serialize_product_requirement(req: ProjectProductRequirement) -> dict:
    product = req.product
    return {
        "ID": req.ID,
        "PROJECT_ID": req.PROJECT_ID,
        "PRODUCT_ID": req.PRODUCT_ID,
        "PRODUCT_CODE": product.PRODUCT_CODE if product else None,
        "PRODUCT_NAME": product.PRODUCT_NAME if product else None,
        "UNIT": product.UNIT if product else None,
        "CATEGORY_ID": product.CATEGORY_ID if product else None,
        "REQUIRED_QTY": req.REQUIRED_QTY,
    }


@router.get("/projects/{project_id}/product-requirements", dependencies=[Depends(require("project.view"))])
def list_product_requirements(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    rows = db.query(ProjectProductRequirement).filter(ProjectProductRequirement.PROJECT_ID == project_id).all()
    return [_serialize_product_requirement(r) for r in rows]


@router.post("/projects", dependencies=[Depends(require("project.create"))])
def create_project(data: ProjectCreate, db: Session = Depends(get_db)):
    cat = db.query(ProjectCategory).filter(ProjectCategory.ID == data.CATEGORY_ID).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    existing = db.query(Project).filter(
        Project.VENDOR_ID == data.VENDOR_ID,
        Project.CATEGORY_ID == data.CATEGORY_ID,
        Project.NAME == data.NAME
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Project '{data.NAME}' already exists in this category")

    project = Project(
        CATEGORY_ID=data.CATEGORY_ID,
        NAME=data.NAME,
        DESCRIPTION=data.DESCRIPTION,
        BOM_MODE=data.BOM_MODE,
        ASSIGNMENT_MODE=data.ASSIGNMENT_MODE or "PARALLEL",
        ESTIMATED_TOTAL_DAYS=0.0,
        VENDOR_ID=data.VENDOR_ID
    )
    db.add(project)
    db.flush()

    # Auto-create a default quotation template for this project (Project
    # Quotation Management, phase 1) — every project gets its own independent
    # quotation document, seeded from the standard default layout.
    _company = get_company_settings(db, data.VENDOR_ID)
    _content = build_default_quotation_content(project, _company)
    _qtn = ProjectQuotationTemplate(
        PROJECT_ID=project.ID,
        VENDOR_ID=data.VENDOR_ID,
        QUOTATION_NUMBER=default_quotation_number(project, _company, now_ist().date()),
        QUOTATION_DATE=now_ist().date(),
        CONTENT_JSON=json.dumps(_content),
    )
    _qtn.RENDERED_HTML = render_quotation_html(_qtn, _company)
    db.add(_qtn)

    if data.tasks:
        _create_tasks_and_groups(db, project, data.tasks, data.task_groups, data.VENDOR_ID)
        db.flush()
        calculate_project_estimated_duration(db, project)

    if data.product_requirements:
        _create_product_requirements(db, project, data.product_requirements, data.VENDOR_ID)
        db.flush()

    try:
        db.commit()
        db.refresh(project)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create project")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create project")
    return {"message": "Project created", "ID": project.ID}


@router.put("/projects/{project_id}", dependencies=[Depends(require("project.update"))])
def update_project(project_id: str, data: ProjectUpdate, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if data.NAME is not None:
        project.NAME = data.NAME
    if data.DESCRIPTION is not None:
        project.DESCRIPTION = data.DESCRIPTION
    if data.BOM_MODE is not None:
        project.BOM_MODE = data.BOM_MODE
    if data.ASSIGNMENT_MODE is not None:
        project.ASSIGNMENT_MODE = data.ASSIGNMENT_MODE
    if data.CATEGORY_ID is not None:
        cat = db.query(ProjectCategory).filter(ProjectCategory.ID == data.CATEGORY_ID).first()
        if not cat:
            raise HTTPException(status_code=404, detail="Category not found")
        project.CATEGORY_ID = data.CATEGORY_ID
    tasks_skipped_reason = None
    if data.tasks is not None:
        # Bulk deletes bypass ORM cascade and rely on DB-level ON DELETE
        # rules. TaskTemplateRequirement has an FK *into* task_template, so
        # it cascades automatically — but TaskGroup does NOT (it's the
        # other direction: TaskTemplate points at TaskGroup). Without this
        # explicit delete, every group would be orphaned (zero members,
        # never cleaned up) on every wizard save.
        # task_groups replacement is coupled to the same `tasks is not
        # None` branch since stale group task_indexes can't stay
        # referentially sane against a fresh task list.
        #
        # BUT: once a customer's production schedule has actually been
        # approved for this project, real CustomerProjectTask rows exist
        # that reference the CURRENT TaskTemplate rows via an ON DELETE
        # RESTRICT FK — deleting them out from under live, already-
        # assigned production work would either crash with an unhandled
        # IntegrityError (what actually happened in production before this
        # guard existed) or, if the FK were ever relaxed, silently orphan
        # real assignment history. So once live tasks exist, the task/
        # group list is frozen — every OTHER field on this same request
        # (name, description, category, assignment mode, inventory
        # requirements) still saves normally; only the task/group
        # replacement is skipped, and the response says so explicitly
        # rather than silently dropping the edit.
        if _project_has_live_production_tasks(db, project_id):
            tasks_skipped_reason = (
                "This project already has production tasks generated and assigned for at least one "
                "customer order, so its task list can no longer be edited here — doing so would break that "
                "live production data. Every other change in this save (name, description, category, "
                "inventory requirements, etc.) was still applied."
            )
        else:
            db.query(TaskGroup).filter(TaskGroup.PROJECT_ID == project_id).delete()
            db.query(TaskTemplate).filter(TaskTemplate.PROJECT_ID == project_id).delete()
            db.flush()
            _create_tasks_and_groups(db, project, data.tasks, data.task_groups, data.VENDOR_ID)
            db.flush()
            calculate_project_estimated_duration(db, project)
    if data.product_requirements is not None:
        db.query(ProjectProductRequirement).filter(ProjectProductRequirement.PROJECT_ID == project_id).delete()
        db.flush()
        _create_product_requirements(db, project, data.product_requirements, data.VENDOR_ID)
        db.flush()
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update project")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update project")
    return {
        "message": "Project updated",
        "tasks_skipped_reason": tasks_skipped_reason,
    }


@router.delete("/projects/{project_id}", dependencies=[Depends(require("project.delete"))])
def delete_project(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    # Clean up task template CF values before cascade deletes the tasks
    task_ids = [str(row[0]) for row in db.query(TaskTemplate.ID).filter(TaskTemplate.PROJECT_ID == project_id).all()]
    if task_ids:
        db.query(CustomFieldTableValue).filter(
            CustomFieldTableValue.TABLE_NAME == "task_template",
            CustomFieldTableValue.TABLE_ROW_ID.in_(task_ids),
        ).delete(synchronize_session=False)
    # Clean up project pricing CF values before cascade deletes the 1:1 pricing row
    pricing_row = db.query(ProjectPricing).filter(ProjectPricing.PROJECT_ID == project_id).first()
    if pricing_row:
        db.query(CustomFieldTableValue).filter(
            CustomFieldTableValue.TABLE_NAME == "project_pricing",
            CustomFieldTableValue.TABLE_ROW_ID == str(pricing_row.ID),
        ).delete(synchronize_session=False)
    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "project",
        CustomFieldTableValue.TABLE_ROW_ID == str(project_id),
    ).delete(synchronize_session=False)
    try:
        db.delete(project)
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete project")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete project")
    return {"message": "Project deleted"}


# =========================
# PROJECT PRICING
# =========================

def _compute_final_price(original, packing, transport, installation, service, additional, tax, discount) -> float:
    return float(original or 0) + float(packing or 0) + float(transport or 0) + float(installation or 0) \
        + float(service or 0) + float(additional or 0) + float(tax or 0) - float(discount or 0)


def _push_pricing_to_quotation(db: Session, pricing: ProjectPricing) -> None:
    """Best-effort: whenever pricing is saved, push FINAL_PRICE into that
    project's quotation (if one exists) so the Amount column stays in sync
    automatically — never lets a sync failure affect the pricing save that
    already succeeded."""
    try:
        quotation = db.query(ProjectQuotationTemplate).filter(
            ProjectQuotationTemplate.PROJECT_ID == pricing.PROJECT_ID
        ).first()
        if not quotation:
            return
        company = get_company_settings(db, pricing.VENDOR_ID)
        if sync_final_price_into_quotation(quotation, pricing, company):
            db.commit()
    except Exception:
        db.rollback()
        log.warning("Could not sync pricing into quotation for project %s", pricing.PROJECT_ID, exc_info=True)


def _serialize_pricing(p: ProjectPricing, project_name: str = None, category_id: str = None) -> dict:
    return {
        "ID": p.ID,
        "PROJECT_ID": p.PROJECT_ID,
        "PROJECT_NAME": project_name,
        "CATEGORY_ID": category_id,
        "VENDOR_ID": p.VENDOR_ID,
        "CURRENCY": p.CURRENCY,
        "ORIGINAL_PRICE": float(p.ORIGINAL_PRICE) if p.ORIGINAL_PRICE is not None else None,
        "MINIMUM_NEGOTIATION_PRICE": float(p.MINIMUM_NEGOTIATION_PRICE) if p.MINIMUM_NEGOTIATION_PRICE is not None else None,
        "NEGOTIATION_PERCENT": float(p.NEGOTIATION_PERCENT) if p.NEGOTIATION_PERCENT is not None else None,
        "PACKING_CHARGE": float(p.PACKING_CHARGE) if p.PACKING_CHARGE is not None else None,
        "TRANSPORTATION_CHARGE": float(p.TRANSPORTATION_CHARGE) if p.TRANSPORTATION_CHARGE is not None else None,
        "INSTALLATION_CHARGE": float(p.INSTALLATION_CHARGE) if p.INSTALLATION_CHARGE is not None else None,
        "SERVICE_CHARGE": float(p.SERVICE_CHARGE) if p.SERVICE_CHARGE is not None else None,
        "ADDITIONAL_CHARGES": float(p.ADDITIONAL_CHARGES) if p.ADDITIONAL_CHARGES is not None else None,
        "TAX_AMOUNT": float(p.TAX_AMOUNT) if p.TAX_AMOUNT is not None else None,
        "DISCOUNT_AMOUNT": float(p.DISCOUNT_AMOUNT) if p.DISCOUNT_AMOUNT is not None else None,
        "FINAL_PRICE": float(p.FINAL_PRICE) if p.FINAL_PRICE is not None else None,
        "REMARKS": p.REMARKS,
        "IS_ACTIVE": p.IS_ACTIVE,
        "CREATED_AT": p.CREATED_AT.isoformat() if p.CREATED_AT else None,
        "UPDATED_AT": p.UPDATED_AT.isoformat() if p.UPDATED_AT else None,
    }


@router.get("/project-pricing", dependencies=[Depends(require("project.view", "project.pricing.view"))])
def list_project_pricing(
    project_id: Optional[str] = Query(None),
    vendor_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(ProjectPricing, Project).join(Project, ProjectPricing.PROJECT_ID == Project.ID)
    if vendor_id is not None:
        q = q.filter(ProjectPricing.VENDOR_ID == vendor_id)
    if project_id:
        q = q.filter(ProjectPricing.PROJECT_ID == project_id)
    if is_active is not None:
        q = q.filter(ProjectPricing.IS_ACTIVE == is_active)
    if search:
        term = f"%{search}%"
        q = q.filter(Project.NAME.ilike(term))
    rows = q.order_by(Project.NAME).all()
    return [_serialize_pricing(p, proj.NAME, proj.CATEGORY_ID) for p, proj in rows]


@router.get("/project-pricing/{pricing_id}", dependencies=[Depends(require("project.view", "project.pricing.view"))])
def get_project_pricing(pricing_id: str, db: Session = Depends(get_db)):
    p = db.query(ProjectPricing).filter(ProjectPricing.ID == pricing_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pricing not found")
    project = db.query(Project).filter(Project.ID == p.PROJECT_ID).first()
    return _serialize_pricing(p, project.NAME if project else None, project.CATEGORY_ID if project else None)


@router.post("/project-pricing", dependencies=[Depends(require("project.create", "project.pricing.create"))])
def create_project_pricing(data: ProjectPricingCreate, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == data.PROJECT_ID).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    existing = db.query(ProjectPricing).filter(ProjectPricing.PROJECT_ID == data.PROJECT_ID).first()
    if existing:
        raise HTTPException(status_code=400, detail="Pricing already exists for this project")

    if data.MINIMUM_NEGOTIATION_PRICE is not None and data.MINIMUM_NEGOTIATION_PRICE > data.ORIGINAL_PRICE:
        raise HTTPException(status_code=400, detail="Minimum Negotiation Price cannot exceed Original Price")

    pricing = ProjectPricing(
        PROJECT_ID=data.PROJECT_ID,
        VENDOR_ID=data.VENDOR_ID,
        CURRENCY=data.CURRENCY,
        ORIGINAL_PRICE=data.ORIGINAL_PRICE,
        MINIMUM_NEGOTIATION_PRICE=data.MINIMUM_NEGOTIATION_PRICE,
        NEGOTIATION_PERCENT=data.NEGOTIATION_PERCENT,
        PACKING_CHARGE=data.PACKING_CHARGE,
        TRANSPORTATION_CHARGE=data.TRANSPORTATION_CHARGE,
        INSTALLATION_CHARGE=data.INSTALLATION_CHARGE,
        SERVICE_CHARGE=data.SERVICE_CHARGE,
        ADDITIONAL_CHARGES=data.ADDITIONAL_CHARGES,
        TAX_AMOUNT=data.TAX_AMOUNT,
        DISCOUNT_AMOUNT=data.DISCOUNT_AMOUNT,
        REMARKS=data.REMARKS,
        IS_ACTIVE=data.IS_ACTIVE,
    )
    pricing.FINAL_PRICE = _compute_final_price(
        data.ORIGINAL_PRICE, data.PACKING_CHARGE, data.TRANSPORTATION_CHARGE, data.INSTALLATION_CHARGE,
        data.SERVICE_CHARGE, data.ADDITIONAL_CHARGES, data.TAX_AMOUNT, data.DISCOUNT_AMOUNT,
    )
    try:
        db.add(pricing)
        db.commit()
        db.refresh(pricing)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create project pricing")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create project pricing")
    _push_pricing_to_quotation(db, pricing)
    return {"message": "Pricing created", "ID": pricing.ID}


@router.put("/project-pricing/{pricing_id}", dependencies=[Depends(require("project.update", "project.pricing.update"))])
def update_project_pricing(pricing_id: str, data: ProjectPricingUpdate, db: Session = Depends(get_db)):
    pricing = db.query(ProjectPricing).filter(ProjectPricing.ID == pricing_id).first()
    if not pricing:
        raise HTTPException(status_code=404, detail="Pricing not found")

    for field in (
        "CURRENCY", "ORIGINAL_PRICE", "MINIMUM_NEGOTIATION_PRICE", "NEGOTIATION_PERCENT",
        "PACKING_CHARGE", "TRANSPORTATION_CHARGE", "INSTALLATION_CHARGE", "SERVICE_CHARGE",
        "ADDITIONAL_CHARGES", "TAX_AMOUNT", "DISCOUNT_AMOUNT", "REMARKS", "IS_ACTIVE",
    ):
        value = getattr(data, field)
        if value is not None:
            setattr(pricing, field, value)

    if pricing.MINIMUM_NEGOTIATION_PRICE is not None and pricing.MINIMUM_NEGOTIATION_PRICE > pricing.ORIGINAL_PRICE:
        raise HTTPException(status_code=400, detail="Minimum Negotiation Price cannot exceed Original Price")

    pricing.FINAL_PRICE = _compute_final_price(
        pricing.ORIGINAL_PRICE, pricing.PACKING_CHARGE, pricing.TRANSPORTATION_CHARGE, pricing.INSTALLATION_CHARGE,
        pricing.SERVICE_CHARGE, pricing.ADDITIONAL_CHARGES, pricing.TAX_AMOUNT, pricing.DISCOUNT_AMOUNT,
    )
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update project pricing")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update project pricing")
    _push_pricing_to_quotation(db, pricing)
    return {"message": "Pricing updated"}


@router.delete("/project-pricing/{pricing_id}", dependencies=[Depends(require("project.delete", "project.pricing.delete"))])
def delete_project_pricing(pricing_id: str, db: Session = Depends(get_db)):
    pricing = db.query(ProjectPricing).filter(ProjectPricing.ID == pricing_id).first()
    if not pricing:
        raise HTTPException(status_code=404, detail="Pricing not found")
    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "project_pricing",
        CustomFieldTableValue.TABLE_ROW_ID == str(pricing_id),
    ).delete(synchronize_session=False)
    try:
        db.delete(pricing)
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete project pricing")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete project pricing")
    return {"message": "Pricing deleted"}


_PRICING_STD_COLS = {
    "PROJECT NAME", "CURRENCY", "ORIGINAL PRICE", "MINIMUM NEGOTIATION PRICE", "NEGOTIATION PERCENTAGE",
    "PACKING CHARGE", "TRANSPORTATION CHARGE", "INSTALLATION CHARGE", "SERVICE CHARGE",
    "ADDITIONAL CHARGES", "TAX AMOUNT", "DISCOUNT AMOUNT", "REMARKS", "S.NO", "S.N", "SN", "",
}

_PRICING_NUMERIC_FIELDS = (
    ("ORIGINAL PRICE", "ORIGINAL_PRICE", True),
    ("MINIMUM NEGOTIATION PRICE", "MINIMUM_NEGOTIATION_PRICE", False),
    ("NEGOTIATION PERCENTAGE", "NEGOTIATION_PERCENT", False),
    ("PACKING CHARGE", "PACKING_CHARGE", False),
    ("TRANSPORTATION CHARGE", "TRANSPORTATION_CHARGE", False),
    ("INSTALLATION CHARGE", "INSTALLATION_CHARGE", False),
    ("SERVICE CHARGE", "SERVICE_CHARGE", False),
    ("ADDITIONAL CHARGES", "ADDITIONAL_CHARGES", False),
    ("TAX AMOUNT", "TAX_AMOUNT", False),
    ("DISCOUNT AMOUNT", "DISCOUNT_AMOUNT", False),
)

# Columns that are genuinely nullable on ProjectPricing — every other numeric
# field is NOT NULL with a default of 0, so a blank cell must become 0, not
# None (an explicit None would override the column default and violate the
# NOT NULL constraint on commit).
_PRICING_NULLABLE_FIELDS = {"MINIMUM_NEGOTIATION_PRICE", "NEGOTIATION_PERCENT"}


@router.post("/project-pricing/bulk-upload", dependencies=[Depends(require("project.create", "project.pricing.create", "project.pricing.import"))])
async def bulk_upload_project_pricing(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "Pricing")

    cf_fields = _cf_fields_for_table("project_pricing", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols = [h for h in headers if h not in _PRICING_STD_COLS and h in cf_by_upper]

    inserted = updated = skipped = 0
    errors: List[dict] = []
    synced_pricings: List[ProjectPricing] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        project_name = _cell(record, "PROJECT NAME")
        if not project_name:
            errors.append({"row": row_num, "field": "Project Name", "message": "Project Name is required"})
            continue

        project = db.query(Project).filter(
            Project.VENDOR_ID == vendor_id, Project.NAME == project_name
        ).first()
        if not project:
            errors.append({"row": row_num, "field": "Project Name", "message": f"Project '{project_name}' not found"})
            continue

        currency = _cell(record, "CURRENCY") or "INR"
        remarks = _cell(record, "REMARKS") or None

        numeric_vals = {}
        row_error = False
        for col, field, required in _PRICING_NUMERIC_FIELDS:
            raw_val = _cell(record, col)
            if not raw_val:
                if required:
                    errors.append({"row": row_num, "field": col.title(), "message": f"{col.title()} is required"})
                    row_error = True
                numeric_vals[field] = None if field in _PRICING_NULLABLE_FIELDS else 0
                continue
            try:
                numeric_vals[field] = float(raw_val)
            except ValueError:
                errors.append({"row": row_num, "field": col.title(), "message": "Must be a number"})
                row_error = True
        if row_error:
            continue

        original_price = numeric_vals["ORIGINAL_PRICE"] or 0
        min_negotiation = numeric_vals["MINIMUM_NEGOTIATION_PRICE"]
        if min_negotiation is not None and min_negotiation > original_price:
            errors.append({
                "row": row_num, "field": "Minimum Negotiation Price",
                "message": "Minimum Negotiation Price cannot exceed Original Price",
            })
            continue

        # Validate CFs (required + type)
        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({
                    "row": row_num, "field": cf_f.FIELD_NAME,
                    "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing',
                })
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        final_price = _compute_final_price(
            original_price, numeric_vals["PACKING_CHARGE"], numeric_vals["TRANSPORTATION_CHARGE"],
            numeric_vals["INSTALLATION_CHARGE"], numeric_vals["SERVICE_CHARGE"], numeric_vals["ADDITIONAL_CHARGES"],
            numeric_vals["TAX_AMOUNT"], numeric_vals["DISCOUNT_AMOUNT"],
        )

        existing = db.query(ProjectPricing).filter(ProjectPricing.PROJECT_ID == project.ID).first()
        if existing:
            existing.CURRENCY = currency
            existing.REMARKS = remarks
            for field in numeric_vals:
                setattr(existing, field, numeric_vals[field])
            existing.FINAL_PRICE = final_price
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(existing.ID, "project_pricing", cf_id, val, db)
            updated += 1
            synced_pricings.append(existing)
        else:
            new_pricing = ProjectPricing(
                PROJECT_ID=project.ID, VENDOR_ID=vendor_id, CURRENCY=currency, REMARKS=remarks,
                **numeric_vals,
            )
            new_pricing.FINAL_PRICE = final_price
            db.add(new_pricing)
            db.flush()
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(new_pricing.ID, "project_pricing", cf_id, val, db)
            inserted += 1
            synced_pricings.append(new_pricing)

    db.commit()
    for p in synced_pricings:
        _push_pricing_to_quotation(db, p)
    total = inserted + updated + skipped + len(errors)
    msg = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {
        "message": msg,
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": total, "errors": errors,
    }


# =========================
# BOM PARSE
# =========================

@router.post("/projects/parse-bom", dependencies=[Depends(require("project.view"))])
async def parse_bom(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
):
    content = await file.read()
    filename = file.filename or ""

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        rows_raw = list(reader)
        name_col = next(
            (c for c in (rows_raw[0].keys() if rows_raw else [])
             if c.strip().upper() in ("ASSEMBLY", "NAME", "ITEM", "PART NAME", "DESCRIPTION")),
            None
        )
        rows = [
            {"name": str(r.get(name_col, "")).strip(), "sequence": i}
            for i, r in enumerate(rows_raw, 1)
            if str(r.get(name_col, "")).strip()
        ]
        return {"sheets": None, "rows": rows}

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets = wb.sheetnames

    if not sheet_name and len(sheets) > 1:
        return {"sheets": sheets, "rows": []}

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = []
    rows = []
    name_col_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper() if c else "" for c in row]
            for idx, h in enumerate(headers):
                if h in ("ASSEMBLY", "NAME", "ITEM", "PART NAME", "DESCRIPTION"):
                    name_col_idx = idx
                    break
            if name_col_idx is None and headers:
                name_col_idx = 0
            continue
        if all(c is None for c in row):
            continue
        if name_col_idx is not None and name_col_idx < len(row):
            val = str(row[name_col_idx]).strip() if row[name_col_idx] else ""
            if val:
                rows.append({"name": val, "sequence": len(rows)})
    return {"sheets": sheets, "rows": rows}


# =========================
# TASK TEMPLATES
# =========================

@router.get("/task-templates", dependencies=[Depends(require("project.view", "project.task_templates.view"))])
def list_task_templates(
    project_id: Optional[str] = Query(None),
    vendor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    q = db.query(TaskTemplate)
    if project_id:
        q = q.filter(TaskTemplate.PROJECT_ID == project_id)
    if vendor_id is not None:
        q = q.filter(TaskTemplate.VENDOR_ID == vendor_id)
    tasks = q.order_by(TaskTemplate.SEQUENCE_NUMBER).all()
    return _enrich_tasks(tasks, db)


@router.post("/task-templates", dependencies=[Depends(require("project.create", "project.task_templates.create"))])
def create_task_template(data: TaskTemplateCreate, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == data.PROJECT_ID).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _validate_unique_sequence(db, data.PROJECT_ID, data.SEQUENCE_NUMBER)
    task = TaskTemplate(
        PROJECT_ID=data.PROJECT_ID,
        NAME=data.NAME,
        DESCRIPTION=data.DESCRIPTION,
        DURATION_VALUE=data.DURATION_VALUE,
        DURATION_UNIT=data.DURATION_UNIT,
        SEQUENCE_NUMBER=data.SEQUENCE_NUMBER,
        TASK_SCOPE=data.TASK_SCOPE or "PROJECT",
        VENDOR_ID=data.VENDOR_ID
    )
    db.add(task)
    db.flush()
    _create_requirement_rows(task.ID, data.requirements, db)
    calculate_project_estimated_duration(db, project)
    try:
        db.commit()
        db.refresh(task)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create task template")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create task template")
    return {"message": "Task created", "ID": task.ID}


@router.post("/task-templates/bulk-create", dependencies=[Depends(require("project.create", "project.task_templates.create"))])
def bulk_create_task_templates():
    raise HTTPException(status_code=501, detail="Use POST /projects with embedded tasks instead")


@router.put("/task-templates/{task_id}", dependencies=[Depends(require("project.update", "project.task_templates.update"))])
def update_task_template(task_id: str, data: TaskTemplateUpdate, db: Session = Depends(get_db)):
    task = db.query(TaskTemplate).filter(TaskTemplate.ID == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if data.NAME is not None:
        task.NAME = data.NAME
    if data.DESCRIPTION is not None:
        task.DESCRIPTION = data.DESCRIPTION
    if data.DURATION_VALUE is not None:
        task.DURATION_VALUE = data.DURATION_VALUE
    if data.DURATION_UNIT is not None:
        task.DURATION_UNIT = data.DURATION_UNIT
    if data.SEQUENCE_NUMBER is not None and data.SEQUENCE_NUMBER != task.SEQUENCE_NUMBER:
        _validate_unique_sequence(db, task.PROJECT_ID, data.SEQUENCE_NUMBER, exclude_task_id=task.ID)
        task.SEQUENCE_NUMBER = data.SEQUENCE_NUMBER
    if data.TASK_SCOPE is not None:
        task.TASK_SCOPE = data.TASK_SCOPE
    if data.requirements is not None:
        db.query(TaskTemplateRequirement).filter(TaskTemplateRequirement.TASK_TEMPLATE_ID == task.ID).delete()
        db.flush()
        _create_requirement_rows(task.ID, data.requirements, db)
    db.flush()
    project = db.query(Project).filter(Project.ID == task.PROJECT_ID).first()
    if project:
        calculate_project_estimated_duration(db, project)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update task template")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update task template")
    return {"message": "Task updated"}


@router.delete("/task-templates/{task_id}", dependencies=[Depends(require("project.delete", "project.task_templates.delete"))])
def delete_task_template(task_id: str, db: Session = Depends(get_db)):
    task = db.query(TaskTemplate).filter(TaskTemplate.ID == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    # CustomerProjectTask.TASK_TEMPLATE_ID has an ON DELETE RESTRICT FK —
    # deleting a task that already has real, assigned production work
    # against it would otherwise crash with an unhandled IntegrityError
    # (500). See _project_has_live_production_tasks()'s own docstring for
    # the full explanation (same underlying issue as update_project()'s
    # task/group replacement guard).
    if db.query(CustomerProjectTask.ID).filter(CustomerProjectTask.TASK_TEMPLATE_ID == task_id).first():
        raise HTTPException(
            status_code=400,
            detail=f"'{task.NAME}' already has production tasks generated and assigned for at least one "
                   "customer order, so it can no longer be deleted.",
        )
    project_id = task.PROJECT_ID
    db.query(CustomFieldTableValue).filter(
        CustomFieldTableValue.TABLE_NAME == "task_template",
        CustomFieldTableValue.TABLE_ROW_ID == str(task_id),
    ).delete(synchronize_session=False)
    db.delete(task)
    db.flush()
    project = db.query(Project).filter(Project.ID == project_id).first()
    if project:
        calculate_project_estimated_duration(db, project)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete task template")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete task template")
    return {"message": "Task deleted"}


@router.patch("/task-templates/reorder", dependencies=[Depends(require("project.update", "project.task_templates.update", "project.task_templates.reorder"))])
def reorder_tasks(items: List[ReorderItem], db: Session = Depends(get_db)):
    for item in items:
        task = db.query(TaskTemplate).filter(TaskTemplate.ID == item.id).first()
        if task:
            task.SEQUENCE_NUMBER = item.sequence_number
    db.commit()
    return {"message": "Tasks reordered"}


# =========================
# TASK GROUPS
# =========================
# Standalone CRUD used both by the /projects row-action "Group" modal and,
# for edits on an already-created project, by the wizard's Task Group step.
# Group/dependency configuration is exclusively managed here — never via
# the Task Template endpoints above.

def _validate_group_membership(db: Session, project_id: str, member_ids: list, exclude_group_id: str = None):
    """A task must exist, belong to this SAME project, and not already
    belong to a DIFFERENT group — the core "one group per task" rule."""
    if not member_ids:
        return
    members = db.query(TaskTemplate).filter(TaskTemplate.ID.in_(member_ids)).all()
    found_ids = {t.ID for t in members}
    missing = set(member_ids) - found_ids
    if missing:
        raise HTTPException(status_code=400, detail="One or more selected tasks were not found.")
    wrong_project = [t for t in members if t.PROJECT_ID != project_id]
    if wrong_project:
        raise HTTPException(
            status_code=400,
            detail="Tasks from another project cannot be assigned to this project's group.",
        )
    already_grouped = [
        t for t in members
        if t.TASK_GROUP_ID and t.TASK_GROUP_ID != exclude_group_id
    ]
    if already_grouped:
        names = ", ".join(t.NAME for t in already_grouped)
        raise HTTPException(
            status_code=400,
            detail=f"These tasks already belong to another group: {names}.",
        )


@router.get("/projects/{project_id}/task-groups", dependencies=[Depends(require("project.view", "project.task_groups.view"))])
def list_task_groups(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    groups = (
        db.query(TaskGroup)
          .filter(TaskGroup.PROJECT_ID == project_id)
          .order_by(TaskGroup.SEQUENCE_NUMBER)
          .all()
    )
    return _enrich_task_groups(groups, db)


@router.post("/projects/{project_id}/task-groups", dependencies=[Depends(require("project.create", "project.task_groups.create"))])
def create_task_group(project_id: str, data: TaskGroupCreate, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    member_ids = list(dict.fromkeys(data.task_template_ids))  # de-dupe, preserve order
    _validate_group_membership(db, project_id, member_ids)

    rule = _normalize_dependency_rule(data.DEPENDENCY_RULE)
    depends_on_id = _resolve_group_dependency(rule, data.DEPENDS_ON_TASK_TEMPLATE_ID, member_ids)
    existing_count = db.query(TaskGroup).filter(TaskGroup.PROJECT_ID == project_id).count()

    group = TaskGroup(
        PROJECT_ID=project_id,
        VENDOR_ID=project.VENDOR_ID,
        NAME=data.NAME or None,
        DEPENDENCY_RULE=rule,
        DEPENDS_ON_TASK_TEMPLATE_ID=depends_on_id,
        SEQUENCE_NUMBER=existing_count,
    )
    db.add(group)
    db.flush()

    if member_ids:
        db.query(TaskTemplate).filter(TaskTemplate.ID.in_(member_ids)).update(
            {"TASK_GROUP_ID": group.ID}, synchronize_session=False
        )
        db.flush()

    calculate_project_estimated_duration(db, project)

    try:
        db.commit()
        db.refresh(group)
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "create task group")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "create task group")

    return {"message": "Task group created", "ID": group.ID}


@router.put("/projects/{project_id}/task-groups/{group_id}", dependencies=[Depends(require("project.update", "project.task_groups.update"))])
def update_task_group(project_id: str, group_id: str, data: TaskGroupUpdate, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    group = db.query(TaskGroup).filter(TaskGroup.ID == group_id, TaskGroup.PROJECT_ID == project_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Task group not found")

    if data.NAME is not None:
        group.NAME = data.NAME or None

    rule = _normalize_dependency_rule(data.DEPENDENCY_RULE) if data.DEPENDENCY_RULE is not None else group.DEPENDENCY_RULE

    if data.task_template_ids is not None:
        member_ids = list(dict.fromkeys(data.task_template_ids))
        _validate_group_membership(db, project_id, member_ids, exclude_group_id=group_id)

        # Full-replace membership: ungroup this group's current members,
        # then (re)assign the submitted list. A task in both the old and
        # new list is simply reassigned to the same group — a no-op.
        db.query(TaskTemplate).filter(TaskTemplate.TASK_GROUP_ID == group_id).update(
            {"TASK_GROUP_ID": None}, synchronize_session=False
        )
        if member_ids:
            db.query(TaskTemplate).filter(TaskTemplate.ID.in_(member_ids)).update(
                {"TASK_GROUP_ID": group_id}, synchronize_session=False
            )
        db.flush()
    else:
        member_ids = [
            t.ID for t in db.query(TaskTemplate).filter(TaskTemplate.TASK_GROUP_ID == group_id).all()
        ]

    # Always re-resolve from the EFFECTIVE (new-if-provided, else existing)
    # rule/target/membership together, rather than patching fields
    # independently — e.g. switching rule away from ONE must clear a
    # stale target even if DEPENDS_ON_TASK_TEMPLATE_ID wasn't resubmitted,
    # and _resolve_group_dependency already does exactly that for ALL/ANY.
    effective_depends_on = (
        data.DEPENDS_ON_TASK_TEMPLATE_ID if data.DEPENDS_ON_TASK_TEMPLATE_ID is not None
        else group.DEPENDS_ON_TASK_TEMPLATE_ID
    )
    group.DEPENDENCY_RULE = rule
    group.DEPENDS_ON_TASK_TEMPLATE_ID = _resolve_group_dependency(rule, effective_depends_on, member_ids)

    db.flush()
    calculate_project_estimated_duration(db, project)

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "update task group")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "update task group")

    return {"message": "Task group updated"}


@router.delete("/projects/{project_id}/task-groups/{group_id}", dependencies=[Depends(require("project.delete", "project.task_groups.delete"))])
def delete_task_group(project_id: str, group_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.ID == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    group = db.query(TaskGroup).filter(TaskGroup.ID == group_id, TaskGroup.PROJECT_ID == project_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Task group not found")

    # Explicit ungroup (rather than relying solely on the DB's own
    # ON DELETE SET NULL) so this same session's subsequent duration
    # recalculation reads fresh TASK_GROUP_ID values instead of any
    # stale identity-mapped TaskTemplate objects. Task templates
    # themselves are never deleted here — only their group assignment
    # is cleared.
    db.query(TaskTemplate).filter(TaskTemplate.TASK_GROUP_ID == group_id).update(
        {"TASK_GROUP_ID": None}, synchronize_session=False
    )
    db.delete(group)
    db.flush()
    calculate_project_estimated_duration(db, project)

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise_db_error(e, "delete task group")
    except Exception as e:
        db.rollback()
        raise_db_error(e, "delete task group")

    return {"message": "Task group deleted"}


# =========================
# BULK UPLOAD — SHARED HELPERS
# =========================

def _cf_fields_for_table(table_name: str, vendor_id: int, db: Session):
    """Return CustomField rows for a table, sorted by SORT_ORDER."""
    return (
        db.query(CustomField)
          .filter(CustomField.TABLE_NAME == table_name, CustomField.VENDOR_ID == vendor_id)
          .order_by(CustomField.SORT_ORDER, CustomField.FIELD_NAME)
          .all()
    )


def _upsert_cf_bulk(row_id: str, table_name: str, cf_field_id: str, value, db: Session):
    """Insert or update a single custom-field value row."""
    stored = value if value else None
    existing = (
        db.query(CustomFieldTableValue)
          .filter(
              CustomFieldTableValue.TABLE_NAME == table_name,
              CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
              CustomFieldTableValue.CUSTOM_FIELD_ID == cf_field_id,
          )
          .first()
    )
    if existing:
        existing.CUSTOM_FIELD_VALUE = stored
    elif stored is not None:
        db.add(CustomFieldTableValue(
            TABLE_NAME=table_name,
            TABLE_ROW_ID=str(row_id),
            CUSTOM_FIELD_ID=cf_field_id,
            CUSTOM_FIELD_VALUE=stored,
        ))


def _validate_cf_value(field, raw_val) -> Optional[str]:
    """Validate a raw bulk-upload value against the field's type and options.
    Returns an error message string if invalid, or None if valid/empty."""
    import re as _re
    from datetime import date as _d, datetime as _dt

    if raw_val is None or str(raw_val).strip() == "":
        return None  # emptiness is handled by the required check

    val = str(raw_val).strip()
    ft  = field.FIELD_TYPE

    if ft == "NUMBER":
        try:
            float(val)
        except ValueError:
            return "Must be a number"

    elif ft == "EMAIL":
        if not _re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", val):
            return "Must be a valid email address (e.g. user@example.com)"

    elif ft == "PHONE":
        if not _re.fullmatch(r"\+?[\d\s\-().]{7,20}", val):
            return "Must be a valid phone number"

    elif ft == "DATE":
        try:
            _d.fromisoformat(val)
        except ValueError:
            return "Must be a valid date (YYYY-MM-DD)"

    elif ft == "DATETIME":
        try:
            _dt.fromisoformat(val)
        except ValueError:
            return "Must be a valid date/time (YYYY-MM-DDTHH:MM)"

    elif ft in ("SELECT", "RADIO"):
        allowed = field.OPTIONS or []
        if allowed and val not in allowed:
            return f'Invalid option "{val}". Allowed values: {", ".join(allowed)}'

    elif ft == "CHECKBOX":
        allowed = set(field.OPTIONS or [])
        if allowed:
            items = [v.strip() for v in val.split(",") if v.strip()] if isinstance(raw_val, str) else [val]
            bad = [v for v in items if v not in allowed]
            if bad:
                return f'Invalid option(s): {", ".join(bad)}. Allowed: {", ".join(field.OPTIONS or [])}'

    return None


def _cf_row_changed(row_id: str, table_name: str, cf_vals_by_id: dict, db: Session) -> bool:
    """Return True if any value in cf_vals_by_id {cf_id: new_val} differs from the stored value."""
    for cf_id, new_val in cf_vals_by_id.items():
        row = (
            db.query(CustomFieldTableValue)
              .filter(
                  CustomFieldTableValue.TABLE_NAME == table_name,
                  CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
                  CustomFieldTableValue.CUSTOM_FIELD_ID == cf_id,
              )
              .first()
        )
        old_s = str(row.CUSTOM_FIELD_VALUE) if (row and row.CUSTOM_FIELD_VALUE is not None) else ""
        new_s = str(new_val) if new_val else ""
        if old_s != new_s:
            return True
    return False


def _parse_bulk_xl(content: bytes, required_sheet: str):
    """Parse an Excel workbook for bulk upload.

    Requires the workbook to contain a sheet named exactly `required_sheet`
    (case-sensitive). Raises HTTPException 400 if the sheet is absent.
    Returns (headers_upper, data_rows).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if required_sheet not in wb.sheetnames:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise HTTPException(
            status_code=400,
            detail=f'Sheet "{required_sheet}" not found in the uploaded file. '
                   f'Available sheets: {available}. '
                   f'Please use the Template download to get a correctly named workbook.',
        )
    ws = wb[required_sheet]
    headers: Optional[List[str]] = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper() if c is not None else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(row)
    return headers, rows


def _cell(record: dict, *keys) -> str:
    """Extract and strip a value from an upper-cased record dict."""
    for k in keys:
        v = record.get(k.upper())
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return ""


# =========================
# PROJECT CATEGORIES BULK UPLOAD
# =========================

_CAT_STD_COLS = {"CATEGORY NAME", "DESCRIPTION", "S.NO", "S.N", "SN", ""}


@router.post("/project-categories/bulk-upload", dependencies=[Depends(require("project.create", "project.categories.create", "project.categories.import"))])
async def bulk_upload_categories(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "Categories")

    cf_fields = _cf_fields_for_table("project_category", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols = [h for h in headers if h not in _CAT_STD_COLS and h in cf_by_upper]

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        cat_name = _cell(record, "CATEGORY NAME")
        desc     = _cell(record, "DESCRIPTION") or None

        if not cat_name:
            errors.append({"row": row_num, "field": "Category Name", "message": "Category Name is required"})
            continue

        # Validate CFs (required + type)
        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({
                    "row": row_num, "field": cf_f.FIELD_NAME,
                    "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing',
                })
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        existing = (
            db.query(ProjectCategory)
              .filter(ProjectCategory.VENDOR_ID == vendor_id, ProjectCategory.NAME == cat_name)
              .first()
        )

        if existing:
            row_changed = (desc or "") != (existing.DESCRIPTION or "")
            if row_changed or _cf_row_changed(existing.ID, "project_category", cf_vals, db):
                if row_changed:
                    existing.DESCRIPTION = desc
                for cf_id, val in cf_vals.items():
                    _upsert_cf_bulk(existing.ID, "project_category", cf_id, val, db)
                updated += 1
            else:
                skipped += 1
        else:
            new_cat = ProjectCategory(NAME=cat_name, DESCRIPTION=desc, VENDOR_ID=vendor_id)
            db.add(new_cat)
            db.flush()
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(new_cat.ID, "project_category", cf_id, val, db)
            inserted += 1

    db.commit()
    total = inserted + updated + skipped + len(errors)
    msg   = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {
        "message": msg,
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": total, "errors": errors,
    }


# =========================
# PROJECTS BULK UPLOAD
# =========================

_PROJ_STD_COLS = {"CATEGORY NAME", "PROJECT NAME", "DESCRIPTION", "S.NO", "S.N", "SN", ""}


@router.post("/projects/bulk-upload", dependencies=[Depends(require("project.create"))])
async def bulk_upload_projects(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "Projects")

    cf_fields  = _cf_fields_for_table("project", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols    = [h for h in headers if h not in _PROJ_STD_COLS and h in cf_by_upper]

    # Pre-build category lookup map (case-insensitive)
    cat_name_map = {
        c.NAME.strip().lower(): c
        for c in db.query(ProjectCategory).filter(ProjectCategory.VENDOR_ID == vendor_id).all()
    }

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        cat_name  = _cell(record, "CATEGORY NAME")
        proj_name = _cell(record, "PROJECT NAME")
        desc      = _cell(record, "DESCRIPTION") or None

        if not proj_name:
            errors.append({"row": row_num, "field": "Project Name", "message": "Project Name is required"})
            continue
        if not cat_name:
            errors.append({"row": row_num, "field": "Category Name", "message": "Category Name is required"})
            continue

        cat = cat_name_map.get(cat_name.lower())
        if not cat:
            errors.append({
                "row": row_num, "field": "Category Name",
                "message": f'Category "{cat_name}" not found',
            })
            continue

        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({
                    "row": row_num, "field": cf_f.FIELD_NAME,
                    "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing',
                })
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        existing = (
            db.query(Project)
              .filter(
                  Project.VENDOR_ID == vendor_id,
                  Project.CATEGORY_ID == cat.ID,
                  Project.NAME == proj_name,
              )
              .first()
        )

        if existing:
            row_changed = (
                (desc or "") != (existing.DESCRIPTION or "") or
                cat.ID != existing.CATEGORY_ID
            )
            if row_changed or _cf_row_changed(existing.ID, "project", cf_vals, db):
                if row_changed:
                    existing.DESCRIPTION = desc
                    existing.CATEGORY_ID = cat.ID
                for cf_id, val in cf_vals.items():
                    _upsert_cf_bulk(existing.ID, "project", cf_id, val, db)
                updated += 1
            else:
                skipped += 1
        else:
            new_proj = Project(
                CATEGORY_ID=cat.ID,
                NAME=proj_name,
                DESCRIPTION=desc,
                BOM_MODE="MANUAL",
                ESTIMATED_TOTAL_DAYS=0.0,
                VENDOR_ID=vendor_id,
            )
            db.add(new_proj)
            db.flush()
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(new_proj.ID, "project", cf_id, val, db)
            inserted += 1

    db.commit()
    total = inserted + updated + skipped + len(errors)
    msg   = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {
        "message": msg,
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": total, "errors": errors,
    }


# =========================
# TASK TEMPLATES BULK UPLOAD
# =========================

_TASK_STD_COLS = {
    "PROJECT NAME", "TASK NAME", "DESCRIPTION",
    "DURATION VALUE", "DURATION UNIT", "DEPARTMENT", "ROLE", "SEQUENCE",
    "EXPERIENCE LEVEL", "REQUIRED COUNT",
    "S.NO", "S.N", "SN", "",
}
_VALID_DUR_UNITS = {"HOURS", "DAYS", "WEEKS", "MONTHS", "YEARS"}
_VALID_EXPERIENCE_LEVELS = {"FRESHER", "INTERMEDIATE", "EXPERIENCED"}


@router.post("/task-templates/bulk-upload", dependencies=[Depends(require("project.create", "project.task_templates.create", "project.task_templates.import"))])
async def bulk_upload_task_templates(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "Tasks")

    cf_fields   = _cf_fields_for_table("task_template", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols     = [h for h in headers if h not in _TASK_STD_COLS and h in cf_by_upper]

    # Pre-build lookup maps
    proj_name_map = {
        p.NAME.strip().lower(): p
        for p in db.query(Project).filter(Project.VENDOR_ID == vendor_id).all()
    }
    dept_name_map = {
        d.NAME.strip().lower(): d
        for d in db.query(Department).filter(Department.VENDOR_ID == vendor_id).all()
    }
    role_name_map = {
        r.NAME.strip().lower(): r
        for r in db.query(Role).filter(Role.VENDOR_ID == vendor_id).all()
    }

    inserted = updated = skipped = 0
    errors: List[dict] = []
    modified_proj_ids: set = set()
    # Lazy sequence counter per project for auto-assigned sequences
    proj_next_seq: dict = {}

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}

        proj_name = _cell(record, "PROJECT NAME")
        task_name = _cell(record, "TASK NAME")
        desc      = _cell(record, "DESCRIPTION") or None
        dur_val_s = _cell(record, "DURATION VALUE")
        dur_unit  = _cell(record, "DURATION UNIT").upper() or "DAYS"
        dept_name = _cell(record, "DEPARTMENT")
        role_name = _cell(record, "ROLE")
        seq_s     = _cell(record, "SEQUENCE")
        exp_raw   = _cell(record, "EXPERIENCE LEVEL")
        req_cnt_s = _cell(record, "REQUIRED COUNT")

        # Required field checks
        if not proj_name:
            errors.append({"row": row_num, "field": "Project Name", "message": "Project Name is required"})
            continue
        if not task_name:
            errors.append({"row": row_num, "field": "Task Name", "message": "Task Name is required"})
            continue

        proj = proj_name_map.get(proj_name.lower())
        if not proj:
            errors.append({
                "row": row_num, "field": "Project Name",
                "message": f'Project "{proj_name}" not found',
            })
            continue

        # Duration value
        try:
            dur_val = float(dur_val_s) if dur_val_s else 1.0
            if dur_val <= 0:
                raise ValueError
        except ValueError:
            errors.append({
                "row": row_num, "field": "Duration Value",
                "message": f'Invalid duration value "{dur_val_s}" — must be a positive number',
            })
            continue

        if dur_unit not in _VALID_DUR_UNITS:
            dur_unit = "DAYS"

        # Department (optional; error if provided but not found)
        dept_id = None
        if dept_name:
            d = dept_name_map.get(dept_name.lower())
            if not d:
                errors.append({
                    "row": row_num, "field": "Department",
                    "message": f'Department "{dept_name}" does not exist',
                })
                continue
            dept_id = d.ID

        # Role (optional; error if provided but not found)
        role_id = None
        if role_name:
            r = role_name_map.get(role_name.lower())
            if not r:
                errors.append({
                    "row": row_num, "field": "Role",
                    "message": f'Role "{role_name}" does not exist',
                })
                continue
            role_id = r.ID

        # Manpower requirement (optional, single "primary" requirement per
        # row, matching this sheet's existing one-row-per-task shape — a
        # task with several requirements is configured via the UI, and
        # bulk upload never touches requirements it didn't ask about).
        # Only managed when the row actually says something about it, so
        # re-uploading a sheet that only changed e.g. Description never
        # silently wipes a UI-configured multi-requirement task.
        has_requirement_signal = bool(dept_name or role_name or exp_raw or req_cnt_s)
        exp_level = exp_raw.upper() if exp_raw.upper() in _VALID_EXPERIENCE_LEVELS else "INTERMEDIATE"
        try:
            req_count = int(float(req_cnt_s)) if req_cnt_s else 1
            if req_count < 1:
                req_count = 1
        except ValueError:
            req_count = 1

        # Sequence
        seq: Optional[int] = None
        if seq_s:
            try:
                seq = int(float(seq_s))
            except ValueError:
                pass
        if seq is None:
            # Auto-assign: start from current task count for this project
            if proj.ID not in proj_next_seq:
                count = db.query(TaskTemplate).filter(TaskTemplate.PROJECT_ID == proj.ID).count()
                proj_next_seq[proj.ID] = count
            seq = proj_next_seq[proj.ID]
            proj_next_seq[proj.ID] += 1

        # Custom fields (required + type validation)
        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            val  = _cell(record, col) or None
            if cf_f.IS_REQUIRED and not val:
                errors.append({
                    "row": row_num, "field": cf_f.FIELD_NAME,
                    "message": f'Required custom field "{cf_f.FIELD_NAME}" is missing',
                })
                cf_error = True
            elif val:
                type_err = _validate_cf_value(cf_f, val)
                if type_err:
                    errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": type_err})
                    cf_error = True
            cf_vals[cf_f.ID] = val
        if cf_error:
            continue

        existing = (
            db.query(TaskTemplate)
              .filter(
                  TaskTemplate.PROJECT_ID == proj.ID,
                  TaskTemplate.NAME == task_name,
              )
              .first()
        )

        if existing:
            existing_req = (
                db.query(TaskTemplateRequirement)
                  .filter(TaskTemplateRequirement.TASK_TEMPLATE_ID == existing.ID)
                  .order_by(TaskTemplateRequirement.CREATED_AT.asc())
                  .first()
                if has_requirement_signal else None
            )
            requirement_changed = has_requirement_signal and (
                not existing_req
                or dept_id != existing_req.DEPARTMENT_ID
                or role_id != existing_req.ROLE_ID
                or exp_level != existing_req.EXPERIENCE_LEVEL
                or req_count != existing_req.REQUIRED_COUNT
            )
            row_changed = (
                (desc or "") != (existing.DESCRIPTION or "")
                or abs(float(dur_val) - float(existing.DURATION_VALUE or 1.0)) > 0.001
                or dur_unit != existing.DURATION_UNIT
            )
            if row_changed or requirement_changed or _cf_row_changed(existing.ID, "task_template", cf_vals, db):
                if row_changed:
                    existing.DESCRIPTION = desc
                    existing.DURATION_VALUE = dur_val
                    existing.DURATION_UNIT  = dur_unit
                if requirement_changed:
                    if existing_req:
                        existing_req.DEPARTMENT_ID    = dept_id
                        existing_req.ROLE_ID           = role_id
                        existing_req.EXPERIENCE_LEVEL  = exp_level
                        existing_req.REQUIRED_COUNT    = req_count
                    else:
                        db.add(TaskTemplateRequirement(
                            TASK_TEMPLATE_ID=existing.ID,
                            DEPARTMENT_ID=dept_id,
                            ROLE_ID=role_id,
                            EXPERIENCE_LEVEL=exp_level,
                            REQUIRED_COUNT=req_count,
                        ))
                for cf_id, val in cf_vals.items():
                    _upsert_cf_bulk(existing.ID, "task_template", cf_id, val, db)
                modified_proj_ids.add(proj.ID)
                updated += 1
            else:
                skipped += 1
        else:
            new_task = TaskTemplate(
                PROJECT_ID      = proj.ID,
                NAME            = task_name,
                DESCRIPTION     = desc,
                DURATION_VALUE  = dur_val,
                DURATION_UNIT   = dur_unit,
                SEQUENCE_NUMBER = seq,
                VENDOR_ID       = vendor_id,
            )
            db.add(new_task)
            db.flush()
            if has_requirement_signal:
                db.add(TaskTemplateRequirement(
                    TASK_TEMPLATE_ID=new_task.ID,
                    DEPARTMENT_ID=dept_id,
                    ROLE_ID=role_id,
                    EXPERIENCE_LEVEL=exp_level,
                    REQUIRED_COUNT=req_count,
                ))
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(new_task.ID, "task_template", cf_id, val, db)
            modified_proj_ids.add(proj.ID)
            inserted += 1

    # Recalculate estimated duration for every touched project
    for pid in modified_proj_ids:
        proj_obj = db.query(Project).filter(Project.ID == pid).first()
        if proj_obj:
            calculate_project_estimated_duration(db, proj_obj)

    db.commit()
    total = inserted + updated + skipped + len(errors)
    msg   = f"Upload complete: {inserted} inserted, {updated} updated, {skipped} skipped"
    if errors:
        msg += f", {len(errors)} error(s)"
    return {
        "message": msg,
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": total, "errors": errors,
    }


# =========================
# SEED
# =========================

@router.post("/seed-project-templates", dependencies=[Depends(get_current_admin)])
def seed_templates(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    from app.services.seed_data import PROJECT_TEMPLATE_CATALOG
    created_cats = 0
    created_projects = 0
    for entry in PROJECT_TEMPLATE_CATALOG:
        cat_name = entry.get("category")
        template_name = entry.get("name")
        if not cat_name or not template_name:
            continue
        cat = db.query(ProjectCategory).filter(
            ProjectCategory.VENDOR_ID == vendor_id,
            ProjectCategory.NAME == cat_name
        ).first()
        if not cat:
            cat = ProjectCategory(NAME=cat_name, VENDOR_ID=vendor_id)
            db.add(cat)
            db.flush()
            created_cats += 1
        existing = db.query(Project).filter(
            Project.VENDOR_ID == vendor_id,
            Project.CATEGORY_ID == cat.ID,
            Project.NAME == template_name
        ).first()
        if not existing:
            db.add(Project(
                CATEGORY_ID=cat.ID,
                NAME=template_name,
                DESCRIPTION=entry.get("description"),
                ESTIMATED_TOTAL_DAYS=0.0,
                VENDOR_ID=vendor_id
            ))
            created_projects += 1
    db.commit()
    return {
        "message": f"Seed complete: {created_cats} categories, {created_projects} projects created"
    }
