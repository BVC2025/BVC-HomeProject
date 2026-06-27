"""
Product Master & Supplier-Product Pricing routes.

ProductMaster is the vendor-scoped product catalogue used by the
procurement module. SupplierProduct is the three-way junction
(Vendor × Supplier × Product) that stores quoted prices; every
price change is appended to SupplierProductPriceHistory.
"""

import io
import uuid
from datetime import datetime, date
from typing import Optional, List

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from sqlalchemy.orm import Session

from app.database.database import get_db
from app.models.models import Supplier, CustomField, CustomFieldTableValue
from app.models.inventory_models import InventoryCategory, ProductMaster
from app.models.supplier_models import (
    SupplierProduct,
    SupplierProductPriceHistory,
    PurchaseRecommendation,
    SupplierRanking,
)
from app.schemas.supplier_product_schema import (
    ProductCreate, ProductUpdate,
    CategoryCreate, CategoryUpdate,
    SupplierProductCreate, SupplierProductPriceUpdate,
)

router = APIRouter(tags=["Supplier Products"])


# ─────────────────────────────────────────────────────────────────────
# Shared bulk-upload helpers  (copied from project_template.py pattern)
# ─────────────────────────────────────────────────────────────────────

def _cf_fields_for_table(table_name: str, vendor_id: int, db: Session):
    return (
        db.query(CustomField)
        .filter(CustomField.TABLE_NAME == table_name, CustomField.VENDOR_ID == vendor_id)
        .order_by(CustomField.SORT_ORDER, CustomField.FIELD_NAME)
        .all()
    )


def _upsert_cf_bulk(row_id: str, table_name: str, cf_field_id: str, value, db: Session):
    stored = value if value else None
    existing = (
        db.query(CustomFieldTableValue)
        .filter(
            CustomFieldTableValue.TABLE_NAME == table_name,
            CustomFieldTableValue.TABLE_ROW_ID == str(row_id),
            CustomFieldTableValue.CUSTOM_FIELD_ID == cf_field_id,
        )
        .first()
    )
    if existing:
        existing.CUSTOM_FIELD_VALUE = stored
    elif stored is not None:
        db.add(CustomFieldTableValue(
            TABLE_NAME=table_name,
            TABLE_ROW_ID=str(row_id),
            CUSTOM_FIELD_ID=cf_field_id,
            CUSTOM_FIELD_VALUE=stored,
        ))


def _parse_bulk_xl(content: bytes, required_sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if required_sheet not in wb.sheetnames:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise HTTPException(
            status_code=400,
            detail=f'Sheet "{required_sheet}" not found. Available: {available}. '
                   f'Use the Template download to get a correctly named workbook.',
        )
    ws = wb[required_sheet]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper() if c is not None else "" for c in row]
            continue
        if all(c is None for c in row):
            continue
        rows.append(row)
    return headers, rows


def _cell(record: dict, *keys) -> str:
    for k in keys:
        v = record.get(k.upper())
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    return ""


# ─────────────────────────────────────────────────────────────────────
# InventoryCategory CRUD
# ─────────────────────────────────────────────────────────────────────

@router.get("/inventory-categories")
def list_categories(
    vendor_id: int = Query(1),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(InventoryCategory).filter(InventoryCategory.VENDOR_ID == vendor_id)
    if search:
        q = q.filter(InventoryCategory.NAME.ilike(f"%{search}%"))
    rows = q.order_by(InventoryCategory.SORT_ORDER, InventoryCategory.NAME).all()
    return [
        {
            "ID": r.ID, "VENDOR_ID": r.VENDOR_ID,
            "NAME": r.NAME, "CODE": r.CODE, "DESCRIPTION": r.DESCRIPTION,
            "SORT_ORDER": r.SORT_ORDER, "IS_ACTIVE": r.IS_ACTIVE,
            "PRODUCT_COUNT": len(r.products),
            "CREATED_AT": r.CREATED_AT.isoformat() if r.CREATED_AT else None,
        }
        for r in rows
    ]


@router.post("/inventory-categories")
def create_category(payload: CategoryCreate, db: Session = Depends(get_db)):
    existing = db.query(InventoryCategory).filter(
        InventoryCategory.VENDOR_ID == payload.VENDOR_ID,
        InventoryCategory.NAME == payload.NAME,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Category '{payload.NAME}' already exists")
    cat = InventoryCategory(**payload.dict())
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return {"message": "Category created", "ID": cat.ID}


@router.get("/inventory-categories/{cat_id}")
def get_category(cat_id: str, db: Session = Depends(get_db)):
    cat = db.query(InventoryCategory).filter(InventoryCategory.ID == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    return {
        "ID": cat.ID, "VENDOR_ID": cat.VENDOR_ID,
        "NAME": cat.NAME, "CODE": cat.CODE, "DESCRIPTION": cat.DESCRIPTION,
        "SORT_ORDER": cat.SORT_ORDER, "IS_ACTIVE": cat.IS_ACTIVE,
        "CREATED_AT": cat.CREATED_AT.isoformat() if cat.CREATED_AT else None,
        "UPDATED_AT": cat.UPDATED_AT.isoformat() if cat.UPDATED_AT else None,
    }


@router.put("/inventory-categories/{cat_id}")
def update_category(cat_id: str, payload: CategoryUpdate, db: Session = Depends(get_db)):
    cat = db.query(InventoryCategory).filter(InventoryCategory.ID == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    for k, v in payload.dict(exclude_none=True).items():
        setattr(cat, k, v)
    db.commit()
    return {"message": "Category updated"}


@router.delete("/inventory-categories/{cat_id}")
def delete_category(cat_id: str, db: Session = Depends(get_db)):
    cat = db.query(InventoryCategory).filter(InventoryCategory.ID == cat_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    if cat.products:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete category '{cat.NAME}' — it has {len(cat.products)} product(s)"
        )
    db.delete(cat)
    db.commit()
    return {"message": "Category deleted"}


# ─────────────────────────────────────────────────────────────────────
# InventoryCategory — bulk-upload, template, export
# ─────────────────────────────────────────────────────────────────────

@router.get("/inventory-categories/bulk-template")
def category_bulk_template(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    cf_fields = _cf_fields_for_table("inventory_category", vendor_id, db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InventoryCategories"
    standard_headers = ["Name", "Code", "Description", "Sort Order"]
    cf_headers = [f.FIELD_NAME for f in cf_fields]
    ws.append(standard_headers + cf_headers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_categories_template.xlsx"},
    )


@router.get("/inventory-categories/export/excel")
def category_export_excel(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    cats = (
        db.query(InventoryCategory)
        .filter(InventoryCategory.VENDOR_ID == vendor_id)
        .order_by(InventoryCategory.SORT_ORDER, InventoryCategory.NAME)
        .all()
    )
    cf_fields = _cf_fields_for_table("inventory_category", vendor_id, db)
    cat_ids = [c.ID for c in cats]
    cf_values_rows = (
        db.query(CustomFieldTableValue)
        .filter(
            CustomFieldTableValue.TABLE_NAME == "inventory_category",
            CustomFieldTableValue.TABLE_ROW_ID.in_(cat_ids),
        )
        .all()
    ) if cat_ids else []
    cfv_map: dict = {}
    for v in cf_values_rows:
        cfv_map.setdefault(v.TABLE_ROW_ID, {})[v.CUSTOM_FIELD_ID] = v.CUSTOM_FIELD_VALUE

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InventoryCategories"
    headers = ["S.No", "Name", "Code", "Description", "Sort Order", "Active", "Product Count", "Created At"]
    headers += [f.FIELD_NAME for f in cf_fields]
    ws.append(headers)

    for i, c in enumerate(cats, 1):
        row_data = [
            i, c.NAME, c.CODE or "", c.DESCRIPTION or "",
            c.SORT_ORDER or 0, "Yes" if c.IS_ACTIVE else "No",
            len(c.products) if c.products is not None else 0,
            c.CREATED_AT.strftime("%Y-%m-%d") if c.CREATED_AT else "",
        ]
        for f in cf_fields:
            val = cfv_map.get(str(c.ID), {}).get(f.ID, "")
            row_data.append(val if val is not None else "")
        ws.append(row_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_categories.xlsx"},
    )


@router.post("/inventory-categories/bulk-upload")
def category_bulk_upload(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = file.file.read()
    headers, rows = _parse_bulk_xl(content, "InventoryCategories")

    cf_fields = _cf_fields_for_table("inventory_category", vendor_id, db)
    cf_header_map = {f.FIELD_NAME.upper(): f for f in cf_fields}

    inserted = updated = skipped = 0
    errors: list = []

    for row_idx, raw in enumerate(rows, start=2):
        record = {headers[j].upper(): (raw[j] if j < len(raw) else None) for j in range(len(headers))}
        name = _cell(record, "NAME")
        if not name:
            errors.append({"row": row_idx, "field": "Name", "message": "Name is required"})
            skipped += 1
            continue

        code = _cell(record, "CODE").upper() or None
        try:
            sort_order = int(_cell(record, "SORT ORDER", "SORT_ORDER") or 0)
        except ValueError:
            sort_order = 0

        existing = (
            db.query(InventoryCategory)
            .filter(InventoryCategory.VENDOR_ID == vendor_id, InventoryCategory.NAME == name)
            .first()
        )

        try:
            if existing:
                existing.CODE = code or existing.CODE
                existing.DESCRIPTION = _cell(record, "DESCRIPTION") or existing.DESCRIPTION
                existing.SORT_ORDER = sort_order
                db.flush()
                row_id = existing.ID
                updated += 1
            else:
                new_cat = InventoryCategory(
                    ID=str(uuid.uuid4()),
                    VENDOR_ID=vendor_id,
                    NAME=name,
                    CODE=code,
                    DESCRIPTION=_cell(record, "DESCRIPTION") or None,
                    SORT_ORDER=sort_order,
                    IS_ACTIVE=True,
                )
                db.add(new_cat)
                db.flush()
                row_id = new_cat.ID
                inserted += 1

            for cf_name_upper, cf_field in cf_header_map.items():
                raw_val = record.get(cf_name_upper)
                if raw_val is not None:
                    _upsert_cf_bulk(row_id, "inventory_category", cf_field.ID, str(raw_val).strip(), db)

        except Exception as exc:
            db.rollback()
            errors.append({"row": row_idx, "message": str(exc)})
            skipped += 1
            continue

    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


# ─────────────────────────────────────────────────────────────────────
# ProductMaster CRUD
# ─────────────────────────────────────────────────────────────────────

def _serialize_product(p: ProductMaster, db: Session, include_suppliers: bool = False) -> dict:
    result = {
        "ID": p.ID, "VENDOR_ID": p.VENDOR_ID, "CATEGORY_ID": p.CATEGORY_ID,
        "DEPARTMENT_ID": p.DEPARTMENT_ID,
        "DEPARTMENT_NAME": p.department.NAME if p.department else None,
        "PRODUCT_CODE": p.PRODUCT_CODE, "PRODUCT_NAME": p.PRODUCT_NAME,
        "DESCRIPTION": p.DESCRIPTION, "HSN_CODE": p.HSN_CODE, "UNIT": p.UNIT,
        "IMAGE_URL": p.IMAGE_URL, "SPECIFICATIONS": p.SPECIFICATIONS,
        "STATUS": p.STATUS,
        "CATEGORY_NAME": p.category.NAME if p.category else None,
        "CREATED_AT": p.CREATED_AT.isoformat() if p.CREATED_AT else None,
        "UPDATED_AT": p.UPDATED_AT.isoformat() if p.UPDATED_AT else None,
    }
    if include_suppliers:
        result["suppliers"] = [
            {
                "SUPPLIER_PRODUCT_ID": sp.ID,
                "SUPPLIER_ID": sp.SUPPLIER_ID,
                "UNIT_PRICE": float(sp.UNIT_PRICE),
                "CURRENCY": sp.CURRENCY,
                "MOQ": sp.MOQ,
                "LEAD_TIME_DAYS": sp.LEAD_TIME_DAYS,
                "AVAILABLE_QTY": sp.AVAILABLE_QTY,
                "STATUS": sp.STATUS,
                "IS_PREFERRED": sp.IS_PREFERRED,
            }
            for sp in p.supplier_products if sp.STATUS == "ACTIVE"
        ]
        rec = p.recommendation
        result["recommendation"] = {
            "RECOMMENDED_SUPPLIER_ID": rec.RECOMMENDED_SUPPLIER_ID,
            "RECOMMENDED_PRICE": float(rec.RECOMMENDED_PRICE),
            "REASON": rec.RECOMMENDATION_REASON,
            "ALTERNATIVES": rec.ALTERNATIVE_SUPPLIER_IDS,
        } if rec else None
    return result


@router.get("/products")
def list_products(
    vendor_id: int = Query(1),
    category_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("PRODUCT_NAME"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    q = db.query(ProductMaster).filter(ProductMaster.VENDOR_ID == vendor_id)
    if category_id:
        q = q.filter(ProductMaster.CATEGORY_ID == category_id)
    if status:
        q = q.filter(ProductMaster.STATUS == status.upper())
    if search:
        term = f"%{search}%"
        q = q.filter(
            ProductMaster.PRODUCT_NAME.ilike(term) |
            ProductMaster.PRODUCT_CODE.ilike(term)
        )
    total = q.count()
    sort_col = getattr(ProductMaster, sort_by, ProductMaster.PRODUCT_NAME)
    rows = q.order_by(sort_col).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [_serialize_product(r, db) for r in rows],
    }


@router.post("/products")
def create_product(payload: ProductCreate, db: Session = Depends(get_db)):
    existing = db.query(ProductMaster).filter(
        ProductMaster.VENDOR_ID == payload.VENDOR_ID,
        ProductMaster.PRODUCT_CODE == payload.PRODUCT_CODE,
    ).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Product code '{payload.PRODUCT_CODE}' already exists"
        )
    product = ProductMaster(**payload.dict())
    db.add(product)
    db.commit()
    db.refresh(product)
    return {"message": "Product created", "ID": product.ID}


@router.get("/products/{product_id}")
def get_product(product_id: str, db: Session = Depends(get_db)):
    p = db.query(ProductMaster).filter(ProductMaster.ID == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    result = _serialize_product(p, db, include_suppliers=True)
    # Attach custom fields
    result["custom_fields"] = [
        {"ID": f.ID, "FIELD_NAME": f.FIELD_NAME, "FIELD_TYPE": f.FIELD_TYPE,
         "IS_REQUIRED": f.IS_REQUIRED, "SORT_ORDER": f.SORT_ORDER, "OPTIONS": f.OPTIONS}
        for f in _cf_fields_for_table("product_master", p.VENDOR_ID, db)
    ]
    result["custom_field_values"] = [
        {"CUSTOM_FIELD_ID": v.CUSTOM_FIELD_ID, "VALUE": v.CUSTOM_FIELD_VALUE}
        for v in db.query(CustomFieldTableValue).filter(
            CustomFieldTableValue.TABLE_NAME == "product_master",
            CustomFieldTableValue.TABLE_ROW_ID == str(product_id),
        ).all()
    ]
    return result


@router.put("/products/{product_id}")
def update_product(product_id: str, payload: ProductUpdate, db: Session = Depends(get_db)):
    p = db.query(ProductMaster).filter(ProductMaster.ID == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    if payload.PRODUCT_CODE and payload.PRODUCT_CODE != p.PRODUCT_CODE:
        clash = db.query(ProductMaster).filter(
            ProductMaster.VENDOR_ID == p.VENDOR_ID,
            ProductMaster.PRODUCT_CODE == payload.PRODUCT_CODE,
            ProductMaster.ID != product_id,
        ).first()
        if clash:
            raise HTTPException(status_code=400, detail="Product code already in use")
    for k, v in payload.dict(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return {"message": "Product updated"}


@router.delete("/products/{product_id}")
def delete_product(product_id: str, db: Session = Depends(get_db)):
    p = db.query(ProductMaster).filter(ProductMaster.ID == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    active_sp = db.query(SupplierProduct).filter(
        SupplierProduct.PRODUCT_ID == product_id,
        SupplierProduct.STATUS == "ACTIVE",
    ).count()
    if active_sp:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete product — {active_sp} active supplier(s) supply it. "
                   "Deactivate all supplier links first."
        )
    p.STATUS = "INACTIVE"
    db.commit()
    return {"message": "Product deactivated"}


@router.get("/products/{product_id}/suppliers")
def get_product_suppliers(product_id: str, db: Session = Depends(get_db)):
    """All suppliers (with prices and rankings) for a product."""
    p = db.query(ProductMaster).filter(ProductMaster.ID == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")

    rows = (
        db.query(SupplierProduct)
        .filter(SupplierProduct.PRODUCT_ID == product_id)
        .order_by(SupplierProduct.UNIT_PRICE.asc())
        .all()
    )
    rankings = {
        r.SUPPLIER_ID: r
        for r in db.query(SupplierRanking).filter(
            SupplierRanking.PRODUCT_ID == product_id
        ).all()
    }

    return [
        {
            "SUPPLIER_PRODUCT_ID": sp.ID,
            "SUPPLIER_ID": sp.SUPPLIER_ID,
            "UNIT_PRICE": float(sp.UNIT_PRICE),
            "CURRENCY": sp.CURRENCY,
            "MOQ": sp.MOQ,
            "LEAD_TIME_DAYS": sp.LEAD_TIME_DAYS,
            "AVAILABLE_QTY": sp.AVAILABLE_QTY,
            "STATUS": sp.STATUS,
            "IS_PREFERRED": sp.IS_PREFERRED,
            "LAST_PRICE_UPDATED_AT": sp.LAST_PRICE_UPDATED_AT.isoformat() if sp.LAST_PRICE_UPDATED_AT else None,
            "RANK": rankings[sp.SUPPLIER_ID].RANK if sp.SUPPLIER_ID in rankings else None,
            "COMPOSITE_SCORE": rankings[sp.SUPPLIER_ID].COMPOSITE_SCORE if sp.SUPPLIER_ID in rankings else None,
        }
        for sp in rows
    ]


@router.get("/products/{product_id}/recommendation")
def get_product_recommendation(product_id: str, db: Session = Depends(get_db)):
    rec = db.query(PurchaseRecommendation).filter(
        PurchaseRecommendation.PRODUCT_ID == product_id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="No recommendation available for this product")
    return {
        "PRODUCT_ID": rec.PRODUCT_ID,
        "RECOMMENDED_SUPPLIER_ID": rec.RECOMMENDED_SUPPLIER_ID,
        "RECOMMENDED_PRICE": float(rec.RECOMMENDED_PRICE),
        "REASON": rec.RECOMMENDATION_REASON,
        "ALTERNATIVES": rec.ALTERNATIVE_SUPPLIER_IDS,
        "LAST_RECALCULATED_AT": rec.LAST_RECALCULATED_AT.isoformat() if rec.LAST_RECALCULATED_AT else None,
    }


# ─────────────────────────────────────────────────────────────────────
# Supplier-Product Pricing
# ─────────────────────────────────────────────────────────────────────

@router.post("/products/{product_id}/suppliers/{supplier_id}/price")
def set_supplier_price(
    product_id: str,
    supplier_id: int,
    payload: SupplierProductPriceUpdate,
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    """Set or update a supplier's price for a product. Appends price history on change."""
    sp = db.query(SupplierProduct).filter(
        SupplierProduct.VENDOR_ID == vendor_id,
        SupplierProduct.SUPPLIER_ID == supplier_id,
        SupplierProduct.PRODUCT_ID == product_id,
    ).first()

    if sp:
        # Price changed — append history row
        if float(sp.UNIT_PRICE) != payload.UNIT_PRICE:
            history = SupplierProductPriceHistory(
                VENDOR_ID=vendor_id,
                SUPPLIER_PRODUCT_ID=sp.ID,
                OLD_PRICE=sp.UNIT_PRICE,
                NEW_PRICE=payload.UNIT_PRICE,
                CHANGED_BY_ID=payload.CHANGED_BY_ID,
                CHANGED_BY_ROLE=payload.CHANGED_BY_ROLE or "EMPLOYEE",
                CHANGE_REASON=payload.CHANGE_REASON,
                EFFECTIVE_DATE=date.today(),
            )
            db.add(history)
            sp.UNIT_PRICE = payload.UNIT_PRICE
            sp.LAST_PRICE_UPDATED_AT = datetime.utcnow()
        if payload.AVAILABLE_QTY is not None:
            sp.AVAILABLE_QTY = payload.AVAILABLE_QTY
        if payload.LEAD_TIME_DAYS is not None:
            sp.LEAD_TIME_DAYS = payload.LEAD_TIME_DAYS
        if payload.MOQ is not None:
            sp.MOQ = payload.MOQ
    else:
        # New supplier-product link
        supplier = db.query(Supplier).filter(Supplier.ID == supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
        product = db.query(ProductMaster).filter(ProductMaster.ID == product_id).first()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        sp = SupplierProduct(
            VENDOR_ID=vendor_id,
            SUPPLIER_ID=supplier_id,
            PRODUCT_ID=product_id,
            UNIT_PRICE=payload.UNIT_PRICE,
            MOQ=payload.MOQ or 1.0,
            LEAD_TIME_DAYS=payload.LEAD_TIME_DAYS or 7,
            AVAILABLE_QTY=payload.AVAILABLE_QTY,
            STATUS="ACTIVE",
            LAST_PRICE_UPDATED_AT=datetime.utcnow(),
        )
        db.add(sp)

    db.commit()

    # Trigger ranking recalculation (best-effort)
    try:
        from app.services.supplier_ranking_service import recalculate_ranking_for_product
        recalculate_ranking_for_product(db, vendor_id, product_id)
    except Exception as exc:
        print(f"[supplier-products] ranking recalc failed (non-fatal): {exc}")

    return {"message": "Supplier price updated", "SUPPLIER_PRODUCT_ID": sp.ID}


@router.delete("/products/{product_id}/suppliers/{supplier_id}")
def deactivate_supplier_product(
    product_id: str,
    supplier_id: int,
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    sp = db.query(SupplierProduct).filter(
        SupplierProduct.VENDOR_ID == vendor_id,
        SupplierProduct.SUPPLIER_ID == supplier_id,
        SupplierProduct.PRODUCT_ID == product_id,
    ).first()
    if not sp:
        raise HTTPException(status_code=404, detail="Supplier-product link not found")
    sp.STATUS = "INACTIVE"
    db.commit()

    try:
        from app.services.supplier_ranking_service import recalculate_ranking_for_product
        recalculate_ranking_for_product(db, vendor_id, product_id)
    except Exception as exc:
        print(f"[supplier-products] ranking recalc failed (non-fatal): {exc}")

    return {"message": "Supplier-product link deactivated"}


@router.get("/products/{product_id}/suppliers/{supplier_id}/history")
def get_price_history(
    product_id: str,
    supplier_id: int,
    vendor_id: int = Query(1),
    db: Session = Depends(get_db),
):
    sp = db.query(SupplierProduct).filter(
        SupplierProduct.VENDOR_ID == vendor_id,
        SupplierProduct.SUPPLIER_ID == supplier_id,
        SupplierProduct.PRODUCT_ID == product_id,
    ).first()
    if not sp:
        raise HTTPException(status_code=404, detail="Supplier-product link not found")
    rows = (
        db.query(SupplierProductPriceHistory)
        .filter(SupplierProductPriceHistory.SUPPLIER_PRODUCT_ID == sp.ID)
        .order_by(SupplierProductPriceHistory.CREATED_AT.desc())
        .all()
    )
    return [
        {
            "ID": r.ID,
            "OLD_PRICE": float(r.OLD_PRICE),
            "NEW_PRICE": float(r.NEW_PRICE),
            "CHANGED_BY_ID": r.CHANGED_BY_ID,
            "CHANGED_BY_ROLE": r.CHANGED_BY_ROLE,
            "CHANGE_REASON": r.CHANGE_REASON,
            "EFFECTIVE_DATE": r.EFFECTIVE_DATE.isoformat() if r.EFFECTIVE_DATE else None,
            "CREATED_AT": r.CREATED_AT.isoformat() if r.CREATED_AT else None,
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────
# Bulk Upload / Template / Export  (follows project_template.py pattern)
# ─────────────────────────────────────────────────────────────────────

_PROD_STD_COLS = {
    "PRODUCT CODE", "PRODUCT NAME", "CATEGORY NAME", "DEPARTMENT NAME",
    "DESCRIPTION", "HSN CODE", "UNIT", "STATUS", "S.NO", "S.N", "SN", ""
}


@router.get("/products/bulk-template")
def download_product_template(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    std_cols = ["PRODUCT CODE", "PRODUCT NAME", "CATEGORY NAME", "DEPARTMENT NAME",
                "DESCRIPTION", "HSN CODE", "UNIT", "STATUS"]
    cf_fields = _cf_fields_for_table("product_master", vendor_id, db)
    cf_cols = [f.FIELD_NAME for f in cf_fields]
    ws.append(std_cols + cf_cols)
    # Style header row
    from openpyxl.styles import Font, PatternFill
    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_bulk_template.xlsx"},
    )


@router.post("/products/bulk-upload")
async def bulk_upload_products(
    vendor_id: int = Query(1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    headers, data_rows = _parse_bulk_xl(content, "Products")

    cf_fields = _cf_fields_for_table("product_master", vendor_id, db)
    cf_by_upper = {f.FIELD_NAME.upper(): f for f in cf_fields}
    cf_cols = [h for h in headers if h not in _PROD_STD_COLS and h in cf_by_upper]

    from app.models.models import Department

    # Build category and department lookups
    cats = {c.NAME.upper(): c for c in db.query(InventoryCategory).filter(
        InventoryCategory.VENDOR_ID == vendor_id
    ).all()}
    depts = {d.NAME.upper(): d for d in db.query(Department).all()}

    inserted = updated = skipped = 0
    errors: List[dict] = []

    for row_num, raw in enumerate(data_rows, start=2):
        record = {headers[i].upper(): raw[i] for i in range(len(headers))}
        prod_code = _cell(record, "PRODUCT CODE")
        prod_name = _cell(record, "PRODUCT NAME")

        if not prod_code:
            errors.append({"row": row_num, "field": "PRODUCT CODE", "message": "Required"})
            continue
        if not prod_name:
            errors.append({"row": row_num, "field": "PRODUCT NAME", "message": "Required"})
            continue

        cat_name = _cell(record, "CATEGORY NAME")
        category_id = cats[cat_name.upper()].ID if cat_name and cat_name.upper() in cats else None
        dept_name = _cell(record, "DEPARTMENT NAME")
        department_id = depts[dept_name.upper()].ID if dept_name and dept_name.upper() in depts else None

        cf_vals: dict = {}
        cf_error = False
        for col in cf_cols:
            cf_f = cf_by_upper[col]
            raw_val = record.get(col)
            if cf_f.IS_REQUIRED and (raw_val is None or str(raw_val).strip() == ""):
                errors.append({"row": row_num, "field": cf_f.FIELD_NAME, "message": "Required"})
                cf_error = True
            cf_vals[cf_f.ID] = raw_val
        if cf_error:
            continue

        existing = db.query(ProductMaster).filter(
            ProductMaster.VENDOR_ID == vendor_id,
            ProductMaster.PRODUCT_CODE == prod_code,
        ).first()

        if existing:
            existing.PRODUCT_NAME = prod_name
            existing.CATEGORY_ID = category_id or existing.CATEGORY_ID
            existing.DEPARTMENT_ID = department_id or existing.DEPARTMENT_ID
            existing.DESCRIPTION = _cell(record, "DESCRIPTION") or existing.DESCRIPTION
            existing.HSN_CODE = _cell(record, "HSN CODE") or existing.HSN_CODE
            existing.UNIT = _cell(record, "UNIT") or existing.UNIT
            status_val = _cell(record, "STATUS")
            if status_val and status_val.upper() in ("ACTIVE", "INACTIVE", "DISCONTINUED"):
                existing.STATUS = status_val.upper()
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(existing.ID, "product_master", cf_id, val, db)
            updated += 1
        else:
            p = ProductMaster(
                VENDOR_ID=vendor_id,
                PRODUCT_CODE=prod_code,
                PRODUCT_NAME=prod_name,
                CATEGORY_ID=category_id,
                DEPARTMENT_ID=department_id,
                DESCRIPTION=_cell(record, "DESCRIPTION") or None,
                HSN_CODE=_cell(record, "HSN CODE") or None,
                UNIT=_cell(record, "UNIT") or "PCS",
                STATUS=(_cell(record, "STATUS") or "ACTIVE").upper(),
            )
            db.add(p)
            db.flush()
            for cf_id, val in cf_vals.items():
                _upsert_cf_bulk(p.ID, "product_master", cf_id, val, db)
            inserted += 1

    db.commit()
    return {
        "message": f"Upload: {inserted} inserted, {updated} updated, {skipped} skipped",
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total_rows": len(data_rows), "errors": errors,
    }


@router.get("/products/export/excel")
def export_products(vendor_id: int = Query(1), db: Session = Depends(get_db)):
    rows = db.query(ProductMaster).filter(
        ProductMaster.VENDOR_ID == vendor_id
    ).order_by(ProductMaster.PRODUCT_NAME).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["PRODUCT CODE", "PRODUCT NAME", "CATEGORY", "DEPARTMENT",
               "HSN CODE", "UNIT", "STATUS", "CREATED AT"])
    for r in rows:
        ws.append([
            r.PRODUCT_CODE, r.PRODUCT_NAME,
            r.category.NAME if r.category else "",
            r.department.NAME if r.department else "",
            r.HSN_CODE or "", r.UNIT or "",
            r.STATUS or "",
            r.CREATED_AT.isoformat() if r.CREATED_AT else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"},
    )
